# -*- coding: utf-8 -*-
import requests
import json
import time
import random
from datetime import datetime
import concurrent.futures
import threading
import pandas as pd
from bs4 import BeautifulSoup
import os

# 全局日志控制
VERBOSE = True

def log_debug(msg: str):
    if VERBOSE:
        print(msg)

def log_info(msg: str):
    print(msg)

def extract_pure_fund_code(fund_code_with_tag):
    """提取纯基金代码，去掉自定义板块标签"""
    if '[' in fund_code_with_tag:
        return fund_code_with_tag.split('[')[0]
    return fund_code_with_tag

def sort_funds_by_category(fund_list):
    """按板块分类排序基金列表"""
    if not fund_list:
        return fund_list
    
    # 定义板块优先级顺序
    category_priority = {
        # 港股板块
        "港股": 1,"港股科技": 2, "科技": 3,"港股金融": 4,"金融": 5,"港股医药": 6, "医药": 7, "医疗器械": 8,
        # 核心科技板块
         "半导体": 9, "计算机": 10, "电子": 11, "通信": 12, "云计算": 13,"人工智能": 14, "机器人": 15,"消费电子": 16,
        # 金融地产
         "地产": 17, "建筑装饰": 18, "建筑材料": 19,
        # 新兴产业
        "新能源": 20, "光伏": 21, "风电": 22, "储能": 23, "新能源汽车": 24, "汽车": 25,
        # 传统优势板块
        "军工": 26,   "消费": 27, "食品饮料": 28, "家电": 29, 
        # 金属细分板块
        "贵金属": 30, "有色金属": 31, "稀土": 32, 
        # 周期板块
        "化工": 33, "钢铁": 34, "煤炭": 35, "传统能源": 36, "电力": 37, "机械设备": 38, "电气设备": 39,
        # 其他板块
        "农业": 40, "基建": 41, "传媒": 42, "环保": 43, "教育": 44, "物流": 45,
        "纺织服装": 46, "轻工制造": 47, "公用事业": 48, "交通运输": 49, "商业贸易": 50, "休闲服务": 51, "综合": 52, "量化": 53,
        # 基金类型分类
        "ETF基金": 54, "LOF基金": 55, "ETF联接": 56, "混合型": 57, "股票型": 58, "债券型": 59, "货币型": 60, "指数型": 61,
        # 其他分类
        "场内基金": 62, "其他": 63, "未知": 64
    }
    
    def get_category_priority(fund):
        category = fund.get('板块分类', '未知')
        return category_priority.get(category, 999)
    
    # 按板块优先级排序，相同板块内按基金代码排序
    return sorted(fund_list, key=lambda x: (get_category_priority(x), x.get('基金代码', '')))

def sort_by_original_order(fund_list, original_order):
    """按照原始定义的顺序排序基金"""
    if not fund_list or not original_order:
        return fund_list
    
    # 创建基金代码到基金的映射，需要处理自定义标签
    fund_dict = {}
    for fund in fund_list:
        fund_dict[fund['基金代码']] = fund
    
    # 按照原始顺序重新排列
    sorted_list = []
    for code in original_order:
        pure_code = extract_pure_fund_code(code)
        if pure_code in fund_dict:
            sorted_list.append(fund_dict[pure_code])
    
    return sorted_list

class FundCategoryClassifier:
    def __init__(self):
        # 定义板块关键词映射 - 更详细和专业
        self.category_keywords = {
            # 国防军工板块
            "军工": ["军工", "国防", "空天", "航天", "航空", "军事", "军民", "军品", "军需", "军转民", "军民融合", "国防科技", "军工科技"],
            
            # 医药生物板块
            "医药": ["医药", "医疗", "生物", "创新药", "健康", "疫苗", "生物医药", "医疗器械", "生物技术", "基因", "细胞", "抗体", "疫苗", "中药", "西药", "生物制品", "医药服务", "CRO", "CDMO"],
            
            # 科技板块
            "科技": ["科技", "TMT", "通信", "电子", "半导体", "芯片", "大数据", "云计算", "物联网", "5G", "6G", "信息技术", "软件", "互联网", "数字化", "智能化", "数字经济"],
            
            # 人工智能板块
            "人工智能": ["人工智能", "AI", "机器学习", "深度学习", "神经网络", "智能算法", "智能系统", "智能技术"],
            
            # 新能源板块
            "新能源": ["新能源", "光伏", "太阳能", "风电", "储能", "电池", "电动车", "氢能", "核能", "生物质能", "地热能", "潮汐能", "清洁能源", "绿色能源", "碳中和", "碳达峰"],
            
            # 新能源汽车板块
            "新能源汽车": ["新能源汽车", "电动车", "电动汽车", "新能源车", "智能汽车", "电动化"],
            
            # 消费板块
            "消费": ["消费", "白酒", "食品", "饮料", "家电", "零售", "消费升级", "新消费", "可选消费", "必选消费", "奢侈品", "化妆品", "服装", "珠宝", "旅游", "酒店", "餐饮", "商超"],
            
            # 金融板块
            "金融": ["银行", "金融", "保险", "证券", "红利", "银行股", "保险股", "券商", "信托", "期货", "基金", "理财", "资管", "投行", "商业银行", "投资银行", "金融科技", "FinTech"],
            
            # 房地产板块
            "地产": ["地产", "房地产", "建筑", "建材", "地产股", "开发商", "物业管理", "商业地产", "住宅地产", "工业地产", "土地开发", "房地产服务", "建筑装饰"],
            
            # 建筑材料板块
            "建筑材料": ["建筑材料", "建材", "水泥", "玻璃", "钢材", "建材股", "建材制造", "建材服务"],
            
            # 农业板块
            "农业": ["农业", "养殖", "种植", "农产品", "农业股", "种植业", "畜牧业", "渔业", "林业", "种子", "化肥", "农药", "农机", "农业服务", "乡村振兴", "智慧农业"],
            
            # 贵金属板块
            "贵金属": ["贵金属", "黄金", "黄金股", "白银", "铂金", "钯金"],
            "有色金属": ["有色金属", "铜", "铝", "锌", "镍", "钴", "锂", "金属矿业", "矿业股"],
            "稀土": ["稀土", "稀土金属", "稀土元素", "稀土材料"],
            
            # 港股板块
            "港股": ["港股", "恒生", "香港", "港股通", "H股", "红筹股", "蓝筹股", "中概股", "港股科技", "港股消费", "港股金融", "港股地产"],
            "港股科技": ["港股科技", "港股通科技", "恒生科技", "港股互联网", "港股TMT"],
            "港股金融": ["港股金融", "港股通金融", "恒生金融", "港股银行", "港股保险"],
            "港股医药": ["港股医药", "港股通医药", "港股医疗", "港股生物医药"],
            
            # 智能制造板块
            "机器人": ["机器人", "智能制造", "工业4.0", "自动化", "工业机器人", "服务机器人", "特种机器人", "智能制造装备", "工业自动化", "数字化制造", "柔性制造", "精益制造"],
            
            # 基建板块
            "基建": ["基建", "基础设施", "工程", "建筑材料", "基建股", "铁路", "公路", "机场", "港口", "水利", "电力", "通信", "城市基础设施", "PPP", "专项债"],
            
            # 传媒板块
            "传媒": ["传媒", "文化", "娱乐", "游戏", "影视", "出版", "广告", "营销", "新媒体", "短视频", "直播", "电竞", "动漫", "IP", "版权", "数字内容"],
            
            # 环保板块
            "环保": ["环保", "节能", "绿色", "环保股", "污水处理", "大气治理", "固废处理", "土壤修复", "环境监测", "环保设备", "环保服务", "碳交易", "ESG", "可持续发展"],
            
            # 教育板块
            "教育": ["教育", "培训", "在线教育", "K12", "职业教育", "高等教育", "学前教育", "教育信息化", "智慧教育", "教育科技", "教育服务", "教育装备"],
            
            # 物流板块
            "物流": ["物流", "快递", "运输", "供应链", "仓储", "配送", "货运", "物流股", "快递股", "运输股", "供应链管理", "智慧物流", "冷链物流", "国际物流"],
            
            # 基金类型分类
            "混合型": ["混合", "混合型", "灵活配置", "平衡配置", "稳健配置", "积极配置"],
            "股票型": ["股票", "股票型", "股票基金", "股票投资", "权益投资"],
            "债券型": ["债券", "债基", "债券基金", "纯债", "一级债基", "二级债基", "可转债", "信用债", "利率债"],
            "货币型": ["货币", "货币基金", "货币市场", "现金管理", "活期理财"],
            "指数型": ["指数", "ETF", "LOF", "指数基金", "被动投资", "量化", "增强指数", "Smart Beta"],
            
            # 新增专业板块
            "汽车": ["汽车", "汽车股", "整车", "零部件", "汽车电子", "智能汽车", "新能源汽车", "传统汽车", "汽车服务", "汽车金融"],
            "化工": ["化工", "化工股", "基础化工", "精细化工", "新材料", "化学制品", "化学原料", "化工新材料", "特种化工"],
            "钢铁": ["钢铁", "钢铁股", "钢铁行业", "钢铁制品", "钢铁贸易", "钢铁物流", "钢铁服务"],
            "煤炭": ["煤炭", "煤炭股", "煤炭开采", "煤炭贸易", "煤炭物流", "煤炭服务", "煤炭化工"],
            "电力": ["电力", "电力股", "发电", "输电", "配电", "电力设备", "电力服务", "清洁电力", "传统电力"],
            "通信": ["通信", "通信股", "通信设备", "通信服务", "5G", "6G", "光通信", "无线通信", "卫星通信"],
            "计算机": ["计算机", "计算机股", "计算机设备", "计算机软件", "计算机服务", "系统集成", "IT服务", "软件开发"],
            "电子": ["电子", "电子股", "电子设备", "电子元件", "电子材料", "电子制造", "电子服务"],
            "半导体": ["半导体", "半导体股", "芯片", "集成电路", "IC", "晶圆", "封装", "测试", "设计", "制造"],
            "光伏": ["光伏", "光伏股", "太阳能", "光伏设备", "光伏材料", "光伏组件", "光伏电站", "光伏服务"],
            "风电": ["风电", "风电股", "风力发电", "风电设备", "风电材料", "风电场", "风电服务"],
            "储能": ["储能", "储能股", "电池", "储能设备", "储能材料", "储能系统", "储能服务", "电化学储能", "物理储能"],
            "消费电子": ["消费电子", "消费电子股", "智能手机", "平板电脑", "笔记本电脑", "智能穿戴", "智能家居", "消费电子设备"],
            "家电": ["家电", "家电股", "白色家电", "黑色家电", "小家电", "智能家电", "家电制造", "家电服务"],
            "食品饮料": ["食品饮料", "食品饮料股", "食品", "饮料", "酒类", "乳制品", "调味品", "休闲食品", "食品加工", "饮料制造"],
            "纺织服装": ["纺织服装", "纺织服装股", "纺织", "服装", "鞋帽", "家纺", "纺织材料", "服装制造", "纺织服务"],
            "轻工制造": ["轻工制造", "轻工制造股", "造纸", "包装", "印刷", "家具", "轻工材料", "轻工设备", "轻工服务"],
            "机械设备": ["机械设备", "机械设备股", "工程机械", "机床", "仪器仪表", "机械制造", "机械服务", "自动化设备"],
            "电气设备": ["电气设备", "电气设备股", "电力设备", "电气控制", "电气自动化", "电气制造", "电气服务"],
            "建筑装饰": ["建筑装饰", "建筑装饰股", "建筑", "装饰", "装修", "建筑设计", "建筑服务", "装饰材料"],
            "公用事业": ["公用事业", "公用事业股", "供水", "供气", "供热", "公共交通", "公共设施", "公共服务"],
            "交通运输": ["交通运输", "交通运输股", "公路", "铁路", "航空", "水运", "城市交通", "交通服务", "物流运输"],
            "商业贸易": ["商业贸易", "商业贸易股", "零售", "批发", "贸易", "商业服务", "商业地产", "商业运营"],
            "休闲服务": ["休闲服务", "休闲服务股", "旅游", "酒店", "餐饮", "娱乐", "休闲服务", "文化服务"],
            "综合": ["综合", "综合股", "多元化", "综合投资", "综合服务", "综合业务"],
            
            # 场内基金板块 - 更专业的分类
            "场内基金": ["场内基金", "ETF", "LOF", "场内交易", "交易所交易", "场内ETF", "场内LOF"],
            "ETF基金": ["ETF", "交易所交易基金", "指数ETF", "行业ETF", "主题ETF", "宽基ETF", "窄基ETF"],
            "ETF联接": ["ETF联接", "联接基金", "FOF", "基金中基金", "基金组合"],
            "LOF基金": ["LOF", "上市型开放式基金", "场内LOF", "场外LOF"]
        }
    
    def classify_fund(self, fund_name, fund_code=""):
        """根据基金名称分类，优先使用自定义板块标签"""
        if not fund_name:
            return "未知"
        
        # 优先检查基金代码中是否有自定义板块标签 [板块名称]
        if fund_code and '[' in fund_code and ']' in fund_code:
            try:
                start_idx = fund_code.find('[')
                end_idx = fund_code.find(']')
                if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                    custom_category = fund_code[start_idx + 1:end_idx]
                    if custom_category:
                        return custom_category
            except:
                pass
        
        # 对于境外基金和场内基金，使用简化的关键词自动分类
        if fund_code and ('.' in fund_code or fund_code in ['015016', '007280', '012060', '012920', '000834', '270042']):
            fund_name = fund_name.upper()  # 转换为大写便于匹配
        
            # 简化的关键词匹配，仅用于境外基金和场内基金
            if any(keyword in fund_name for keyword in ['科技', 'TMT', '互联网', '通信', '电子']):
                return "科技"
            elif any(keyword in fund_name for keyword in ['金融', '银行', '保险', '证券']):
                return "金融"
            elif any(keyword in fund_name for keyword in ['医药', '医疗', '生物', '健康']):
                return "医药"
            elif any(keyword in fund_name for keyword in ['消费', '食品', '饮料', '白酒']):
                return "消费"
            elif any(keyword in fund_name for keyword in ['军工', '国防', '航空', '航天']):
                return "军工"
            elif any(keyword in fund_name for keyword in ['新能源', '光伏', '风电', '储能']):
                return "新能源"
            elif any(keyword in fund_name for keyword in ['ETF', '指数']):
                return "ETF基金"
            else:
                return "其他"
        
        # 对于自选基金和监控基金，如果没有自定义标签，返回"未知"
        return "未知"
    
    def get_category_description(self, category):
        """获取板块描述"""
        descriptions = {
            # 核心科技板块
            "科技": "科技板块",
            "半导体": "半导体板块",
            # 港股细分板块
            "港股科技": "港股科技板块",
            "港股金融": "港股金融板块", 
            "港股医药": "港股医药板块",
            "港股": "港股板块",
            # 金属细分板块
            "贵金属": "贵金属板块",
            "有色金属": "有色金属板块",
            "稀土": "稀土板块",
            "计算机": "计算机板块",
            "电子": "电子板块",
            "通信": "通信板块",
            "人工智能": "人工智能板块",
            "机器人": "智能制造板块",
            
            # 新兴产业
            "新能源": "新能源板块",
            "光伏": "光伏板块",
            "风电": "风电板块",
            "储能": "储能板块",
            "新能源汽车": "新能源汽车板块",
            "消费电子": "消费电子板块",
            
            # 传统优势板块
            "军工": "国防军工板块",
            "医药": "医药生物板块",
            "消费": "消费板块",
            "食品饮料": "食品饮料板块",
            "家电": "家电板块",
            "汽车": "汽车板块",
            
            # 金融地产
            "金融": "金融板块",
            "地产": "房地产板块",
            "建筑装饰": "建筑装饰板块",
            "建筑材料": "建筑材料板块",
            
            # 周期板块
            "化工": "化工板块",
            "钢铁": "钢铁板块",
            "煤炭": "煤炭板块",
            "电力": "电力板块",
            "机械设备": "机械设备板块",
            "电气设备": "电气设备板块",
            
            # 其他板块
            "农业": "农业板块",
            "黄金": "贵金属板块",
            "港股": "港股板块",
            "基建": "基建板块",
            "传媒": "传媒板块",
            "环保": "环保板块",
            "教育": "教育板块",
            "物流": "物流板块",
            "纺织服装": "纺织服装板块",
            "轻工制造": "轻工制造板块",
            "公用事业": "公用事业板块",
            "交通运输": "交通运输板块",
            "商业贸易": "商业贸易板块",
            "休闲服务": "休闲服务板块",
            "综合": "综合板块",
            
            # 基金类型分类
            "混合型": "混合型基金",
            "股票型": "股票型基金",
            "债券型": "债券型基金",
            "货币型": "货币型基金",
            "指数型": "指数型基金",
            "ETF基金": "交易所交易基金",
            "LOF基金": "上市型开放式基金",
            "ETF联接": "ETF联接基金(FOF)",
            "场内基金": "场内交易基金",
            "其他": "其他类型",
            "未知": "未知类型"
        }
        return descriptions.get(category, "未知类型")

# 定义用户列表常量
SUPPORTED_USERS = ['chaochao', 'yaoyao', 'QDII']

class HoldingsProfitCalculator:
    """持仓收益计算器"""
    def __init__(self):
        """初始化持仓收益计算器"""
        pass
    
    def create_sample_holdings(self):
        """创建示例持仓数据"""
        holdings_data = {
            'chaochao': [
                {'fund_code': '023482', 'fund_name': '万家创新药', 'shares': 800, 'cost_price': 1.1500, 'cost_amount': 920.00},
                {'fund_code': '016573', 'fund_name': '招商银行AH', 'shares': 1200, 'cost_price': 1.0800, 'cost_amount': 1296.00},
                {'fund_code': '004409', 'fund_name': '招商TMT', 'shares': 900, 'cost_price': 1.2500, 'cost_amount': 1125.00},
                {'fund_code': '015740', 'fund_name': '国泰港股通科技', 'shares': 1100, 'cost_price': 1.1800, 'cost_amount': 1298.00},
                {'fund_code': '010364', 'fund_name': '鹏华军工', 'shares': 1000, 'cost_price': 1.2000, 'cost_amount': 1200.00},
                {'fund_code': '013309', 'fund_name': '易方达恒生科技', 'shares': 800, 'cost_price': 1.1000, 'cost_amount': 880.00},
                {'fund_code': '014806', 'fund_name': '国金量化混合', 'shares': 1000, 'cost_price': 1.0500, 'cost_amount': 1050.00},
                {'fund_code': '019571', 'fund_name': '诺安配置混合', 'shares': 900, 'cost_price': 1.0800, 'cost_amount': 972.00},
                {'fund_code': '018388', 'fund_name': '华泰柏瑞港股通红利', 'shares': 1000, 'cost_price': 1.1200, 'cost_amount': 1120.00},
                {'fund_code': '006113', 'fund_name': '汇添富创新药混合A', 'shares': 800, 'cost_price': 1.1500, 'cost_amount': 920.00},
                {'fund_code': '021717', 'fund_name': '招商云计算ETF', 'shares': 1000, 'cost_price': 1.1000, 'cost_amount': 1100.00},
                {'fund_code': '001665', 'fund_name': '平安鑫安混合', 'shares': 900, 'cost_price': 1.0500, 'cost_amount': 945.00},
                {'fund_code': '022435', 'fund_name': '南方中证500', 'shares': 1000, 'cost_price': 1.0800, 'cost_amount': 1080.00},
                {'fund_code': '019919', 'fund_name': '招商中证2000', 'shares': 800, 'cost_price': 1.1000, 'cost_amount': 880.00},
                {'fund_code': '014422', 'fund_name': '弘毅消费混合', 'shares': 1000, 'cost_price': 1.1200, 'cost_amount': 1120.00},
                {'fund_code': '020902', 'fund_name': '招商量化选股', 'shares': 900, 'cost_price': 1.0800, 'cost_amount': 972.00},
                {'fund_code': '021378', 'fund_name': '兴业港股通互联网', 'shares': 800, 'cost_price': 1.1500, 'cost_amount': 920.00},
                {'fund_code': '015401', 'fund_name': '弘毅甄选混合', 'shares': 1000, 'cost_price': 1.1000, 'cost_amount': 1100.00},
                {'fund_code': '016814', 'fund_name': '国联中证煤炭指数C', 'shares': 1000, 'cost_price': 1.7700, 'cost_amount': 1770.00},
                {'fund_code': '012341', 'fund_name': '东财食品饮料指数增强C', 'shares': 800, 'cost_price': 0.6483, 'cost_amount': 518.64},
                {'fund_code': '020671', 'fund_name': '易方达上证科创板芯片指数发起式C', 'shares': 900, 'cost_price': 2.0987, 'cost_amount': 1888.83},
                {'fund_code': '270042', 'fund_name': '广发纳斯达克100ETF联接人民币(QDII)A', 'shares': 100, 'cost_price': 6.8871, 'cost_amount': 688.71},
            ],
            'yaoyao': [
                {'fund_code': '021172', 'fund_name': '华安北证50A', 'shares': 1500, 'cost_price': 1.0500, 'cost_amount': 1575.00},
                {'fund_code': '015945', 'fund_name': '易方达军工混合', 'shares': 1000, 'cost_price': 1.1200, 'cost_amount': 1120.00},
                {'fund_code': '018647', 'fund_name': '易方达家电龙头', 'shares': 800, 'cost_price': 1.0800, 'cost_amount': 864.00},
                {'fund_code': '015897', 'fund_name': '天弘中证化工', 'shares': 1200, 'cost_price': 1.1500, 'cost_amount': 1380.00},
                {'fund_code': '012349', 'fund_name': '天弘恒生科技', 'shares': 900, 'cost_price': 1.2000, 'cost_amount': 1080.00},
                {'fund_code': '013416', 'fund_name': '永赢医疗器械', 'shares': 1000, 'cost_price': 1.1000, 'cost_amount': 1100.00},
                {'fund_code': '002833', 'fund_name': '华夏锦绣混合', 'shares': 800, 'cost_price': 1.0800, 'cost_amount': 864.00},
                {'fund_code': '003547', 'fund_name': '鹏华丰禄债券', 'shares': 1000, 'cost_price': 1.0500, 'cost_amount': 1050.00},
                {'fund_code': '021457', 'fund_name': '易方达恒生红利低波ETF联接A', 'shares': 800, 'cost_price': 1.2938, 'cost_amount': 1035.04},
                {'fund_code': '012725', 'fund_name': '国泰中证畜牧养殖ETF联接C', 'shares': 1000, 'cost_price': 0.7806, 'cost_amount': 780.60},
                {'fund_code': '004253', 'fund_name': '国泰黄金ETF联接C', 'shares': 400, 'cost_price': 2.7374, 'cost_amount': 1094.96},
                {'fund_code': '016814', 'fund_name': '国联中证煤炭指数C', 'shares': 800, 'cost_price': 1.7755, 'cost_amount': 1420.40},
                {'fund_code': '015016', 'fund_name': '华安德国(DAX)联接(QDII)C', 'shares': 700, 'cost_price': 2.0182, 'cost_amount': 1412.74},
                {'fund_code': '007280', 'fund_name': '摩根日本精选股票(QDII)A', 'shares': 800, 'cost_price': 1.8237, 'cost_amount': 1458.96},
                {'fund_code': '012060', 'fund_name': '富国全球消费精选混合(QDII)人民币A', 'shares': 800, 'cost_price': 1.6019, 'cost_amount': 1281.52},
                {'fund_code': '000834', 'fund_name': '大成纳斯达克100ETF联接(QDII)A', 'shares': 200, 'cost_price': 5.0848, 'cost_amount': 1016.96},
            ]
        }
        
        # 保存为Excel文件
        with pd.ExcelWriter('holdings_data.xlsx', engine='openpyxl') as writer:
            for user, holdings in holdings_data.items():
                df = pd.DataFrame(holdings)
                df.to_excel(writer, sheet_name=user, index=False)
        
        log_info("✓ 已创建示例持仓文件: holdings_data.xlsx")
        return holdings_data
    
    def update_holdings_data(self):
        """更新持仓数据文件，确保与当前基金代码列表匹配"""
        log_info("正在更新持仓数据文件...")
        return self.create_sample_holdings()
    
    def load_holdings_from_excel(self, filename='holdings_data.xlsx'):
        """从Excel文件加载持仓数据"""
        try:
            holdings_data = {}
            excel_file = pd.ExcelFile(filename)
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(filename, sheet_name=sheet_name)
                holdings = []
                
                for _, row in df.iterrows():
                    # 检查必需字段是否存在
                    required_fields = ['fund_code', 'fund_name', 'shares', 'cost_price', 'cost_amount']
                    missing_fields = [field for field in required_fields if field not in df.columns]
                    if missing_fields:
                        raise ValueError(f"缺少必需字段: {missing_fields}")
                    
                    # 确保基金代码是6位数字格式
                    fund_code = str(row['fund_code']).zfill(6)
                    
                    # 安全地转换数字字段
                    try:
                        shares = float(row['shares']) if pd.notna(row['shares']) else 0.0
                        cost_price = float(row['cost_price']) if pd.notna(row['cost_price']) else 0.0
                        cost_amount = float(row['cost_amount']) if pd.notna(row['cost_amount']) else 0.0
                    except (ValueError, TypeError) as e:
                        raise ValueError(f"数字字段转换失败: {e}，请检查shares、cost_price、cost_amount字段是否为数字")
                    
                    holding = {
                        'fund_code': fund_code,
                        'fund_name': str(row['fund_name']),
                        'shares': shares,
                        'cost_price': cost_price,
                        'cost_amount': cost_amount
                    }
                    
                    # 添加操作字段（如果存在）
                    operation_fields = ['buy_amount', 'sell_shares', 'convert_shares', 'convert_from_fund_code', 'convert_from_fund_name', 'convert_ratio']
                    for field in operation_fields:
                        if field in df.columns:
                            try:
                                if field in ['convert_from_fund_code']:
                                    holding[field] = str(row[field]).zfill(6) if pd.notna(row[field]) and str(row[field]).strip() else ''
                                elif field in ['convert_from_fund_name']:
                                    holding[field] = str(row[field]) if pd.notna(row[field]) else ''
                                else:
                                    holding[field] = float(row[field]) if pd.notna(row[field]) else 0
                            except (ValueError, TypeError) as e:
                                log_info(f"⚠️ 操作字段 {field} 转换失败: {e}，跳过该字段")
                                if field in ['convert_from_fund_code', 'convert_from_fund_name']:
                                    holding[field] = ''
                                else:
                                    holding[field] = 0
                    
                    holdings.append(holding)
                
                # 按持仓成本降序排序，重仓基金在前
                holdings.sort(key=lambda x: x['cost_amount'], reverse=True)
                holdings_data[sheet_name] = holdings
            
            log_info(f"✓ 已加载持仓文件: {filename}")
            return holdings_data
            
        except Exception as e:
            log_info(f"❌ 加载持仓文件失败: {e}")
            log_info("⚠️ 请检查持仓文件格式是否正确，包含以下必需字段：")
            log_info("   - fund_code: 基金代码")
            log_info("   - fund_name: 基金名称")
            log_info("   - shares: 持仓份额（数字）")
            log_info("   - cost_price: 成本单价（数字）")
            log_info("   - cost_amount: 持仓成本（数字）")
            log_info("   可选操作字段：buy_amount, sell_shares, convert_shares, convert_from_fund_code, convert_from_fund_name, convert_ratio")
            log_info("   请手工修改持仓文件后重新运行程序")
            return None
    
    def validate_holdings_data(self, holdings_data, fund_data_dict):
        """验证持仓数据是否与基金数据匹配"""
        if not holdings_data or not fund_data_dict:
            return False
        
        # 获取所有基金代码
        all_fund_codes = set()
        if 'all' in fund_data_dict:
            for fund in fund_data_dict['all']:
                all_fund_codes.add(fund['基金代码'])
        else:
            for fund_list in fund_data_dict.values():
                for fund in fund_list:
                    all_fund_codes.add(fund['基金代码'])
        
        log_info(f"🔍 验证持仓数据匹配性...")
        log_debug(f"📊 基金数据中的代码数量: {len(all_fund_codes)}")
        log_debug(f"🔍 基金数据中的代码: {sorted(list(all_fund_codes))}")
        
        # 检查持仓数据中的基金代码是否都存在
        missing_codes = []
        for user, holdings in holdings_data.items():
            log_info(f"💰 检查 {user} 的持仓数据 ({len(holdings)} 只基金)")
            for holding in holdings:
                if holding['fund_code'] not in all_fund_codes:
                    missing_codes.append({
                        'user': user,
                        'fund_code': holding['fund_code'],
                        'fund_name': holding['fund_name']
                    })
                    log_info(f"⚠️  持仓数据中的基金代码 {holding['fund_code']} ({holding['fund_name']}) 在基金数据中未找到")
        
        if missing_codes:
            log_info(f"❌ 发现 {len(missing_codes)} 个不匹配的基金代码:")
            for item in missing_codes:
                log_info(f"  - {item['user']}: {item['fund_code']} ({item['fund_name']})")
            return False
        
        log_info("✅ 所有持仓数据都匹配成功")
        return True
    
    def calculate_holdings_profit(self, holdings_data, fund_data_dict):
        """根据已获取的基金数据计算持仓收益"""
        results = {}
        
        # 将基金数据转换为字典格式，便于查找
        # 只使用自选基金数据，排除境外基金和ETF基金
        fund_data = {}
        # 合并钞钞和垚垚的基金数据，排除境外基金
        for user_key in ['chaochao', 'yaoyao']:
            if user_key in fund_data_dict:
                for fund in fund_data_dict[user_key]:
                    # 排除境外基金和ETF基金
                    if fund.get('板块分类') not in ['境外基金', 'ETF基金']:
                        fund_data[fund['基金代码']] = fund
        
        log_info(f"📊 可用于收益计算的基金数量: {len(fund_data)} 只")
        # 详细代码列表仅在调试模式下输出
        available_codes = list(fund_data.keys())
        log_debug(f"🔍 可用基金代码: {available_codes}")
        
        for user, holdings in holdings_data.items():
            log_info(f"💰 计算 {user} 持仓收益 ({len(holdings)} 只基金)")
            user_results = {
                'total_cost': 0,
                'total_current_value': 0,
                'total_profit': 0,
                'total_profit_rate': 0,
                'today_profit': 0,
                'today_profit_rate': 0,
                'holdings': []
            }
            
            for holding in holdings:
                fund_code = holding['fund_code']
                fund_info = fund_data.get(fund_code)
                
                if fund_info:
                    # 获取基金数据，处理可能的"N/A"值
                    try:
                        # 使用标准逻辑：估算净值是当前价格，最新净值是昨日净值
                        current_price = float(fund_info.get('估算净值', 0)) if fund_info.get('估算净值') != 'N/A' else 0
                        yesterday_price = float(fund_info.get('最新净值', 0)) if fund_info.get('最新净值') != 'N/A' else 0
                    except (ValueError, TypeError):
                        current_price = 0
                        yesterday_price = 0
                    
                    # 持仓信息
                    shares = holding['shares']
                    cost_price = holding['cost_price']
                    cost_amount = holding['cost_amount']
                    
                    # 计算当前价值
                    current_value = shares * current_price
                    yesterday_value = shares * yesterday_price
                    
                    # 计算收益
                    total_profit = current_value - cost_amount
                    total_profit_rate = (total_profit / cost_amount) * 100 if cost_amount > 0 else 0
                    
                    # 计算今日收益
                    today_profit = current_value - yesterday_value
                    today_profit_rate = (today_profit / yesterday_value) * 100 if yesterday_value > 0 else 0
                    
                    # 更新总计
                    user_results['total_cost'] += cost_amount
                    user_results['total_current_value'] += current_value
                    user_results['total_profit'] += total_profit
                    user_results['today_profit'] += today_profit
                    
                    # 保存单个持仓结果
                    holding_result = {
                        'fund_code': fund_code,
                        'fund_name': fund_info.get('基金名称', holding['fund_name']),
                        'shares': shares,
                        'cost_price': cost_price,
                        'cost_amount': cost_amount,
                        'current_price': current_price,
                        'current_value': current_value,
                        'total_profit': total_profit,
                        'total_profit_rate': total_profit_rate,
                        'today_profit': today_profit,
                        'today_profit_rate': today_profit_rate,
                        'change_rate': fund_info.get('估算涨跌率', '0.00'),
                        'category': fund_info.get('板块分类', '未知')
                    }
                    user_results['holdings'].append(holding_result)
                    # 精准调试：检查016814在各用户下的持仓与收益是否独立
                    # if fund_code == '016814':
                        # print(f"🔎 调试016814 - 用户:{user} 份额:{shares} 成本价:{cost_price} 成本金额:{cost_amount} 当前价:{current_price} 当前市值:{current_value:.2f} 总收益:{total_profit:.2f}")
                    log_debug(f"✅ 持仓匹配成功: {fund_code} ({fund_info.get('基金名称', holding['fund_name'])})")
                else:
                    log_info(f"⚠️  未找到基金 {fund_code} ({holding['fund_name']}) 的数据")
                    log_debug(f"💡 该基金可能不在当前获取的基金列表中，或者代码不匹配")
            
            # 计算总收益率
            if user_results['total_cost'] > 0:
                user_results['total_profit_rate'] = (user_results['total_profit'] / user_results['total_cost']) * 100
                user_results['today_profit_rate'] = (user_results['today_profit'] / user_results['total_current_value']) * 100
            
            log_info(f"📈 {user}: 成本 {user_results['total_cost']:,.0f}, 收益 {user_results['total_profit']:,.0f} ({user_results['total_profit_rate']:+.1f}%)")
            results[user] = user_results
        
        return results
    
    def enhance_fund_data_with_holdings(self, fund_data_dict, profit_results):
        """将持仓收益信息添加到基金数据中"""
        enhanced_dict = {}
        
        for user, fund_list in fund_data_dict.items():
            if user == 'all':
                # 对于'all'键，需要合并所有用户的基金并添加持仓信息
                all_funds = []
                all_holdings = {}
                
                # 合并所有用户的持仓信息，但保持用户区分
                for user_key in ['chaochao', 'yaoyao']:
                    if user_key in profit_results:
                        user_holdings = profit_results[user_key].get('holdings', [])
                        for holding in user_holdings:
                            fund_code = holding['fund_code']
                            # 使用用户+基金代码作为唯一键，避免覆盖
                            unique_key = f"{user_key}_{fund_code}"
                            all_holdings[unique_key] = {
                                'user': user_key,
                                'holding': holding
                            }
                
                # 为所有基金添加持仓信息（包括境外基金和ETF基金）
                for user_key in ['chaochao', 'yaoyao']:
                    if user_key in fund_data_dict:
                        for fund in fund_data_dict[user_key]:
                            fund_code = fund['基金代码']
                            enhanced_fund = fund.copy()
                            
                            # 查找对应的持仓信息，优先使用当前用户的持仓
                            user_holding_key = f"{user_key}_{fund_code}"
                            holding_info = all_holdings.get(user_holding_key)
                            
                            if holding_info and holding_info['user'] == user_key:
                                holding = holding_info['holding']
                                enhanced_fund.update({
                                    '成本单价': f"{holding['cost_price']:.4f}",
                                    '当日收益': f"{holding['today_profit']:,.2f}",
                                    '持仓收益': f"{holding['total_profit']:,.2f}",
                                    '持仓收益率': f"{holding['total_profit_rate']:+.2f}%"
                                })
                            else:
                                # 如果没有持仓信息，填充默认值
                                enhanced_fund.update({
                                    '成本单价': 'N/A',
                                    '当日收益': 'N/A',
                                    '持仓收益': 'N/A',
                                    '持仓收益率': 'N/A'
                                })
                            
                            all_funds.append(enhanced_fund)
                
                # 添加境外基金和ETF基金（这些基金没有持仓信息，但需要显示在'all'中）
                if 'overseas' in fund_data_dict:
                    for fund in fund_data_dict['overseas']:
                        enhanced_fund = fund.copy()
                        enhanced_fund.update({
                            '成本单价': 'N/A',
                            '当日收益': 'N/A',
                            '持仓收益': 'N/A',
                            '持仓收益率': 'N/A'
                        })
                        all_funds.append(enhanced_fund)
                
                if 'etf' in fund_data_dict:
                    for fund in fund_data_dict['etf']:
                        enhanced_fund = fund.copy()
                        enhanced_fund.update({
                            '成本单价': 'N/A',
                            '当日收益': 'N/A',
                            '持仓收益': 'N/A',
                            '持仓收益率': 'N/A'
                        })
                        all_funds.append(enhanced_fund)
                
                enhanced_dict[user] = all_funds
                continue
            
            enhanced_funds = []
            user_profit_data = profit_results.get(user, {})
            user_holdings = user_profit_data.get('holdings', [])
            
            # 创建持仓查找字典
            holdings_lookup = {holding['fund_code']: holding for holding in user_holdings}
            
            for fund in fund_list:
                fund_code = fund['基金代码']
                enhanced_fund = fund.copy()
                
                # 查找对应的持仓信息，只保留4个关键字段
                holding = holdings_lookup.get(fund_code)
                if holding:
                    enhanced_fund.update({
                        '成本单价': f"{holding['cost_price']:.4f}",
                        '当日收益': f"{holding['today_profit']:,.2f}",
                        '持仓收益': f"{holding['total_profit']:,.2f}",
                        '持仓收益率': f"{holding['total_profit_rate']:+.2f}%"
                    })
                else:
                    # 如果没有持仓信息，填充默认值
                    enhanced_fund.update({
                        '成本单价': 'N/A',
                        '当日收益': 'N/A',
                        '持仓收益': 'N/A',
                        '持仓收益率': 'N/A'
                    })
                
                enhanced_funds.append(enhanced_fund)
            
            enhanced_dict[user] = enhanced_funds
        
        return enhanced_dict
    
    def save_profit_report(self, profit_results, filename=None):
        """保存持仓收益报告"""
        if not filename:
            filename = f"持仓收益报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        
        print(f"调试: 开始生成收益报告，数据包含 {len(profit_results)} 个用户")
        for user, result in profit_results.items():
            print(f"调试: 用户 {user} 有 {len(result.get('holdings', []))} 个持仓")
        
        html_template = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>持仓收益报告</title>
    <style>
body {{
    font-family: 'Microsoft YaHei', Arial, sans-serif;
    margin: 20px;
    background-color: #f5f5f5;
}}
.container {{
    max-width: 1400px;
    margin: 0 auto;
    background-color: white;
    padding: 20px;
    border-radius: 10px;
    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
}}
h1, h2 {{
    color: #333;
    text-align: center;
}}
.summary-card {{
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    color: white;
    padding: 20px;
    border-radius: 10px;
    margin: 20px 0;
    text-align: center;
}}
.profit-positive {{
    color: #dc3545;
    font-weight: bold;
}}
.profit-negative {{
    color: #28a745;
    font-weight: bold;
}}
.profit-neutral {{
    color: #6c757d;
}}
table {{
    width: 100%;
    border-collapse: collapse;
    margin: 20px 0;
    background-color: white;
}}
th {{
    background-color: #007bff;
    color: white;
    padding: 12px 8px;
    text-align: center;
    font-weight: bold;
}}
td {{
    padding: 10px 8px;
    text-align: center;
    border-bottom: 1px solid #ddd;
}}
tr:nth-child(even) {{
    background-color: #f8f9fa;
}}
tr:hover {{
    background-color: #e9ecef;
}}
.timestamp {{
    text-align: center;
    color: #6c757d;
    font-size: 14px;
    margin-bottom: 20px;
}}
    </style>
</head>
<body>
    <div class="container">
        <h1>持仓收益报告</h1>
        <div class="timestamp">生成时间: {timestamp}</div>
        
        {summary_sections}
        
        {detail_sections}
    </div>
</body>
</html>"""
        
        # 生成汇总部分
        summary_sections = ""
        detail_sections = ""
        
        for user, result in profit_results.items():
            user_name = "钞钞" if user == "chaochao" else "垚垚"
            
            # 汇总卡片
            summary_sections += f"""
            <div class="summary-card">
                <h2>{user_name}的持仓汇总</h2>
                <p><strong>总投入:</strong> {result['total_cost']:,.2f}</p>
                <p><strong>当前市值:</strong> {result['total_current_value']:,.2f}</p>
                <p><strong>总收益:</strong> <span class="{'profit-positive' if result['total_profit'] > 0 else 'profit-negative' if result['total_profit'] < 0 else 'profit-neutral'}">{result['total_profit']:,.2f} ({result['total_profit_rate']:+.2f}%)</span></p>
                <p><strong>今日收益:</strong> <span class="{'profit-positive' if result['today_profit'] > 0 else 'profit-negative' if result['today_profit'] < 0 else 'profit-neutral'}">{result['today_profit']:,.2f} ({result['today_profit_rate']:+.2f}%)</span></p>
            </div>
            """
            
            # 详细表格
            if result['holdings']:  # 只有当有持仓数据时才显示表格
                detail_sections += f"""
                <h2>{user_name}的持仓明细</h2>
                <table>
                    <thead>
                        <tr>
                            <th>基金代码</th>
                            <th>基金名称</th>
                            <th>持仓份额</th>
                            <th>成本价</th>
                            <th>当前价格</th>
                            <th>当前市值</th>
                            <th>总收益</th>
                            <th>总收益率</th>
                            <th>今日收益</th>
                            <th>今日涨跌</th>
                        </tr>
                    </thead>
                    <tbody>
                """
                
                for holding in result['holdings']:
                    detail_sections += f"""
                        <tr>
                            <td>{holding['fund_code']}</td>
                            <td>{holding['fund_name']}</td>
                            <td>{holding['shares']:,.0f}</td>
                            <td>{holding['cost_price']:.4f}</td>
                            <td>{holding['current_price']:.4f}</td>
                            <td>{holding['current_value']:,.2f}</td>
                            <td class="{'profit-positive' if holding['total_profit'] > 0 else 'profit-negative' if holding['total_profit'] < 0 else 'profit-neutral'}">{holding['total_profit']:,.2f}</td>
                            <td class="{'profit-positive' if holding['total_profit_rate'] > 0 else 'profit-negative' if holding['total_profit_rate'] < 0 else 'profit-neutral'}">{holding['total_profit_rate']:+.2f}%</td>
                            <td class="{'profit-positive' if holding['today_profit'] > 0 else 'profit-negative' if holding['today_profit'] < 0 else 'profit-neutral'}">{holding['today_profit']:,.2f}</td>
                            <td class="{'profit-positive' if float(holding['change_rate']) > 0 else 'profit-negative' if float(holding['change_rate']) < 0 else 'profit-neutral'}">{holding['change_rate']}%</td>
                        </tr>
                    """
                
                detail_sections += """
                    </tbody>
                </table>
                """
            else:
                detail_sections += f"""
                <h2>{user_name}的持仓明细</h2>
                <p style="text-align: center; color: #6c757d; padding: 20px;">暂无持仓数据</p>
                """
        
        # 生成HTML文件
        html_content = html_template.format(
            timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            summary_sections=summary_sections,
            detail_sections=detail_sections
        )
        
        with open(filename, 'w', encoding='utf-8-sig') as f:
            f.write(html_content)
        
        log_info(f"✓ 持仓收益报告已保存: {filename}")
        return filename
    
    def backup_holdings_data(self, holdings_data, backup_filename='holdings_backup.xlsx'):
        """备份持仓数据"""
        try:
            with pd.ExcelWriter(backup_filename, engine='openpyxl') as writer:
                for user, holdings in holdings_data.items():
                    df = pd.DataFrame(holdings)
                    df.to_excel(writer, sheet_name=user, index=False)
            
            log_info(f"✓ 持仓数据已备份到: {backup_filename}")
            return True
        except Exception as e:
            log_info(f"备份持仓数据失败: {e}")
            return False
    
    def update_holdings_from_trading_data(self, holdings_data, trading_data, fund_nav_data, fund_data_dict=None):
        """根据交易数据更新持仓信息"""
        try:
            # 先备份原始数据
            if not self.backup_holdings_data(holdings_data):
                log_info("⚠️ 备份失败，停止更新")
                return False
            
            updated_holdings = {}
            cleared_funds = set()  # 记录需要清仓的基金
            trades_by_user = {}    # 收集本次交易记录
            
            # 创建基金代码到基金名称的映射，用于更新基金名称
            fund_name_map = {}
            if fund_data_dict:
                for user_key in ['chaochao', 'yaoyao', 'all']:
                    if user_key in fund_data_dict:
                        for fund in fund_data_dict[user_key]:
                            fund_code = fund.get('基金代码', '')
                            fund_name = fund.get('基金名称', '')
                            if fund_code and fund_name:
                                fund_name_map[fund_code] = fund_name
                log_info(f"📋 交易更新时基金名称映射表包含 {len(fund_name_map)} 个基金")

            # 构建昨日净值与估算净值映射，优先使用 fund_data_dict 中的数据
            yesterday_nav_map = {}
            estimate_nav_map = {}
            if fund_data_dict:
                for user_key in ['chaochao', 'yaoyao', 'all']:
                    if user_key in fund_data_dict:
                        for fund in fund_data_dict[user_key]:
                            code = fund.get('基金代码', '')
                            if code:
                                try:
                                    y_nav = float(fund.get('最新净值', 0)) if fund.get('最新净值', 'N/A') != 'N/A' else 0
                                except Exception:
                                    y_nav = 0
                                try:
                                    e_nav = float(fund.get('估算净值', 0)) if fund.get('估算净值', 'N/A') != 'N/A' else 0
                                except Exception:
                                    e_nav = 0
                                if y_nav > 0:
                                    yesterday_nav_map[code] = y_nav
                                if e_nav > 0:
                                    estimate_nav_map[code] = e_nav

            # 不进行人为强制覆盖昨日净值，完全以实时抓取/列表数据为准
            
            for user, holdings in holdings_data.items():
                updated_holdings[user] = []
                user_holdings = {holding['fund_code']: holding for holding in holdings}
                # 记录本轮已新增/更新的基金代码，避免重复行
                added_codes = set()
                trades_by_user[user] = []
                
                # 处理交易数据
                for trade in trading_data:
                    if trade.get('user') != user:
                        continue
                    
                    fund_code = str(trade['fund_code']).zfill(6)
                    # 使用最新的基金名称，如果找不到则使用交易数据中的名称
                    old_fund_name = trade.get('fund_name', '')
                    fund_name = fund_name_map.get(fund_code, old_fund_name)
                    if fund_name != old_fund_name and old_fund_name:
                        log_info(f"✓ {user} 交易操作中更新基金名称: {fund_code} {old_fund_name} -> {fund_name}")
                    shares = float(trade.get('shares', 0))
                    cost_price = float(trade.get('cost_price', 0))
                    cost_amount = float(trade.get('cost_amount', 0))
                    buy_amount = float(trade.get('buy_amount', 0))
                    sell_shares = float(trade.get('sell_shares', 0))
                    convert_shares = float(trade.get('convert_shares', 0))
                    convert_from_fund_code = str(trade.get('convert_from_fund_code', '')).zfill(6) if trade.get('convert_from_fund_code') else ''
                    convert_from_fund_name = trade.get('convert_from_fund_name', '')
                    convert_ratio = float(trade.get('convert_ratio', 1))
                    
                    # 获取基金净值：估算净值用于当日估值，昨日净值用于建仓成本
                    nav_estimate = 0
                    try:
                        nav_estimate = float(fund_nav_data.get(fund_code, 0)) if fund_nav_data else 0
                    except Exception:
                        nav_estimate = 0
                    if nav_estimate <= 0:
                        nav_estimate = estimate_nav_map.get(fund_code, 0)
                    nav_yesterday = yesterday_nav_map.get(fund_code, 0)
                    if nav_estimate <= 0 and nav_yesterday <= 0:
                        log_info(f"⚠️ 基金 {fund_code} 净值数据不可用，跳过更新")
                        continue
                    
                    # 处理建仓（仅处理买入建仓，转换建仓在转换逻辑中处理）
                    if buy_amount > 0 and (fund_code not in user_holdings or (user_holdings.get(fund_code, {}).get('shares', 0) or 0) <= 0):
                        # 买入建仓
                        base_price = nav_yesterday if nav_yesterday > 0 else (nav_estimate if nav_estimate > 0 else 0)
                        new_shares = buy_amount / base_price if base_price > 0 else 0
                        new_shares = round(new_shares, 2)
                        new_cost_price = round(base_price, 4)
                        new_cost_amount = buy_amount
                        if base_price <= 0 or new_shares <= 0:
                            log_info(f"⚠️ {user} 买入建仓跳过 {fund_code}: 无效价格或份额(base={base_price}, shares={new_shares})")
                            continue
                        # 记录交易
                        trades_by_user[user].append(
                            self._build_trade_record(fund_code, fund_name, 'buy', new_shares, base_price, new_cost_amount)
                        )
                        
                        new_holding = {
                            'fund_code': fund_code,
                            'fund_name': fund_name,
                            'shares': new_shares,
                            'cost_price': new_cost_price,
                            'cost_amount': new_cost_amount,
                            # 保留其他列的默认值
                            'buy_amount': 0,
                            'sell_shares': 0,
                            'convert_shares': 0,
                            'convert_from_fund_code': '',
                            'convert_from_fund_name': '',
                            'convert_ratio': 1
                        }
                        # 去重：若已存在于新增集合，则合并
                        if fund_code in added_codes:
                            # 合并到已添加的记录（查找并加权合并）
                            for item in updated_holdings[user]:
                                if item.get('fund_code') == fund_code:
                                    total_cost = item['cost_amount'] + new_cost_amount
                                    total_shares = item['shares'] + new_shares
                                    item['shares'] = round(total_shares, 2)
                                    item['cost_amount'] = round(total_cost, 2)
                                    item['cost_price'] = round((total_cost / total_shares) if total_shares > 0 else item['cost_price'], 4)
                                    break
                        else:
                            updated_holdings[user].append(new_holding)
                            added_codes.add(fund_code)
                        log_info(f"✓ {user} 买入建仓 {fund_code}: {new_shares:.2f}份 @{new_cost_price:.4f}")
                    
                    # 处理加仓
                    elif fund_code in user_holdings and buy_amount > 0:
                        existing = user_holdings[fund_code]
                        # 加仓份额按当日估算净值计算
                        price_for_add = nav_estimate if nav_estimate > 0 else (nav_yesterday if nav_yesterday > 0 else 0)
                        new_shares = buy_amount / price_for_add if price_for_add > 0 else 0
                        total_shares = existing['shares'] + new_shares
                        total_cost = existing['cost_amount'] + buy_amount
                        new_cost_price = total_cost / total_shares
                        
                        existing['shares'] = total_shares
                        existing['cost_price'] = new_cost_price
                        existing['cost_amount'] = total_cost
                        # 记录交易
                        trades_by_user[user].append(
                            self._build_trade_record(fund_code, fund_name, 'buy', new_shares, price_for_add, buy_amount)
                        )
                        log_info(f"✓ {user} 加仓 {fund_code}: +{new_shares:.2f}份，新成本价 {new_cost_price:.4f}")
                    
                    # 处理减仓
                    elif fund_code in user_holdings and sell_shares > 0:
                        existing = user_holdings[fund_code]
                        # 以当日估值/昨日净值作为成交价
                        sell_price = nav_estimate if nav_estimate > 0 else (nav_yesterday if nav_yesterday > 0 else existing.get('cost_price', 0))
                        sell_amount = sell_shares * sell_price
                        if sell_shares >= existing['shares']:
                            # 清仓
                            cleared_funds.add((user, fund_code))
                            trades_by_user[user].append(
                                self._build_trade_record(fund_code, fund_name, 'sell', sell_shares, sell_price, sell_amount)
                            )
                            log_info(f"✓ {user} 清仓 {fund_code}: 卖出 {sell_shares:.2f}份")
                        else:
                            # 减仓，重新计算成本价
                            remaining_shares = existing['shares'] - sell_shares
                            # 公式：(昨日净值 - x) * (持仓份额 - 卖出份额) = (昨日净值 - 成本单价) * 持仓份额
                            # 解出 x = 昨日净值 - (昨日净值 - 成本单价) * 持仓份额 / (持仓份额 - 卖出份额)
                            old_cost_price = existing['cost_price']
                            price_for_recalc = nav_yesterday if nav_yesterday > 0 else (nav_estimate if nav_estimate > 0 else old_cost_price)
                            new_cost_price = price_for_recalc - (price_for_recalc - old_cost_price) * existing['shares'] / remaining_shares
                            new_cost_amount = remaining_shares * new_cost_price
                            
                            existing['shares'] = remaining_shares
                            existing['cost_price'] = new_cost_price
                            existing['cost_amount'] = new_cost_amount
                            trades_by_user[user].append(
                                self._build_trade_record(fund_code, fund_name, 'sell', sell_shares, sell_price, sell_amount)
                            )
                            log_info(f"✓ {user} 减仓 {fund_code}: -{sell_shares:.2f}份，新成本价 {new_cost_price:.4f}")
                    
                    # 处理基金转换
                    if convert_from_fund_code and convert_from_fund_code in user_holdings:
                        from_fund = user_holdings[convert_from_fund_code]
                        # 如果转出份额略大于可用份额，则使用可用份额进行转换（避免转换失败）
                        actual_convert_shares = min(convert_shares, from_fund['shares'])
                        if actual_convert_shares > 0:
                            # 获取转出基金的净值（优先使用昨日净值，因为转换操作是上一交易日发生的）
                            from_fund_nav = 0
                            # 优先使用昨日净值
                            from_fund_nav = yesterday_nav_map.get(convert_from_fund_code, 0)
                            if from_fund_nav <= 0:
                                # 如果昨日净值不可用，再尝试当日估值净值
                                try:
                                    from_fund_nav = float(fund_nav_data.get(convert_from_fund_code, 0)) if fund_nav_data else 0
                                except Exception:
                                    from_fund_nav = 0
                                if from_fund_nav <= 0:
                                    from_fund_nav = estimate_nav_map.get(convert_from_fund_code, 0)
                            if from_fund_nav <= 0:
                                log_info(f"⚠️ {user} 基金转换失败: {convert_from_fund_code} 净值数据不可用")
                                continue
                            
                            # 计算转出金额 = 实际转出份额 * 转出基金净值
                            convert_amount = actual_convert_shares * from_fund_nav
                            trades_by_user[user].append(
                                self._build_trade_record(convert_from_fund_code, fund_name_map.get(convert_from_fund_code, convert_from_fund_code), 'convert_out', actual_convert_shares, from_fund_nav, convert_amount)
                            )
                            
                            # 转换出 - 需要重算成本单价
                            remaining_shares = from_fund['shares'] - actual_convert_shares
                            if remaining_shares > 0:
                                # 重算成本单价：(昨日净值 - x) * (持仓份额-转出份额) = (昨日净值 - 成本单价) * 持仓份额
                                # 解出 x = 昨日净值 - (昨日净值 - 成本单价) * 持仓份额 / (持仓份额 - 转出份额)
                                old_cost_price = from_fund['cost_price']
                                new_cost_price = from_fund_nav - (from_fund_nav - old_cost_price) * from_fund['shares'] / remaining_shares
                                new_cost_amount = remaining_shares * new_cost_price
                                
                                from_fund['shares'] = remaining_shares
                                from_fund['cost_price'] = new_cost_price
                                from_fund['cost_amount'] = new_cost_amount
                                
                                log_info(f"✓ {user} 基金转换出: {convert_from_fund_code} {actual_convert_shares:.2f}份(金额{convert_amount:.2f}元)，剩余份额成本价 {new_cost_price:.4f}")
                            else:
                                # 全部转换出，清仓
                                cleared_funds.add((user, convert_from_fund_code))
                                log_info(f"✓ {user} 基金转换清仓: {convert_from_fund_code} {actual_convert_shares:.2f}份(金额{convert_amount:.2f}元)")
                            
                            # 转换入 - 使用昨日净值计算份额（操作是上一交易日发生的）
                            in_price = nav_yesterday if nav_yesterday > 0 else (nav_estimate if nav_estimate > 0 else 0)
                            # 份额计算也使用昨日净值，因为转换操作是上一交易日发生的
                            shares_price_for_calc = nav_yesterday if nav_yesterday > 0 else (nav_estimate if nav_estimate > 0 else 0)
                            new_shares = (convert_amount / shares_price_for_calc) * convert_ratio if shares_price_for_calc > 0 else 0
                            new_shares = round(new_shares, 2)
                            new_cost_price = round(in_price, 4)
                            new_cost_amount = new_shares * new_cost_price
                            if new_shares <= 0 or new_cost_price <= 0:
                                log_info(f"⚠️ {user} 基金转换建仓跳过 {fund_code}: 无效价格或份额(price={new_cost_price}, shares={new_shares})")
                                continue
                            trades_by_user[user].append(
                                self._build_trade_record(fund_code, fund_name, 'convert_in', new_shares, new_cost_price, new_cost_amount)
                            )
                            
                            if fund_code not in user_holdings:
                                # 转换建仓
                                new_holding = {
                                    'fund_code': fund_code,
                                    'fund_name': fund_name,
                                    'shares': new_shares,
                                    'cost_price': new_cost_price,
                                    'cost_amount': new_cost_amount,
                                    # 保留其他列的默认值
                                    'buy_amount': 0,
                                    'sell_shares': 0,
                                    'convert_shares': 0,
                                    'convert_from_fund_code': '',
                                    'convert_from_fund_name': '',
                                    'convert_ratio': 1
                                }
                                if fund_code in added_codes:
                                    for item in updated_holdings[user]:
                                        if item.get('fund_code') == fund_code:
                                            total_cost = item['cost_amount'] + new_cost_amount
                                            total_shares = item['shares'] + new_shares
                                            item['shares'] = round(total_shares, 2)
                                            item['cost_amount'] = round(total_cost, 2)
                                            item['cost_price'] = round((total_cost / total_shares) if total_shares > 0 else item['cost_price'], 4)
                                            break
                                else:
                                    updated_holdings[user].append(new_holding)
                                    added_codes.add(fund_code)
                                log_info(f"✓ {user} 基金转换建仓: {convert_from_fund_code} -> {fund_code}, {actual_convert_shares:.2f}份->{new_shares:.2f}份(金额{convert_amount:.2f}元)")
                            else:
                                # 转换加仓 - 与买入加仓逻辑一致
                                existing = user_holdings[fund_code]
                                total_shares = existing['shares'] + new_shares
                                total_cost = existing['cost_amount'] + new_cost_amount
                                new_cost_price = total_cost / total_shares
                                
                                existing['shares'] = total_shares
                                existing['cost_price'] = new_cost_price
                                existing['cost_amount'] = total_cost
                                log_info(f"✓ {user} 基金转换加仓: {convert_from_fund_code} -> {fund_code}, {actual_convert_shares:.2f}份->{new_shares:.2f}份(金额{convert_amount:.2f}元)，新成本价 {new_cost_price:.4f}")
                        else:
                            log_info(f"⚠️ {user} 基金转换失败: {convert_from_fund_code} 无可用份额 ({from_fund['shares']:.2f} <= 0)")
                
                # 添加未清仓的基金（保留变更字段但清空数据）
                for fund_code, holding in user_holdings.items():
                    if (user, fund_code) not in cleared_funds and fund_code not in added_codes:
                        # 获取最新的基金名称
                        latest_fund_name = fund_name_map.get(fund_code, holding['fund_name'])
                        if latest_fund_name != holding['fund_name']:
                            log_info(f"✓ {user} 更新基金名称: {fund_code} {holding['fund_name']} -> {latest_fund_name}")
                        
                        # 创建新的持仓记录，保留所有字段但清空操作数据
                        clean_holding = {
                            'fund_code': holding['fund_code'],
                            'fund_name': latest_fund_name,  # 使用最新的基金名称
                            'shares': round(holding['shares'], 2),  # 保留2位小数
                            'cost_price': round(holding['cost_price'], 4),  # 成本单价保留4位小数
                            'cost_amount': round(holding['cost_amount'], 2),  # 保留2位小数
                            'buy_amount': 0,  # 清空操作数据
                            'sell_shares': 0,
                            'convert_shares': 0,
                            'convert_from_fund_code': '',
                            'convert_from_fund_name': '',
                            'convert_ratio': 1
                        }
                        updated_holdings[user].append(clean_holding)
                        added_codes.add(fund_code)
                
                # 过滤无效/空值行，并按持仓成本降序排序，重仓基金在前
                filtered_rows = []
                for item in updated_holdings[user]:
                    code = str(item.get('fund_code', '')).strip()
                    shares = float(item.get('shares', 0) or 0)
                    cost_amount = float(item.get('cost_amount', 0) or 0)
                    if not code or code == '000000':
                        continue
                    if shares <= 0 or cost_amount < 0:
                        continue
                    # 标准化份额与成本单价的小数位
                    item['shares'] = round(shares, 2)
                    if 'cost_price' in item and item['cost_price']:
                        try:
                            item['cost_price'] = round(float(item['cost_price']), 4)
                        except Exception:
                            pass
                    item['cost_amount'] = round(cost_amount, 2)
                    filtered_rows.append(item)
                filtered_rows.sort(key=lambda x: x['cost_amount'], reverse=True)
                updated_holdings[user] = filtered_rows
            
            # 保存更新后的持仓数据
            holdings_excel_file = 'holdings_data.xlsx'  # 持仓数据文件名
            trade_excel_file = 'trade_data.xlsx'  # 交易记录文件名
            self.save_updated_holdings(updated_holdings, holdings_excel_file)
            log_info("✓ 持仓数据更新完成，操作字段已清除")
            # 保存交易记录（同时保存到JSON和Excel）
            self._save_trades_json(trades_by_user, directory='trades', excel_file=trade_excel_file)
            return True
            
        except Exception as e:
            log_info(f"更新持仓数据失败: {e}")
            return False
    
    def save_updated_holdings(self, holdings_data, filename='holdings_data.xlsx'):
        """保存更新后的持仓数据"""
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                for user, holdings in holdings_data.items():
                    df = pd.DataFrame(holdings)
                    df.to_excel(writer, sheet_name=user, index=False)
            
            log_info(f"✓ 更新后的持仓数据已保存到: {filename}")
            return True
        except Exception as e:
            log_info(f"保存更新后的持仓数据失败: {e}")
            return False
    
    # ===== 持仓交易记录辅助方法 =====
    def _prev_trading_day(self, current_date=None):
        """获取上一交易日（简单按工作日回退）"""
        from datetime import timedelta
        if current_date is None:
            current_date = datetime.now().date()
        prev_day = current_date - timedelta(days=1)
        while prev_day.weekday() >= 5:  # 5,6 为周末
            prev_day = prev_day - timedelta(days=1)
        return prev_day
    
    def _build_trade_record(self, fund_code, fund_name=None, action=None, shares=None, price=None, amount=None, trade_date=None):
        """构造标准交易记录（精简版：仅保留核心字段）"""
        if trade_date is None:
            trade_date = self._prev_trading_day().strftime('%Y-%m-%d')
        else:
            # 确保date格式为yyyy-MM-dd（去除时间部分）
            if isinstance(trade_date, str):
                trade_date = trade_date.split()[0]  # 只取日期部分
            elif hasattr(trade_date, 'strftime'):
                trade_date = trade_date.strftime('%Y-%m-%d')
            else:
                trade_date = str(trade_date).split()[0]
        # tx_id格式：date-fund_code-action
        # 注意：同一天同一基金可以有多种操作（如买入和卖出），每种操作会有不同的tx_id
        tx_id = f"{trade_date}-{fund_code}-{action}"
        return {
            'fund_code': fund_code,
            'date': trade_date,
            'action': action,
            'shares': round(float(shares), 2) if shares is not None else 0.0,
            'tx_id': tx_id
        }
    
    def _load_trades_json(self, user, directory='trades'):
        """加载指定用户的交易JSON（若不存在返回空列表）"""
        try:
            filepath = os.path.join(directory, f"{user}.json")
            if not os.path.exists(filepath):
                return []
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return data.get('trades', []) if isinstance(data, dict) else []
        except Exception as e:
            log_info(f"⚠️ 无法读取交易文件: {e}")
            return []
    
    def create_trade_data_template(self, excel_file='trade_data.xlsx'):
        """创建交易记录Excel模板文件（按用户分sheet），如果文件已存在则只添加缺失的sheet"""
        try:
            headers = ['fund_code', 'date', 'action', 'shares', 'tx_id']  # 移除user列，因为sheet名已表示用户
            df = pd.DataFrame(columns=headers)
            
            # 如果文件已存在，先备份，然后只添加缺失的sheet
            if os.path.exists(excel_file):
                log_info(f"📄 文件 {excel_file} 已存在，将只添加缺失的sheet，保留现有数据...")
                # 先备份
                self.backup_trades_data(directory='trades', excel_file=excel_file)
                
                # 使用openpyxl加载现有文件
                from openpyxl import load_workbook
                wb = load_workbook(excel_file)
                existing_sheets = set(wb.sheetnames)
                
                # 检查哪些sheet缺失
                missing_sheets = [user for user in SUPPORTED_USERS if user not in existing_sheets]
                
                if missing_sheets:
                    log_info(f"📋 发现缺失的sheet: {', '.join(missing_sheets)}")
                    # 添加缺失的sheet
                    for user in missing_sheets:
                        ws = wb.create_sheet(user)
                        ws.append(headers)
                    wb.save(excel_file)
                    log_info(f"✓ 已添加缺失的sheet: {', '.join(missing_sheets)}")
                else:
                    log_info(f"✓ 所有必需的sheet已存在，无需添加")
                
                # 显示当前所有sheet
                log_info(f"📋 当前Excel文件包含的sheet: {', '.join(wb.sheetnames)}")
                return True
            else:
                # 文件不存在，创建新文件
                log_info(f"📄 文件 {excel_file} 不存在，正在创建新模板...")
                with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                    for user in SUPPORTED_USERS:
                        df.to_excel(writer, sheet_name=user, index=False)
                log_info(f"✓ 已创建交易记录模板文件: {excel_file} (包含{', '.join(SUPPORTED_USERS)}三个sheet)")
                return True
        except Exception as e:
            log_info(f"❌ 创建交易记录模板失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def backup_trades_data(self, directory='trades', excel_file='trade_data.xlsx'):
        """备份交易记录数据（JSON和Excel），每个文件只保留一份备份"""
        try:
            import shutil
            backup_success = True
            
            # 备份JSON文件
            for user in SUPPORTED_USERS:
                json_file = os.path.join(directory, f"{user}.json")
                if os.path.exists(json_file):
                    backup_json = os.path.join(directory, f"{user}_backup.json")
                    try:
                        shutil.copy2(json_file, backup_json)
                        log_info(f"✓ JSON备份: {json_file} -> {backup_json}")
                    except Exception as e:
                        log_info(f"⚠️ JSON备份失败 {json_file}: {e}")
                        backup_success = False
            
            # 备份Excel文件
            if os.path.exists(excel_file):
                backup_excel = excel_file.replace('.xlsx', '_backup.xlsx')
                try:
                    shutil.copy2(excel_file, backup_excel)
                    log_info(f"✓ Excel备份: {excel_file} -> {backup_excel}")
                except Exception as e:
                    log_info(f"⚠️ Excel备份失败 {excel_file}: {e}")
                    backup_success = False
            
            return backup_success
        except Exception as e:
            log_info(f"❌ 备份交易记录失败: {e}")
            return False
    
    def restore_trades_from_backup(self, excel_file='trade_data.xlsx', directory='trades'):
        """从备份恢复交易记录数据（JSON和Excel）"""
        try:
            import shutil
            restore_success = True
            
            # 恢复JSON文件
            for user in SUPPORTED_USERS:
                backup_json = os.path.join(directory, f"{user}_backup.json")
                json_file = os.path.join(directory, f"{user}.json")
                if os.path.exists(backup_json):
                    try:
                        shutil.copy2(backup_json, json_file)
                        log_info(f"✓ JSON恢复: {backup_json} -> {json_file}")
                    except Exception as e:
                        log_info(f"⚠️ JSON恢复失败 {backup_json}: {e}")
                        restore_success = False
                else:
                    log_info(f"⚠️ 未找到备份文件: {backup_json}")
            
            # 恢复Excel文件
            backup_excel = excel_file.replace('.xlsx', '_backup.xlsx')
            if os.path.exists(backup_excel):
                try:
                    shutil.copy2(backup_excel, excel_file)
                    log_info(f"✓ Excel恢复: {backup_excel} -> {excel_file}")
                except Exception as e:
                    log_info(f"⚠️ Excel恢复失败 {backup_excel}: {e}")
                    restore_success = False
            else:
                log_info(f"⚠️ 未找到备份文件: {backup_excel}")
            
            if restore_success:
                log_info("✅ 所有备份文件已恢复")
            else:
                log_info("⚠️ 部分备份文件恢复失败，请检查日志")
            
            return restore_success
        except Exception as e:
            log_info(f"❌ 恢复交易记录失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _save_trades_json(self, trades_by_user, directory='trades', excel_file='trade_data.xlsx'):
        """将交易记录写入JSON和Excel，按用户拆文件，增量合并"""
        try:
            # 先备份现有数据
            self.backup_trades_data(directory, excel_file)
            
            os.makedirs(directory, exist_ok=True)
            all_trades_for_excel = []  # 收集所有交易记录用于Excel备份
            
            for user, records in trades_by_user.items():
                existing = self._load_trades_json(user, directory)
                # 按 tx_id 去重合并
                merged = {item.get('tx_id'): item for item in existing if item.get('tx_id')}
                for rec in records:
                    merged[rec.get('tx_id')] = rec
                final_list = list(merged.values())
                # 统一处理date格式为yyyy-MM-dd（去除时间部分）
                for trade in final_list:
                    if 'date' in trade and trade['date']:
                        date_val = trade['date']
                        if isinstance(date_val, str):
                            trade['date'] = date_val.split()[0]  # 只取日期部分
                        elif hasattr(date_val, 'strftime'):
                            trade['date'] = date_val.strftime('%Y-%m-%d')
                        else:
                            trade['date'] = str(date_val).split()[0]
                # 按日期排序
                final_list.sort(key=lambda x: (x.get('date', ''), x.get('fund_code', ''), x.get('action', '')))
                payload = {
                    'user': user,
                    'updated_at': datetime.now().isoformat(),
                    'trades': final_list
                }
                filepath = os.path.join(directory, f"{user}.json")
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump(payload, f, ensure_ascii=False, indent=2)
                log_info(f"✓ 交易记录已保存JSON: {filepath} ({len(final_list)} 条)")
                
                # 收集所有合并后的交易记录（用于Excel备份，包含历史记录）
                # final_list中的date已经统一处理为yyyy-MM-dd格式
                for rec in final_list:
                    all_trades_for_excel.append({
                        'user': user,
                        'fund_code': rec.get('fund_code', ''),
                        'date': rec.get('date', ''),
                        'action': rec.get('action', ''),
                        'shares': rec.get('shares', 0),
                        'tx_id': rec.get('tx_id', '')
                    })
            
            # 保存到Excel（按用户分sheet，增量追加）
            if all_trades_for_excel and excel_file:
                # 如果文件不存在，先创建模板
                if not os.path.exists(excel_file):
                    self.create_trade_data_template(excel_file)
                
                if os.path.exists(excel_file):
                    try:
                        from openpyxl import load_workbook
                        wb = load_workbook(excel_file)
                        
                        # 兼容已有Excel：如果存在"交易记录"sheet，迁移数据到用户sheet
                        if '交易记录' in wb.sheetnames:
                            log_info("📋 检测到旧格式Excel，正在迁移数据到用户sheet...")
                            try:
                                df_old = pd.read_excel(excel_file, sheet_name='交易记录')
                                old_trades = df_old.to_dict('records')
                                # 按用户分组迁移
                                for trade in old_trades:
                                    user = trade.get('user', '').strip()
                                    if user in SUPPORTED_USERS:
                                        # 添加到对应用户的记录中
                                        all_trades_for_excel.append({
                                            'user': user,
                                            'fund_code': str(trade.get('fund_code', '')).strip().zfill(6),
                                            'date': str(trade.get('date', '')).strip().split()[0],
                                            'action': str(trade.get('action', '')).strip(),
                                            'shares': float(trade.get('shares', 0)) if pd.notna(trade.get('shares')) else 0.0,
                                            'tx_id': str(trade.get('tx_id', '')).strip()
                                        })
                                log_info(f"✓ 已迁移 {len(old_trades)} 条旧记录")
                            except Exception as e:
                                log_info(f"⚠️ 迁移旧数据失败: {e}")
                            # 删除旧的交易记录sheet
                            wb.remove(wb['交易记录'])
                        
                        # 按用户分组处理
                        trades_by_user_excel = {}
                        for rec in all_trades_for_excel:
                            user = rec.get('user', '')
                            if user not in trades_by_user_excel:
                                trades_by_user_excel[user] = []
                            # 移除user字段（sheet名已表示用户）
                            trade_rec = {
                                'fund_code': rec.get('fund_code', ''),
                                'date': rec.get('date', ''),
                                'action': rec.get('action', ''),
                                'shares': rec.get('shares', 0),
                                'tx_id': rec.get('tx_id', '')
                            }
                            trades_by_user_excel[user].append(trade_rec)
                        
                        # 为每个用户处理sheet
                        headers = ['fund_code', 'date', 'action', 'shares', 'tx_id']
                        total_count = 0
                        for user in SUPPORTED_USERS:
                            # 读取现有记录
                            existing_trades = []
                            try:
                                if user in wb.sheetnames:
                                    df_existing = pd.read_excel(excel_file, sheet_name=user)
                                    existing_trades = df_existing.to_dict('records')
                                    # 统一date格式
                                    for trade in existing_trades:
                                        if 'date' in trade and trade['date']:
                                            if pd.notna(trade['date']):
                                                if isinstance(trade['date'], pd.Timestamp):
                                                    trade['date'] = trade['date'].strftime('%Y-%m-%d')
                                                elif isinstance(trade['date'], str):
                                                    trade['date'] = trade['date'].split()[0]
                                                else:
                                                    trade['date'] = str(trade['date']).split()[0]
                            except:
                                pass  # sheet不存在或为空
                            
                            # 合并：按tx_id去重
                            merged_trades = {}
                            for item in existing_trades:
                                tx_id = item.get('tx_id', '')
                                if tx_id:
                                    merged_trades[tx_id] = item
                                else:
                                    # 如果没有tx_id，使用date-fund_code-action作为key
                                    date = item.get('date', '')
                                    fund_code = item.get('fund_code', '')
                                    action = item.get('action', '')
                                    if date and fund_code and action:
                                        key = f"{date}-{fund_code}-{action}"
                                        merged_trades[key] = item
                            
                            # 添加新记录
                            for rec in trades_by_user_excel.get(user, []):
                                tx_id = rec.get('tx_id', '')
                                if tx_id:
                                    merged_trades[tx_id] = rec
                                else:
                                    date = rec.get('date', '')
                                    fund_code = rec.get('fund_code', '')
                                    action = rec.get('action', '')
                                    if date and fund_code and action:
                                        key = f"{date}-{fund_code}-{action}"
                                        rec['tx_id'] = key
                                        merged_trades[key] = rec
                            
                            # 排序
                            final_list = list(merged_trades.values())
                            final_list.sort(key=lambda x: (x.get('date', ''), x.get('fund_code', '')))
                            
                            # 更新或创建sheet
                            if user in wb.sheetnames:
                                wb.remove(wb[user])
                            ws = wb.create_sheet(user)
                            ws.append(headers)
                            for trade in final_list:
                                ws.append([trade.get(h, '') for h in headers])
                            
                            total_count += len(final_list)
                            log_info(f"✓ {user}: {len(final_list)} 条记录")
                        
                        wb.save(excel_file)
                        log_info(f"✓ 交易记录已备份到Excel: {excel_file} (按用户分sheet, 共 {total_count} 条)")
                    except Exception as e:
                        log_info(f"⚠️ 保存交易记录到Excel失败: {e}")
                        import traceback
                        traceback.print_exc()
            
            return True
        except Exception as e:
            log_info(f"❌ 保存交易记录失败: {e}")
            return False
    
    def import_trades_from_excel(self, excel_file='trade_data.xlsx', directory='trades'):
        """从Excel导入交易记录到JSON文件（全量覆盖），支持按用户分sheet或旧的统一sheet"""
        try:
            # 先备份现有JSON数据
            self.backup_trades_data(directory, excel_file)
            
            if not os.path.exists(excel_file):
                log_info(f"⚠️ Excel文件不存在: {excel_file}")
                return False
            
            # 按用户分组处理
            os.makedirs(directory, exist_ok=True)
            trades_by_user = {}
            
            # 尝试从用户sheet读取（新格式）
            has_user_sheets = False
            for user in SUPPORTED_USERS:
                try:
                    df = pd.read_excel(excel_file, sheet_name=user)
                    if not df.empty:
                        has_user_sheets = True
                        if user not in trades_by_user:
                            trades_by_user[user] = []
                        
                        for _, row in df.iterrows():
                            # 处理date字段：如果是Timestamp，转换为字符串，只保留yyyy-MM-dd格式
                            date_value = row.get('date', '')
                            if pd.notna(date_value):
                                if isinstance(date_value, pd.Timestamp):
                                    date_str = date_value.strftime('%Y-%m-%d')
                                else:
                                    date_str = str(date_value).strip().split()[0]  # 只取日期部分，去除时间
                            else:
                                date_str = ''
                            
                            fund_code = str(row.get('fund_code', '')).strip().zfill(6)
                            action = str(row.get('action', '')).strip()
                            
                            # 验证必要字段
                            if not fund_code or fund_code == '000000':
                                continue
                            if not date_str or date_str == 'nan':
                                continue
                            if not action or action == 'nan':
                                continue
                            
                            # 如果没有tx_id或tx_id为空，根据date-fund_code-action生成（精简格式）
                            tx_id = str(row.get('tx_id', '')).strip()
                            if not tx_id or tx_id == 'nan' or tx_id == '':
                                tx_id = f"{date_str}-{fund_code}-{action}"
                            
                            trade_record = {
                                'fund_code': fund_code,
                                'date': date_str,
                                'action': action,
                                'shares': float(row.get('shares', 0)) if pd.notna(row.get('shares')) else 0.0,
                                'tx_id': tx_id
                            }
                            
                            trades_by_user[user].append(trade_record)
                except:
                    pass  # sheet不存在，继续尝试其他sheet
            
            # 兼容旧格式：如果用户sheet不存在，尝试从"交易记录"sheet读取
            if not has_user_sheets:
                try:
                    df = pd.read_excel(excel_file, sheet_name='交易记录')
                    if not df.empty:
                        log_info("📋 检测到旧格式Excel（统一sheet），正在读取...")
                        for _, row in df.iterrows():
                            user = str(row.get('user', '')).strip()
                            if not user or user == 'nan' or user not in SUPPORTED_USERS:
                                continue
                            
                            if user not in trades_by_user:
                                trades_by_user[user] = []
                            
                            # 处理date字段
                            date_value = row.get('date', '')
                            if pd.notna(date_value):
                                if isinstance(date_value, pd.Timestamp):
                                    date_str = date_value.strftime('%Y-%m-%d')
                                else:
                                    date_str = str(date_value).strip().split()[0]
                            else:
                                date_str = ''
                            
                            fund_code = str(row.get('fund_code', '')).strip().zfill(6)
                            action = str(row.get('action', '')).strip()
                            
                            # 验证必要字段
                            if not fund_code or fund_code == '000000':
                                continue
                            if not date_str or date_str == 'nan':
                                continue
                            if not action or action == 'nan':
                                continue
                            
                            # 如果没有tx_id或tx_id为空，根据date-fund_code-action生成
                            tx_id = str(row.get('tx_id', '')).strip()
                            if not tx_id or tx_id == 'nan' or tx_id == '':
                                tx_id = f"{date_str}-{fund_code}-{action}"
                            
                            trade_record = {
                                'fund_code': fund_code,
                                'date': date_str,
                                'action': action,
                                'shares': float(row.get('shares', 0)) if pd.notna(row.get('shares')) else 0.0,
                                'tx_id': tx_id
                            }
                            
                            trades_by_user[user].append(trade_record)
                except:
                    pass  # 旧格式sheet也不存在
            
            # 全量覆盖到JSON文件（不合并，直接覆盖）
            total_count = 0
            for user, trades in trades_by_user.items():
                # 统一处理date格式为yyyy-MM-dd（去除时间部分）
                for trade in trades:
                    if 'date' in trade and trade['date']:
                        date_val = trade['date']
                        if isinstance(date_val, str):
                            trade['date'] = date_val.split()[0]  # 只取日期部分
                        elif hasattr(date_val, 'strftime'):
                            trade['date'] = date_val.strftime('%Y-%m-%d')
                        else:
                            trade['date'] = str(date_val).split()[0]
                
                # 按日期排序
                trades.sort(key=lambda x: (x.get('date', ''), x.get('fund_code', ''), x.get('action', '')))
                
                payload = {
                    'user': user,
                    'updated_at': datetime.now().isoformat(),
                    'trades': trades
                }
                filepath = os.path.join(directory, f"{user}.json")
                with open(filepath, 'w', encoding='utf-8') as f:
                    json.dump(payload, f, ensure_ascii=False, indent=2)
                
                total_count += len(trades)
                log_info(f"✓ {user}: 全量导入 {len(trades)} 条交易记录（覆盖原有数据）")
            
            log_info(f"✅ 从Excel全量导入交易记录完成，共导入 {total_count} 条记录")
            return True
            
        except Exception as e:
            log_info(f"❌ 从Excel导入交易记录失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def export_trades_to_excel(self, excel_file='trade_data.xlsx', directory='trades'):
        """从JSON文件全量导出交易记录到Excel（全量覆盖）"""
        try:
            # 先备份现有Excel数据
            if os.path.exists(excel_file):
                import shutil
                backup_excel = excel_file.replace('.xlsx', '_backup.xlsx')
                try:
                    shutil.copy2(excel_file, backup_excel)
                    log_info(f"✓ Excel备份: {excel_file} -> {backup_excel}")
                except Exception as e:
                    log_info(f"⚠️ Excel备份失败 {excel_file}: {e}")
            
            # 按用户收集交易记录
            trades_by_user = {}
            for user in SUPPORTED_USERS:
                json_file = os.path.join(directory, f"{user}.json")
                if os.path.exists(json_file):
                    try:
                        with open(json_file, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                        trades = data.get('trades', []) if isinstance(data, dict) else []
                        
                        # 统一处理date格式为yyyy-MM-dd（去除时间部分）
                        for trade in trades:
                            if 'date' in trade and trade['date']:
                                date_val = trade['date']
                                if isinstance(date_val, str):
                                    trade['date'] = date_val.split()[0]  # 只取日期部分
                                elif hasattr(date_val, 'strftime'):
                                    trade['date'] = date_val.strftime('%Y-%m-%d')
                                else:
                                    trade['date'] = str(date_val).split()[0]
                        
                        # 按日期排序
                        trades.sort(key=lambda x: (x.get('date', ''), x.get('fund_code', '')))
                        trades_by_user[user] = trades
                        log_info(f"✓ 读取 {user}.json: {len(trades)} 条记录")
                    except Exception as e:
                        log_info(f"⚠️ 读取 {json_file} 失败: {e}")
                        trades_by_user[user] = []
            
            if not any(trades_by_user.values()):
                log_info("⚠️ 没有找到任何交易记录")
                return False
            
            # 按用户分sheet写入Excel
            headers = ['fund_code', 'date', 'action', 'shares', 'tx_id']  # 移除user列
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                for user in SUPPORTED_USERS:
                    trades = trades_by_user.get(user, [])
                    if trades:
                        # 转换为DataFrame（移除user字段，因为sheet名已表示用户）
                        trade_list = []
                        for trade in trades:
                            trade_list.append({
                                'fund_code': trade.get('fund_code', ''),
                                'date': trade.get('date', ''),
                                'action': trade.get('action', ''),
                                'shares': trade.get('shares', 0),
                                'tx_id': trade.get('tx_id', '')
                            })
                        df = pd.DataFrame(trade_list, columns=headers)
                        df.to_excel(writer, sheet_name=user, index=False)
                    else:
                        # 即使没有数据也创建空sheet
                        df = pd.DataFrame(columns=headers)
                        df.to_excel(writer, sheet_name=user, index=False)
            
            total_count = sum(len(trades) for trades in trades_by_user.values())
            log_info(f"✅ 从JSON全量导出交易记录完成，共导出 {total_count} 条记录到 {excel_file} (按用户分sheet)")
            return True
            
        except Exception as e:
            log_info(f"❌ 从JSON导出交易记录失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def update_fund_names_in_holdings(self, holdings_data, fund_data_dict, filename='holdings_data.xlsx', verbose=True):
        """更新持仓文件中的基金名称"""
        try:
            if not fund_data_dict:
                if verbose:
                    log_info("⚠️ 没有基金数据，跳过基金名称更新")
                return False
            
            # 创建基金代码到基金名称的映射
            fund_name_map = {}
            for user_key in ['chaochao', 'yaoyao', 'all']:
                if user_key in fund_data_dict:
                    for fund in fund_data_dict[user_key]:
                        fund_code = fund.get('基金代码', '')
                        fund_name = fund.get('基金名称', '')
                        if fund_code and fund_name:
                            fund_name_map[fund_code] = fund_name
            
            if verbose:
                log_info(f"📋 基金名称映射表包含 {len(fund_name_map)} 个基金")
            
            if not fund_name_map:
                if verbose:
                    log_info("⚠️ 没有找到有效的基金名称映射")
                return False
            
            # 基金名称更新不需要单独备份，因为后续交易操作会统一备份
            
            updated_holdings = {}
            name_updated_count = 0
            
            for user, holdings in holdings_data.items():
                updated_holdings[user] = []
                for holding in holdings:
                    fund_code = holding['fund_code']
                    old_name = holding['fund_name']
                    new_name = fund_name_map.get(fund_code, old_name)
                    
                    if new_name != old_name:
                        if verbose:
                            log_info(f"✓ {user} 更新基金名称: {fund_code} {old_name} -> {new_name}")
                        name_updated_count += 1
                        holding['fund_name'] = new_name
                    
                    updated_holdings[user].append(holding)
                
                # 按持仓成本降序排序
                updated_holdings[user].sort(key=lambda x: x['cost_amount'], reverse=True)
            
            if name_updated_count > 0:
                # 保存更新后的持仓数据
                self.save_updated_holdings(updated_holdings, filename)
                if verbose:
                    log_info(f"✓ 已更新 {name_updated_count} 个基金名称")
                return True
            else:
                if verbose:
                    log_info("✓ 所有基金名称都是最新的，无需更新")
                return True
                
        except Exception as e:
            log_info(f"更新基金名称失败: {e}")
            return False

class OptimizedFundTracker:
    def __init__(self, max_workers=10):
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            "Accept": "*/*",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Referer": "https://fund.eastmoney.com/",
            "Origin": "https://fund.eastmoney.com"
        })
        self.classifier = FundCategoryClassifier()
        self.max_workers = max_workers
        self._lock = threading.Lock()
        # 抓取重试参数
        self.retry_attempts = 3
        self.retry_backoff_seconds = 0.4
    
    def _parse_jsonp(self, text):
        """解析 jsonpgz(JSONP) 响应，返回 dict 或 None。"""
        if not text:
            return None
        text = text.strip()
        if not text.startswith("jsonpgz("):
            return None
        # 取括号内内容
        start = text.find("(")
        end = text.rfind(")")
        if start == -1 or end == -1 or end <= start:
            return None
        json_str = text[start + 1:end].strip()
        # 处理空/null
        if not json_str or json_str == "null":
            return None
        try:
            data = json.loads(json_str)
            # 仅当为 dict 且包含基金代码时认为有效
            if isinstance(data, dict) and data.get("fundcode"):
                return data
        except Exception:
            return None
        return None
    
    def _fetch_single_fund(self, fund_code):
        """获取单个基金的实时信息"""
        # 优先 https，其次 http，并附加时间戳参数避免缓存
        ts = str(int(time.time() * 1000))
        url_candidates = [
            f"https://fundgz.1234567.com.cn/js/{fund_code}.js?rt={ts}",
            f"http://fundgz.1234567.com.cn/js/{fund_code}.js?rt={ts}"
        ]
        last_error = None
        for url in url_candidates:
            for attempt in range(self.retry_attempts):
                try:
                    response = self.session.get(url, timeout=8)
                    if response.status_code == 200:
                        data = self._parse_jsonp(response.text)
                        if data:
                            # 检查是否有今日净值
                            jzrq = data.get('jzrq', '')
                            today = datetime.now().strftime('%Y-%m-%d')
                            
                            # 如果净值日期是今天，使用今日净值替换估算净值
                            if jzrq == today:
                                data['gsz'] = data.get('dwjz', data.get('gsz', ''))
                                data['gszzl'] = '0.00'  # 今日净值涨跌率为0
                                data['gztime'] = f"{today} 15:00"  # 更新时间
                            
                            # 检查估算净值是否为空或无效，如果是则尝试备用方案
                            gsz = data.get('gsz', '')
                            if not gsz or gsz == '' or gsz == 'NaN' or gsz == 'null':
                                log_info(f"⚠️ {fund_code} 估算净值为空，尝试备用方案...")
                                
                                # 首先尝试使用最新净值作为估算净值
                                dwjz = data.get('dwjz', '')
                                if dwjz and dwjz != '' and dwjz != 'NaN' and dwjz != 'null':
                                    data['gsz'] = dwjz
                                    data['gszzl'] = '0.00'  # 使用最新净值时涨跌率为0
                                    data['gztime'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                    log_info(f"✅ {fund_code} 使用最新净值作为估算净值: {dwjz}")
                                else:
                                    # 如果最新净值也没有，尝试其他备用方案
                                    backup_data = self._fetch_fund_backup(fund_code)
                                    if backup_data:
                                        # 使用备用数据补充缺失的字段
                                        if not gsz and backup_data.get('gsz'):
                                            data['gsz'] = backup_data['gsz']
                                        if not data.get('gszzl') or data.get('gszzl') == 'NaN':
                                            data['gszzl'] = backup_data.get('gszzl', '0.00')
                                        if not data.get('gztime') or data.get('gztime') == '0001-01-01 00:00':
                                            data['gztime'] = backup_data.get('gztime', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                                        log_info(f"✅ {fund_code} 备用方案获取成功")
                                    else:
                                        log_info(f"⚠️ {fund_code} 所有备用方案均失败，使用最新净值")
                                        if dwjz:
                                            data['gsz'] = dwjz
                                            data['gszzl'] = '0.00'
                                            data['gztime'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            
                            return data
                    else:
                        last_error = f"HTTP {response.status_code}"
                except Exception as e:
                    last_error = str(e)
                # 指数退避 + 抖动
                time.sleep(self.retry_backoff_seconds * (2 ** attempt) + random.uniform(0, 0.2))
        with self._lock:
            if last_error:
                log_info(f"❌ {fund_code}: {last_error}")
            else:
                log_info(f"❌ {fund_code}: 未知错误")
        return None
    
    def _fetch_fund_backup(self, fund_code):
        """备用方案获取基金数据"""
        try:
            # 方案1: 尝试使用天天基金网的历史净值接口
            history_url = f"https://fund.eastmoney.com/f10/F10DataApi.aspx?type=lsjz&code={fund_code}&page=1&per=1"
            response = self.session.get(history_url, timeout=5)
            if response.status_code == 200:
                # 解析历史净值数据
                content = response.text
                if 'var apidata=' in content:
                    # 提取JSON数据
                    start = content.find('var apidata=') + 12
                    end = content.find(';', start)
                    if end > start:
                        json_str = content[start:end]
                        try:
                            import json
                            data = json.loads(json_str)
                            if data and 'content' in data and data['content']:
                                # 获取最新净值
                                latest_nav = data['content'][0].get('单位净值', '0')
                                if latest_nav and latest_nav != '0':
                                    return {
                                        'gsz': latest_nav,
                                        'gszzl': '0.00',  # 历史净值无法计算实时涨跌
                                        'gztime': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                    }
                        except:
                            pass
            
            # 方案2: 尝试使用新浪财经接口
            sina_url = f"https://hq.sinajs.cn/list=f_{fund_code}"
            response = self.session.get(sina_url, timeout=5)
            if response.status_code == 200:
                content = response.text
                if 'var hq_str_' in content:
                    # 解析新浪数据格式
                    parts = content.split('"')[1].split(',')
                    if len(parts) >= 3:
                        nav = parts[2]  # 净值
                        if nav and nav != '0':
                            return {
                                'gsz': nav,
                                'gszzl': '0.00',
                                'gztime': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
            
            # 方案3: 尝试使用腾讯财经接口
            tencent_url = f"https://qt.gtimg.cn/q=sz{fund_code}"
            response = self.session.get(tencent_url, timeout=5)
            if response.status_code == 200:
                content = response.text
                if 'v_' in content:
                    # 解析腾讯数据格式
                    parts = content.split('"')[1].split('~')
                    if len(parts) >= 3:
                        nav = parts[3]  # 净值
                        if nav and nav != '0':
                            return {
                                'gsz': nav,
                                'gszzl': '0.00',
                                'gztime': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            }
            
            # 方案4: 尝试使用东方财富网基金详情页
            eastmoney_url = f"https://fund.eastmoney.com/{fund_code}.html"
            response = self.session.get(eastmoney_url, timeout=5)
            if response.status_code == 200:
                content = response.text
                # 查找净值信息
                import re
                nav_pattern = r'<span class="ui-num">(\d+\.\d+)</span>'
                nav_match = re.search(nav_pattern, content)
                if nav_match:
                    nav = nav_match.group(1)
                    return {
                        'gsz': nav,
                        'gszzl': '0.00',
                        'gztime': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
            
            # 方案5: 尝试使用基金历史净值API
            nav_api_url = f"https://fund.eastmoney.com/f10/F10DataApi.aspx?type=lsjz&code={fund_code}&page=1&per=5"
            response = self.session.get(nav_api_url, timeout=5)
            if response.status_code == 200:
                content = response.text
                if 'var apidata=' in content:
                    start = content.find('var apidata=') + 12
                    end = content.find(';', start)
                    if end > start:
                        json_str = content[start:end]
                        try:
                            import json
                            data = json.loads(json_str)
                            if data and 'content' in data and data['content']:
                                # 获取最新净值
                                latest_record = data['content'][0]
                                nav = latest_record.get('单位净值', '0')
                                if nav and nav != '0':
                                    return {
                                        'gsz': nav,
                                        'gszzl': '0.00',
                                        'gztime': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                                    }
                        except:
                            pass
            
            log_info(f"⚠️ {fund_code} 所有备用方案均失败")
            return None
            
        except Exception as e:
            log_info(f"⚠️ {fund_code} 备用方案获取失败: {e}")
            return None
    
    def get_funds_realtime(self, fund_codes):
        """获取基金实时数据 - 并发版本"""
        if isinstance(fund_codes, str):
            fund_codes = [fund_codes]
        
        fund_data = []
        today = datetime.now().strftime('%Y-%m-%d')
        today_funds = 0
        estimate_funds = 0
        
        # 使用线程池并发请求
        with concurrent.futures.ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # 提交所有任务
            future_to_code = {executor.submit(self._fetch_single_fund, code): code for code in fund_codes}
            
            # 收集结果
            for future in concurrent.futures.as_completed(future_to_code):
                code = future_to_code[future]
                try:
                    data = future.result()
                    if data:
                        fund_data.append(data)
                        jzrq = data.get('jzrq', '')
                        if jzrq == today:
                            today_funds += 1
                            with self._lock:
                                log_debug(f"✅ {code} (今日净值)")
                        else:
                            estimate_funds += 1
                            with self._lock:
                                log_debug(f"✅ {code} (估算净值)")
                    else:
                        with self._lock:
                            log_debug(f"❌ {code}")
                except Exception as e:
                    with self._lock:
                        log_debug(f"❌ {code}: {e}")
        
        # 打印净值统计（简洁）
        log_info(f"📊 净值: 今日 {today_funds} 只, 估算 {estimate_funds} 只")
        
        return fund_data
    
    def get_fund_history_nav(self, fund_code, days=30):
        """获取基金历史净值数据 - 分页抓取（带冗余缓冲，避免翻页去重后不足）"""
        try:
            log_debug(f"正在获取基金 {fund_code} 的历史净值数据...")
            url = f"http://fund.eastmoney.com/f10/F10DataApi.aspx"
            # 目标天数加入冗余缓冲，避免跨页重复导致数量不足
            buffer_days = max(30, int(days * 0.4))
            target_unique = days + buffer_days
            per_page = 50  # 接口每页尽量取满
            max_pages = 12  # 安全上限，避免过多请求

            unique_by_date = {}
            page = 1
            while len(unique_by_date) < target_unique and page <= max_pages:
                params = {
                    'type': 'lsjz',
                    'code': fund_code,
                    'page': page,
                    'per': per_page,
                    'sdate': '',
                    'edate': ''
                }
                log_debug(f"请求第 {page} 页: {url}?{params}")
                response = self.session.get(url, params=params, timeout=15)
                if response.status_code != 200:
                    log_info(f"获取基金 {fund_code} 历史数据失败: HTTP {response.status_code} (page={page})")
                    break
                page_rows = self._parse_fund_history_html(response.text)
                if not page_rows:
                    # 没有更多数据
                    log_debug(f"第 {page} 页没有数据")
                    break
                log_debug(f"第 {page} 页获取到 {len(page_rows)} 条数据")
                for row in page_rows:
                    unique_by_date[row['date']] = row
                page += 1
                time.sleep(0.12)

            # 去重后按日期降序，截取所需天数
            deduped = list(unique_by_date.values())
            deduped.sort(key=lambda x: x['date'], reverse=True)
            result = deduped[:days]
            log_debug(f"基金 {fund_code} 最终获取到 {len(result)} 条历史数据")
            return result
        except Exception as e:
            log_info(f"获取基金 {fund_code} 历史数据失败: {e}")
            return []
    
    def _parse_fund_history_html(self, html_content):
        """解析HTML中的历史净值数据"""
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            table = soup.find('table', class_='w782 comm lsjz')
            
            if not table:
                # 尝试其他可能的表格选择器
                table = soup.find('table')
                if not table:
                    return []
            
            data = []
            rows = table.find_all('tr')[1:]  # 跳过表头
            
            for row in rows:
                cols = row.find_all('td')
                if len(cols) >= 4:
                    date = cols[0].text.strip()
                    nav = cols[1].text.strip()
                    acc_nav = cols[2].text.strip()
                    change_rate = cols[3].text.strip()
                    
                    try:
                        nav_value = float(nav)
                        data.append({
                            'date': date,
                            'nav': nav_value,
                            'acc_nav': acc_nav,
                            'change_rate': change_rate
                        })
                    except ValueError:
                        continue
            
            # 按日期排序（最新的在前）
            data.sort(key=lambda x: x['date'], reverse=True)
            
            # 返回全部解析到的数据（数量由API参数per控制）
            return data
            
        except Exception as e:
            log_info(f"解析历史净值HTML失败: {e}")
            return []
    
    def get_fund_history_nav_simple(self, fund_code, days=7):
        """获取基金历史净值的简化版本 - 使用另一个API端点"""
        try:
            # 使用另一个API端点，可能更稳定
            url = f"http://fund.eastmoney.com/f10/F10DataApi.aspx"
            
            params = {
                'type': 'lsjz',
                'code': fund_code,
                'page': 1,
                'per': days
            }
            
            response = self.session.get(url, params=params, timeout=15)
            if response.status_code == 200:
                return self._parse_fund_history_html(response.text)
            else:
                print(f"获取基金 {fund_code} 历史数据失败: HTTP {response.status_code}")
                return []
                
        except Exception as e:
            print(f"获取基金 {fund_code} 历史数据失败: {e}")
            return []

    def get_overseas_fund_data(self, fund_codes):
        """获取境外基金数据 - 直接使用天天基金网API"""
        overseas_funds = []
        
        for fund_code in fund_codes:
            try:
                # 直接调用天天基金网API获取境外基金数据
                fund_data = self._fetch_overseas_fund(fund_code, fund_code)
                if fund_data:
                    overseas_funds.append(fund_data)
                    log_debug(f"✅ 境外基金: {fund_data.get('name', 'N/A')} ({fund_code})")
                else:
                    log_info(f"❌ 境外基金数据获取失败: {fund_code}")
                    
            except Exception as e:
                log_info(f"❌ 境外基金 {fund_code} 处理失败: {e}")
                continue
        
        return overseas_funds
    
    def get_etf_fund_data(self, etf_codes):
        """获取场内ETF基金数据 - 支持 .SZ/.SH 后缀的ETF代码"""
        etf_funds = []
        
        for etf_code in etf_codes:
            try:
                # 调用ETF基金API
                fund_data = self._fetch_etf_fund(etf_code, etf_code)
                if fund_data:
                    etf_funds.append(fund_data)
                    log_debug(f"📊 ETF基金: {fund_data.get('name', 'N/A')} ({etf_code})")
                else:
                    log_info(f"❌ ETF基金数据获取失败: {etf_code}")
                    
            except Exception as e:
                log_info(f"❌ ETF基金 {etf_code} 处理失败: {e}")
                continue
        
        return etf_funds
    
    def _map_overseas_fund_code(self, fund_code):
        """境外基金代码映射规则"""
        # 移除后缀，获取纯数字代码
        if '.' in fund_code:
            base_code = fund_code.split('.')[0]
        else:
            base_code = fund_code
        
        # 根据后缀类型返回对应的API代码
        if fund_code.endswith('.OF'):
            # 开放式基金，使用天天基金网API
            return base_code
        elif fund_code.endswith('.SZ'):
            # 深圳交易所，使用深交所API
            return f"SZ{base_code}"
        elif fund_code.endswith('.SH'):
            # 上海交易所，使用上交所API
            return f"SH{base_code}"
        else:
            # 默认使用天天基金网API
            return base_code
    
    def _fetch_overseas_fund(self, fund_code, original_code):
        """获取境外基金数据"""
        try:
            # 优先使用天天基金网实时API
            url = f"http://fundgz.1234567.com.cn/js/{fund_code}.js"
            response = self.session.get(url, timeout=10)
            
            if response.status_code == 200:
                # 尝试解析JSONP数据
                data = self._parse_jsonp(response.text)
                if data:
                    return self._format_overseas_fund_data(data, original_code)
            
            # 如果实时API失败，尝试获取历史净值数据
            try:
                nav_data = self.get_fund_history_nav_simple(fund_code, 1)
                if nav_data and len(nav_data) > 0:
                    latest = nav_data[0]
                    return self._format_overseas_fund_data_from_nav(latest, original_code)
            except Exception as e:
                print(f"获取历史净值失败: {e}")
            
            return None
            
        except Exception as e:
            print(f"获取境外基金 {original_code} 失败: {e}")
            return None
    
    def _format_overseas_fund_data(self, data, original_code):
        """格式化境外基金数据（JSONP格式）"""
        return {
            'fundcode': original_code,
            'name': data.get('name', '境外基金'),
            'dwjz': data.get('dwjz', 'N/A'),  # 最新净值
            'gsz': data.get('gsz', 'N/A'),    # 估算净值
            'gszzl': data.get('gszzl', 'N/A'), # 估算涨跌率
            'jzrq': data.get('jzrq', 'N/A'),  # 净值日期
            'gztime': data.get('gztime', 'N/A'), # 估值时间
            'is_overseas': True
        }
    
    def _format_overseas_fund_data_from_nav(self, nav_data, original_code):
        """格式化境外基金数据（净值格式）"""
        # 尝试从多个API获取基金名称，如果都失败则使用预定义名称
        fund_name = self._get_overseas_fund_name(original_code)
        
        return {
            'fundcode': original_code,
            'name': fund_name,
            'dwjz': nav_data.get('nav', 'N/A'),
            'gsz': nav_data.get('nav', 'N/A'),
            'gszzl': '0.00',  # 历史净值无法计算涨跌率
            'jzrq': nav_data.get('date', 'N/A'),
            'gztime': 'N/A',
            'is_overseas': True
        }
    
    def _get_overseas_fund_name(self, fund_code):
        """获取境外基金名称的多种方法"""
        # 方法1：预定义的基金名称映射
        predefined_names = {
            "015016": "华安德国(DAX)联接(QDII)C",
            "007280": "摩根日本精选股票(QDII)A",
            "012060": "富国全球消费精选混合(QDII)人民币A",
            "012920": "易方达全球成长精选混合(QDII)人民币A",
            "000834": "大成纳斯达克100ETF联接(QDII)A",
            "270042": "广发纳斯达克100ETF联接人民币(QDII)A",
            "019671": "广发港股创新药ETF联接(QDII)C",
            "022122": "国泰海通中证香港科技指数发起(QDII)C",
            "006105": "宏利印度股票(QDII)",
            "020712": "华安三菱日联日经225ETF发起式联接(QDII)A",
            "019172": "摩根纳斯达克100指数(QDII)人民币A",
            "006373": "国富全球科技互联混合(QDII)人民币A",
            "024773": "摩根标普港股通低波红利ETF发起式联接C",
            "007911": "大成有色金属期货ETF联接C",
            "025733": "华安国证航天航空行业ETF发起式联接C",
            "017652": "中航华证商飞高端制造产业指数发起C",
            "009225": "天弘中证中美互联网(QDII)A",
            "014982": "华安标普全球石油指数(LOF)C",
            "021190": "南方亚太精选ETF联接(QDII)C"
        }
        
        if fund_code in predefined_names:
            log_debug(f"✅ 使用预定义名称: {predefined_names[fund_code]}")
            return predefined_names[fund_code]
        
        # 方法2：尝试天天基金网实时API
        try:
            url = f"http://fundgz.1234567.com.cn/js/{fund_code}.js"
            log_debug(f"🔍 尝试天天基金网API: {url}")
            response = self.session.get(url, timeout=5)
            if response.status_code == 200:
                data = self._parse_jsonp(response.text)
                if data and data.get('name'):
                    log_debug(f"✅ 天天基金网API成功: {data['name']}")
                    return data['name']
        except Exception as e:
            log_debug(f"⚠️  天天基金网API失败: {e}")
        
        # 方法3：尝试东方财富网基金详情页
        try:
            url = f"http://fund.eastmoney.com/{fund_code}.html"
            log_debug(f"🔍 尝试东方财富网详情页: {url}")
            response = self.session.get(url, timeout=8)
            if response.status_code == 200:
                from bs4 import BeautifulSoup
                soup = BeautifulSoup(response.text, 'html.parser')
                # 查找基金名称
                name_element = soup.find('div', class_='fundDetail-tit')
                if name_element:
                    name = name_element.get_text().strip()
                    if name:
                        log_debug(f"✅ 东方财富网详情页成功: {name}")
                        return name
        except Exception as e:
            log_debug(f"⚠️  东方财富网详情页失败: {e}")
        
        # 方法4：尝试天天基金网基金详情页
        try:
            url = f"http://fund.10jqka.com.cn/{fund_code}/"
            log_debug(f"🔍 尝试同花顺基金详情页: {url}")
            response = self.session.get(url, timeout=8)
            if response.status_code == 200:
                from bs4 import BeautifulSoup
                soup = BeautifulSoup(response.text, 'html.parser')
                # 查找基金名称
                title_element = soup.find('title')
                if title_element:
                    title_text = title_element.get_text()
                    if '基金' in title_text:
                        name = title_text.replace('基金', '').replace('_同花顺', '').strip()
                        if name:
                            log_debug(f"✅ 同花顺基金详情页成功: {name}")
                            return name
        except Exception as e:
            log_debug(f"⚠️  同花顺基金详情页失败: {e}")
        
        # 如果所有方法都失败，使用默认名称
        default_name = f"境外基金{fund_code}"
        log_info(f"⚠️  所有API方法都失败，使用默认名称: {default_name}")
        return default_name
    
    def _fetch_etf_fund(self, etf_code, original_code):
        """获取场内ETF基金数据 - 多接口备用方案"""
        try:
            # 解析ETF代码，获取交易所和代码
            if '.' in etf_code:
                base_code = etf_code.split('.')[0]
                exchange = etf_code.split('.')[1]
            else:
                base_code = etf_code
                exchange = 'SZ'  # 默认深交所
            
            # 方案1：尝试腾讯财经接口（主要备用方案）
            try:
                if exchange == 'SZ':
                    url = f"https://qt.gtimg.cn/q=sz{base_code}"
                elif exchange == 'SH':
                    url = f"https://qt.gtimg.cn/q=sh{base_code}"
                else:
                    url = f"https://qt.gtimg.cn/q=sz{base_code}"
                
                log_debug(f"尝试腾讯财经接口: {url}")
                response = self.session.get(url, timeout=10)
                if response.status_code == 200:
                    data = self._parse_tencent_etf_data(response.text, original_code)
                    if data:
                        log_debug(f"腾讯财经接口成功获取ETF数据: {original_code}")
                        return data
                        
            except Exception as e:
                log_debug(f"腾讯财经接口调用失败: {e}")
            
            # 方案2：尝试东方财富网ETF接口（原方案）
            try:
                if exchange == 'SZ':
                    url = f"https://push2.eastmoney.com/api/qt/stock/get?secid=0.{base_code}&fields=f43,f57,f58,f169,f170,f46,f44,f51,f168,f47,f48,f60,f45"
                elif exchange == 'SH':
                    url = f"https://push2.eastmoney.com/api/qt/stock/get?secid=1.{base_code}&fields=f43,f57,f58,f169,f170,f46,f44,f51,f168,f47,f48,f60,f45"
                else:
                    url = None
                
                if url:
                    log_debug(f"尝试东方财富接口: {url}")
                    response = self.session.get(url, timeout=10)
                    if response.status_code == 200:
                        data = response.json()
                        if data and data.get('rc') == 0:
                            payload = data.get('data') or {}
                            if payload:
                                return self._format_etf_fund_data(payload, original_code)
                        
            except Exception as e:
                log_debug(f"东方财富接口调用失败: {e}")
            
            # 方案3：尝试通过基金接口获取
            try:
                clean_code = base_code
                log_debug(f"尝试通过基金接口获取: {clean_code}")
                fund_data = self._fetch_single_fund(clean_code)
                if fund_data:
                    return self._format_etf_fund_data_from_fund(fund_data, original_code)
            except Exception as e:
                log_debug(f"基金接口调用失败: {e}")
            
            # 方案4：使用静态数据作为最后备用
            log_info(f"所有接口都失败，使用静态数据: {original_code}")
            return self._get_static_etf_data(original_code)
            
        except Exception as e:
            log_info(f"获取ETF基金 {original_code} 失败: {e}")
            return self._get_static_etf_data(original_code)
    
    def _parse_tencent_etf_data(self, response_text, original_code):
        """解析腾讯财经ETF数据"""
        try:
            # 腾讯财经返回格式：v_sz159513="51~纳斯达克100指数ETF~159513~1.554~1.535~1.550~558296~314711~243585~1.553~5922~1.552~4338~1.551~6830~1.550~2551~1.549~1812~1.554~16931~1.555~11911~1.556~6747~1.557~1655~1.558~6523~~2025102413105"
            if '=' in response_text:
                data_part = response_text.split('=')[1].strip('"')
                fields = data_part.split('~')
                
                if len(fields) >= 10:
                    # 字段说明：代码~名称~当前价~昨收价~今开价~成交量~成交额~买一量~买一价~买二量~买二价~买三量~买三价~买四量~买四价~买五量~买五价~卖一量~卖一价~卖二量~卖二价~卖三量~卖三价~卖四量~卖四价~卖五量~卖五价~时间
                    fund_code = fields[2]  # 基金代码在第3个字段
                    fund_name = fields[1]  # 基金名称在第2个字段
                    current_price = float(fields[3]) if fields[3] and fields[3] != '' else 0.0  # 当前价
                    prev_close = float(fields[4]) if fields[4] and fields[4] != '' else 0.0  # 昨收价
                    open_price = float(fields[5]) if fields[5] and fields[5] != '' else 0.0  # 今开价
                    volume = int(float(fields[6])) if fields[6] and fields[6] != '' else 0  # 成交量
                    amount = float(fields[7]) if fields[7] and fields[7] != '' else 0.0  # 成交额
                    
                    # 计算涨跌幅
                    if prev_close > 0:
                        change_rate = ((current_price - prev_close) / prev_close) * 100
                        change_amount = current_price - prev_close
                    else:
                        change_rate = 0.0
                        change_amount = 0.0
                    
                    return {
                        'fundcode': fund_code,
                        'name': fund_name,
                        'dwjz': f"{current_price:.4f}",  # 当前价格作为净值
                        'gsz': f"{current_price:.4f}",    # 当前价格
                        'gszzl': f"{change_rate:.2f}",    # 涨跌幅
                        'jzrq': datetime.now().strftime('%Y-%m-%d'),  # 当前日期
                        'gztime': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),  # 当前时间
                        'prev_close': f"{prev_close:.4f}",  # 昨收盘价
                        'open_price': f"{open_price:.4f}",  # 今开价
                        'volume': volume,  # 成交量
                        'amount': amount,  # 成交额
                        'change_amount': f"{change_amount:.4f}"  # 涨跌额
                    }
            
            return None
            
        except Exception as e:
            log_debug(f"解析腾讯财经ETF数据失败: {e}")
            return None
    
    def _get_static_etf_data(self, original_code):
        """获取静态ETF数据作为最后备用"""
        try:
            # 预定义的ETF静态数据
            static_etf_data = {
                "159513.SZ": {
                    'name': '纳斯达克100指数ETF',
                    'dwjz': '1.5540',
                    'gsz': '1.5540',
                    'gszzl': '1.24',
                    'jzrq': datetime.now().strftime('%Y-%m-%d'),
                    'gztime': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                },
                "513520.SH": {
                    'name': '日经ETF',
                    'dwjz': '1.2350',
                    'gsz': '1.2350',
                    'gszzl': '0.82',
                    'jzrq': datetime.now().strftime('%Y-%m-%d'),
                    'gztime': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                },
                "159919.SZ": {
                    'name': '沪深300ETF',
                    'dwjz': '3.4560',
                    'gsz': '3.4560',
                    'gszzl': '0.15',
                    'jzrq': datetime.now().strftime('%Y-%m-%d'),
                    'gztime': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                },
                "513500.SH": {
                    'name': '标普500ETF',
                    'dwjz': '2.1230',
                    'gsz': '2.1230',
                    'gszzl': '0.45',
                    'jzrq': datetime.now().strftime('%Y-%m-%d'),
                    'gztime': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                },
                "510500.SH": {
                    'name': '中证500ETF',
                    'dwjz': '4.5670',
                    'gsz': '4.5670',
                    'gszzl': '0.32',
                    'jzrq': datetime.now().strftime('%Y-%m-%d'),
                    'gztime': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
            }
            
            if original_code in static_etf_data:
                data = static_etf_data[original_code].copy()
                data['fundcode'] = original_code.split('.')[0]
                log_info(f"使用静态数据: {original_code}")
                return data
            else:
                # 返回默认数据
                return {
                    'fundcode': original_code.split('.')[0],
                    'name': f'ETF{original_code}',
                    'dwjz': '1.0000',
                    'gsz': '1.0000',
                    'gszzl': '0.00',
                    'jzrq': datetime.now().strftime('%Y-%m-%d'),
                    'gztime': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
        except Exception as e:
            log_info(f"获取静态ETF数据失败: {e}")
            return None
    
    def _format_etf_fund_data(self, data, original_code):
        """格式化ETF基金数据（东方财富接口）"""
        try:
            # 获取ETF名称
            etf_name = data.get('f58', '')
            if not etf_name:
                # 如果API没有返回名称，使用预定义的名称
                etf_names = {
                    "159513.SZ": "纳斯达克100指数ETF",
                    "513520.SH": "日经ETF",
                    "159919.SZ": "沪深300ETF"
                }
                etf_name = etf_names.get(original_code, f"ETF{original_code}")
            
            # 获取并规范化当前价格（不同接口字段倍率可能不同，做自适应归一）
            raw_price = data.get('f43', 0) or 0
            def normalize_price(v):
                # 根据实际ETF价格范围进行更精确的归一化
                # 纳斯达克100ETF等通常价格在1-2之间，沪深300ETF在3-4之间
                if v <= 0:
                    return 0.0
                
                # 尝试不同的除数，优先选择合理的价格范围
                candidates = [
                    v / 100.0,      # 原始值可能是分
                    v / 1000.0,     # 原始值可能是厘
                    v / 10000.0,    # 原始值可能是毫
                    v / 1.0         # 原始值可能已经是元
                ]
                
                # 优先选择在合理ETF价格范围内的值（0.1 ~ 10）
                for p in candidates:
                    if 0.1 <= p <= 10:
                        return p
                
                # 如果都不在合理范围内，选择最接近1的值
                best_candidate = min(candidates, key=lambda x: abs(x - 1.0))
                return best_candidate
            
            current_price = normalize_price(raw_price)
            
            # 获取昨收盘价（f60字段）
            raw_prev_close = data.get('f60', 0) or 0
            prev_close_price = normalize_price(raw_prev_close)
            
            # 获取涨跌幅
            change_rate = data.get('f170', 0) / 100.0  # 涨跌幅需要除以100
            
            # 获取涨跌额
            change_amount = data.get('f169', 0) / 100.0  # 涨跌额需要除以100
            
            # 获取成交量
            volume = data.get('f47', 0)
            
            # 获取成交额（单位万）
            raw_amount = data.get('f48', 0) or 0
            # 通常为分*股，接口已在万为单位，做自适应
            amount_candidates = [raw_amount / 10000.0, raw_amount / 100000.0]
            amount = amount_candidates[0] if amount_candidates[0] >= 0.01 else amount_candidates[1]
            
            # 获取当前时间，但需要调整为交易时间
            now = datetime.now()
            from datetime import timedelta
            
            # 判断是否为交易日（周一至周五）
            def is_trading_day(date_obj):
                """判断是否为交易日（简单判断：周一至周五）"""
                return date_obj.weekday() < 5  # 0-4 表示周一到周五
            
            # 获取上一个交易日
            def get_prev_trading_day(current_date):
                """获取上一个交易日"""
                prev_day = current_date - timedelta(days=1)
                while not is_trading_day(prev_day):
                    prev_day = prev_day - timedelta(days=1)
                return prev_day
            
            # 判断当前是否为交易时间
            is_trading_time = False
            if is_trading_day(now):
                # 交易时间：9:30-11:30, 13:00-15:00
                if ((9 <= now.hour < 11) or (now.hour == 11 and now.minute <= 30) or 
                    (13 <= now.hour < 15) or (now.hour == 15 and now.minute == 0)):
                    is_trading_time = True
            
            if is_trading_time:
                # 交易时间内，使用当前时间
                trade_time = now
                trade_date = now.strftime('%Y-%m-%d')
            else:
                # 非交易时间，使用上一个交易日的15:00
                prev_trade_day = get_prev_trading_day(now)
                trade_time = prev_trade_day.replace(hour=15, minute=0, second=0, microsecond=0)
                trade_date = prev_trade_day.strftime('%Y-%m-%d')
            
            # 昨收盘价对应的日期（上一个交易日）
            # 无论是否为当天，昨收盘价日期都应该是trade_date的上一个交易日
            trade_date_obj = datetime.strptime(trade_date, '%Y-%m-%d')
            prev_trade_date = get_prev_trading_day(trade_date_obj).strftime('%Y-%m-%d')
            
            return {
                'fundcode': original_code,
                'name': etf_name,
                'dwjz': f"{current_price:.4f}",  # 最新价格（当前价格）
                'gsz': f"{prev_close_price:.4f}",   # 昨收盘价
                'gszzl': f"{change_rate:.2f}",   # 涨跌率
                'jzrq': trade_date,  # 最新净值日期（交易日期）
                'gztime': trade_time.strftime('%Y-%m-%d %H:%M'),  # 估值时间（交易时间）
                'is_etf': True,
                'volume': volume,
                'amount': amount,
                'change_amount': change_amount,
                'prev_trade_date': prev_trade_date  # 昨收盘价对应的日期
            }
        except Exception as e:
            log_info(f"格式化ETF数据失败: {e}")
            return None
    

    
    def _format_etf_fund_data_from_fund(self, fund_data, original_code):
        """格式化ETF基金数据（基金接口）"""
        try:
            # 将基金数据转换为ETF格式
            etf_data = fund_data.copy()
            etf_data['fundcode'] = original_code
            etf_data['is_etf'] = True
            
            # 添加ETF特有字段
            if 'volume' not in etf_data:
                etf_data['volume'] = 0
            if 'amount' not in etf_data:
                etf_data['amount'] = 0
            if 'change_amount' not in etf_data:
                etf_data['change_amount'] = 0
            
            # 确保名称字段正确
            if 'name' not in etf_data or not etf_data['name']:
                etf_names = {
                    "159513.SZ": "华夏中证机器人ETF",
                    "513520.SH": "华夏中证机器人ETF", 
                    "159919.SZ": "嘉实沪深300ETF"
                }
                etf_data['name'] = etf_names.get(original_code, f"ETF{original_code}")
            
            return etf_data
        except Exception as e:
            log_info(f"转换ETF数据失败: {e}")
            return None

def get_self_selected_funds(max_workers=10):
    """获取自选基金信息 - 优化版本"""
    tracker = OptimizedFundTracker(max_workers=max_workers)
    
    # 钞钞的基金 - 根据实际持仓数据更新
    chaochao_fund_codes = [
        "013172[港股科技]",	# 华夏恒生互联网科技业ETF联接(QDII)C
        "013180[储能]",  # 广发国证新能源车电池ETF联接C
        "012414[消费]",  # 招商中证白酒指数
        "021536[科技]",  # 天弘中证软件服务指数发起C
        "023852[新能源]",  # 富国上证科创板新能源ETF发起式联接C
        "015401[机器人]",  # 弘毅甄选混合
        "015897[化工]",  # 天弘中证化工
        "009883[半导体]",  # 华润元大核心动力混合C
        "025209[半导体]",  # 永赢先锋半导体智选混合发起C
        "015790[军工]",  # 永赢高端装备智选混合发起C
        "025857[储能]",  # 华夏中证电网设备
        "023881[科技]",  # 兴全商业模式混合(LOF)C
        "024663[人工智能]",  # 富国创业板人工智能
    ]
    
    # 垚垚的基金 - 根据实际持仓数据更新
    yaoyao_fund_codes = [
        "013416[医疗器械]",	# 永赢中证全指医疗器械ETF发起联接C
        "004070[证券]",     # 南方中证全指证券公司ETF联接C
        "012725[畜牧养殖]",	# 国泰中证畜牧养殖ETF联接C
        "013309[港股科技]",	# 易方达恒生科技ETF联接(QDII)C
        "001595[金融]",  # 天弘中证银行ETF联接C
        "022365[通信]",  # 永赢智选混合
        "023754[人工智能]", # 永赢信息产业智选混合
        "016387[科技]",  # 永赢低碳环保智选混合发起C
        "016550[传媒]",  # 永赢消费龙头智选混合发起C
        "024738[储能]",  # 永赢新材料智选混合发起C
        "024203[科技]",  # 永赢制造升级混合
        "018647[家用电器]",  # 易方达家电龙头
        "002834[混合型]",  # 华夏新锦绣混合C
        "025209[半导体]",  # 永赢先锋半导体智选混合发起C
        "021172[指数型]",  # 华安北证50
        "004744[指数型]",  # 易方达创业板ETF联接C
    ]
    
    # 境外基金（实际可获取的基金代码）
    overseas_fund_codes = [
        "015016",  # 华安国际龙头(dax)
        "007280",  # 摩根日本精选股票
        "012060",  # 富国全球消费
        "012920",  # 易方达全球精选混合A
        "000834",  # 大成纳斯达克
        "270042",  # 广发纳指100
        "019671",  # 广发港股创新药ETF联接(QDII)C
        "022122",  # 国泰海通中证香港科技指数发起(QDII)C
        "006105",  # 宏利印度股票(QDII)
        "020712",  # 华安三菱日联日经225ETF发起式联接(QDII)A
        "019172",  # 摩根纳斯达克100指数(QDII)人民币A
        "006373",  # 国富全球科技互联混合(QDII)人民币A
        "024773",  # 摩根标普港股通低波红利ETF发起式联接C
        "007911",  # 大成有色金属期货ETF
        "025733",  # 华安国证航天航空行业ETF发起式联接C
        "017652",  # 中航华证商飞高端制造产业指数发起C
        "009225",  # 天弘中证中美互联网(QDII)A
        "014982",  # 华安标普全球石油指数(LOF)C
        "021190",  # 南方亚太精选ETF联接(QDII)C
    ]
    
    # ETF基金代码（场内交易）
    etf_fund_codes = [
        "159513.SZ",  # 纳斯达克100指数ETF
        "513520.SH",  # 日经ETF
        "159919.SZ",  # 沪深300ETF
        "513500.SH",  # 标普500ETF
        "510500.SH",  # 中证500ETF
        "159531.SZ",  # 中证2000ETF
        "520880.SH",  # 港股通创新药ETF
        "513180.SH",  # 恒生科技指数ETF
        "513330.SH",  # 恒生互联网ETF
        "512660.SH",  # 军工ETF
        "512710.SH",  # 军工龙头ETF
        "560710.SH",  # 船舶ETF
        "588000.SH",  # 科创5OETF
        "588200.SH",  # 科创芯片ETF
        "159915.SZ",  # 创业板ETF 
        "515880.SH",  # 通信ETF
        "515050.SH",  # 5G通信ETF 
        "159819.SZ",  # 人工智能ETF
        "516510.SH",  # 云计算ETF
        "562500.SH",  # 机器人ETF
        "159869.SZ",  # 游戏ETF
        "159928.SZ",  # 消费ETF
        "512690.SH",  # 酒ETF
        "518880.SH",  # 黄金ETF 
        "518600.SH",  # 上海金ETF
        "516780.SH",  # 稀士ETF
        "512400.SH",  # 有色金属ETF
        "515220.SH",  # 煤炭ETF
        "515790.SH",  # 光伏ETF
        "159825.SZ",  # 农业ETF
        "512800.SH",  # 银行ETF
        "510880.SH",  # 红利ETF
        "513630.SH",  # 港股红利ETF
        "159755.SZ",  # 电池ETF
        "159875.SZ",  # 新能源ETF
        "159512.SZ",  # 汽车ETF
        "159267.SZ",  # 航天ETF
        "159206.SZ",  # 卫星ETF
    ]
    
    # 合并自选基金代码（不包括境外基金，境外基金单独处理）
    # 提取纯基金代码，去掉自定义板块标签
    chaochao_pure_codes = [extract_pure_fund_code(code) for code in chaochao_fund_codes]
    yaoyao_pure_codes = [extract_pure_fund_code(code) for code in yaoyao_fund_codes]
    
    # 合并所有基金代码并去重
    all_pure_codes = chaochao_pure_codes + yaoyao_pure_codes
    my_fund_codes = list(set(all_pure_codes))  # 去重
    
    log_info("=== 自选基金信息 ===")
    log_info(f"🔍 并发请求: {max_workers} 线程")
    
    start_time = time.time()
    fund_data_raw = tracker.get_funds_realtime(my_fund_codes)
    end_time = time.time()
    
    log_info(f"⏱️  耗时: {end_time - start_time:.1f}s, 成功: {len(fund_data_raw)}/{len(my_fund_codes)} 只")
    log_info("=" * 50)
    
    # 获取境外基金数据
    log_info("\n=== 境外基金信息 ===")
    overseas_fund_data_raw = tracker.get_overseas_fund_data(overseas_fund_codes)
    log_info(f"🌍 境外基金: 成功 {len(overseas_fund_data_raw)}/{len(overseas_fund_codes)} 只")
    log_info("=" * 50)
    
    # 获取ETF基金数据
    log_info("\n=== ETF基金信息 ===")
    etf_fund_data_raw = tracker.get_etf_fund_data(etf_fund_codes)
    log_info(f"📊 ETF基金: 成功 {len(etf_fund_data_raw)}/{len(etf_fund_codes)} 只")
    log_info("=" * 50)
    
    # 按组分类基金数据
    chaochao_fund_data = []
    yaoyao_fund_data = []
    overseas_fund_data = []
    etf_fund_data = []
    
    for fund_info in fund_data_raw:
        if fund_info:
            fund_code = fund_info.get('fundcode', 'N/A')
            fund_name = fund_info.get('name', 'N/A')
            
            # 查找原始基金代码（带自定义标签）
            original_fund_code = ""
            for code in chaochao_fund_codes + yaoyao_fund_codes:
                if code.startswith(fund_code):
                    original_fund_code = code
                    break
            
            category = tracker.classifier.classify_fund(fund_name, original_fund_code)
            category_desc = tracker.classifier.get_category_description(category)
            
            # 检查估算净值日期是否与当前交易日一致
            # 如果日期不一致（如港股闭市），使用最新净值作为估算净值，涨跌幅设为0
            gztime = fund_info.get("gztime", "")
            today = datetime.now().strftime('%Y-%m-%d')
            dwjz = fund_info.get("dwjz", "N/A")
            gsz = fund_info.get("gsz", "N/A")
            gszzl = fund_info.get("gszzl", "N/A")
            
            # 提取估算净值日期（从 gztime 中提取日期部分）
            estimate_date = None
            if gztime and gztime != "N/A":
                try:
                    # gztime 格式可能是 "YYYY-MM-DD HH:mm" 或 "YYYY-MM-DD"
                    if ' ' in gztime:
                        estimate_date = gztime.split(' ')[0]
                    else:
                        estimate_date = gztime
                except:
                    pass
            
            # 如果估算净值日期不是今天，且最新净值有效，则使用最新净值
            if estimate_date and estimate_date != today:
                if dwjz != "N/A" and dwjz != "" and dwjz != "NaN" and dwjz != "null":
                    try:
                        # 验证最新净值是有效数字
                        float(dwjz)
                        gsz = dwjz  # 使用最新净值作为估算净值
                        gszzl = "0.00"  # 涨跌幅设为0
                        # 保持 gztime 为当前日期（用于列头显示），但实际净值是上日的
                        gztime = f"{today} 15:00"
                        log_debug(f"📅 {fund_name} ({fund_code}) 估算净值日期({estimate_date})非今日，使用最新净值({dwjz})，涨跌幅设为0")
                    except (ValueError, TypeError):
                        pass  # 如果最新净值无效，保持原值
            
            change_rate = gszzl
            change_symbol = "+" if change_rate != 'N/A' and float(change_rate) > 0 else ""
            log_debug(f"✅ {fund_name} ({fund_code}) {change_symbol}{change_rate}% [{category}]")
            
            fund_item = {
                "基金代码": fund_code,
                "基金名称": fund_name,
                "板块分类": category,
                "最新净值": dwjz,
                "估算净值": gsz,
                "估算涨跌率": gszzl,
                "净值日期": fund_info.get("jzrq", "N/A"),
                "估值时间": gztime
            }
            
            # 根据基金代码判断属于哪个组，为每个用户创建独立的基金数据项
            if fund_code in chaochao_pure_codes:
                # 为钞钞创建独立的基金数据项
                chaochao_fund_item = fund_item.copy()
                chaochao_fund_data.append(chaochao_fund_item)
            
            if fund_code in yaoyao_pure_codes:
                # 为垚垚创建独立的基金数据项
                yaoyao_fund_item = fund_item.copy()
                yaoyao_fund_data.append(yaoyao_fund_item)
            
            # 境外基金不在这里处理，在专门的境外基金处理循环中处理
    
    # 处理境外基金数据
    for fund_info in overseas_fund_data_raw:
        if fund_info:
            fund_code = fund_info.get('fundcode', 'N/A')
            fund_name = fund_info.get('name', 'N/A')
            
            log_debug(f"🌍 {fund_name} ({fund_code})")
            
            # 获取历史净值数据来计算涨跌幅
            try:
                history = tracker.get_fund_history_nav_simple(fund_code, 2)
                if history and len(history) >= 2:
                    # 最新净值（最近净值）
                    latest_nav = history[0]['nav']
                    # 上日净值
                    prev_nav = history[1]['nav']
                    # 计算涨跌幅
                    if prev_nav > 0:
                        change_rate = ((latest_nav - prev_nav) / prev_nav) * 100
                        change_rate_str = f"{change_rate:.2f}"
                    else:
                        change_rate_str = "0.00"
                    
                    fund_item = {
                        "基金代码": fund_code,
                        "基金名称": fund_name,
                        "板块分类": "境外基金",
                        "最新净值": f"{latest_nav:.4f}",  # 最新净值放在"最新净值"列
                        "估算净值": f"{prev_nav:.4f}",   # 上日净值放在"估算净值"列
                        "估算涨跌率": change_rate_str,
                        "净值日期": history[0]['date'] if len(history) >= 1 else "N/A",  # 最新净值日期
                        "估值时间": history[1]['date'] if len(history) >= 2 else "N/A"   # 上日净值日期
                    }
                else:
                    # 如果没有历史数据，使用原始数据
                    fund_item = {
                        "基金代码": fund_code,
                        "基金名称": fund_name,
                        "板块分类": "境外基金",
                        "最新净值": fund_info.get("dwjz", "N/A"),
                        "估算净值": fund_info.get("gsz", "N/A"),
                        "估算涨跌率": fund_info.get("gszzl", "N/A"),
                        "净值日期": fund_info.get("jzrq", "N/A"),
                        "估值时间": fund_info.get("gztime", "N/A")
                    }
            except Exception as e:
                log_info(f"处理境外基金 {fund_code} 历史数据时出错: {e}")
                fund_item = {
                    "基金代码": fund_code,
                    "基金名称": fund_name,
                    "板块分类": "境外基金",
                    "最新净值": fund_info.get("dwjz", "N/A"),
                    "估算净值": fund_info.get("gsz", "N/A"),
                    "估算涨跌率": fund_info.get("gszzl", "N/A"),
                    "净值日期": fund_info.get("jzrq", "N/A"),
                    "估值时间": fund_info.get("gztime", "N/A")
                }
            
            overseas_fund_data.append(fund_item)
            # 境外基金已经在上面添加到自选基金中了，这里不需要重复添加
    
    # 处理ETF基金数据
    for fund_info in etf_fund_data_raw:
        if fund_info:
            fund_code = fund_info.get('fundcode', 'N/A')
            fund_name = fund_info.get('name', 'N/A')
            
            log_debug(f"📊 {fund_name} ({fund_code})")
            
            # ETF基金：最新净值日期应该是最新净值的日期，而不是当日
            # 对于ETF基金，由于是实时交易，最新净值就是当前价格，日期应该是交易日
            fund_item = {
                "基金代码": fund_code,
                "基金名称": fund_name,
                "板块分类": "ETF基金",
                "最新净值": fund_info.get("dwjz", "N/A"),
                "估算净值": fund_info.get("gsz", "N/A"),
                "估算涨跌率": fund_info.get("gszzl", "N/A"),
                "净值日期": fund_info.get("jzrq", "N/A"),  # 最新净值日期
                "估值时间": fund_info.get("gztime", "N/A"),  # 估值时间
                "昨收盘价日期": fund_info.get("prev_trade_date", "N/A")  # 昨收盘价对应的日期
            }
            etf_fund_data.append(fund_item)
            # ETF基金不添加到自选基金列表中，因为它们是场内基金
    
    # 按原始顺序排序各组基金
    
    # 按板块分类排序各组基金
    chaochao_sorted = sort_funds_by_category(chaochao_fund_data)
    yaoyao_sorted = sort_funds_by_category(yaoyao_fund_data)
    overseas_sorted = sort_funds_by_category(overseas_fund_data)
    etf_sorted = sort_funds_by_category(etf_fund_data)
    
    # 合并所有基金数据到 'all' 键，确保没有重复
    all_funds = []
    all_codes = set()
    
    # 按优先级添加：自选基金 -> 境外基金 -> ETF基金
    for fund in chaochao_sorted + yaoyao_sorted:
        if fund['基金代码'] not in all_codes:
            all_funds.append(fund)
            all_codes.add(fund['基金代码'])
    
    for fund in overseas_sorted:
        if fund['基金代码'] not in all_codes:
            all_funds.append(fund)
            all_codes.add(fund['基金代码'])
    
    for fund in etf_sorted:
        if fund['基金代码'] not in all_codes:
            all_funds.append(fund)
            all_codes.add(fund['基金代码'])
    
    log_info(f"📊 合并后的总基金数量: {len(all_funds)} 只")
    log_info(f"🔍 包含的基金类型: 自选基金({len(chaochao_sorted + yaoyao_sorted)}), 境外基金({len(overseas_sorted)}), 场内基金({len(etf_sorted)})")
    
    return {
        "chaochao": chaochao_sorted,
        "yaoyao": yaoyao_sorted,
        "overseas": overseas_sorted,
        "etf": etf_sorted,
        "all": all_funds
    }

def get_monitor_funds(max_workers=10):
    """获取监控基金信息 - 优化版本"""
    tracker = OptimizedFundTracker(max_workers=max_workers)
    
    monitor_fund_codes = [
        #军工
        "010364[军工]",  # 鹏华军工
        "022243[军工]",  # 中邮军工混合
        "015790[军工]",  # 永赢高端装备智选混合发起C
        "015945[军工]",  # 易方达军工混合
        "013566[军工]",  # 华夏军工混合
        "015599[军工]",  # 国泰国证航天军工
        #医疗
        "006113[港股医药]",  # 汇添富创新药混合A
        "023482[港股医药]",  # 万家港股创新药
        "017633[医疗器械]",  # 汇添富医疗器械
        "024380[港股医药]",  # 平安港股通医疗混合
        "013416[医疗器械]",  # 永赢医疗器械
        "014565[港股医药]",  # 天弘沪深港创新药
        "020398[港股医药]",  # 中银沪港通创新药
        "000591[医疗器械]",  # 中银健康生活
        "001056[医疗器械]",  # 华银健康生活
        "016018[医疗器械]",  # 银河康乐股票C
        #银行
        "016573[金融]",  # 招商银行AH
        "021980[金融]",  # 兴全红利量化选股股票C
        "021457[港股金融]",  # 易方达红利低波A
        "018388[港股金融]",  # 华泰柏瑞港股通红利
        "019026[金融]",  # 易方达金融股票
        "006810[港股金融]",  # 泰康香港银行
        "004070[证券]",  # 南方中证全指证券公司ETF联接C
        "025194[证券]",  #银华中证全指证券公司ETF发起式联接C
        "012420[金融]",  # 广发价值领先混合C
        "024074[金融]",  # 上银国证自由现金流
        "001595[金融]",  # 天弘中证银行ETF联接C
        #通信
        "022365[通信]",  # 永赢智选混合
        "025422[通信]",  # 浦银安盛数字经济混合
        "018291[通信]",  # 广发新兴成长混合C
        "021717[云计算]",  # 招商云计算ETF
        "019170[云计算]",  # 天弘沪港深云计算
        "023881[科技]",  # 兴全商业模式混合(LOF)C
        "014819[科技]",  # 国金新兴价值混合
        "377240[科技]",  # 摩根新兴动力混合
        "024203[科技]",  # 永赢制造升级混合
        "016387[科技]",  # 永赢低碳环保智选混
        "021536[科技]",  # 天弘中证软件服务指数发起C
        "025784[科技]",  # 兴证资管金麒麟兴享优选混合E
        "018994[通信]",  # 中欧数字经济混合
        "021989[通信]",  # 银河中证通信
        "021934[通信]",  # 富国中证通信
        "020900[通信]",  # 天弘中证全指通信设备指数发起C
        "024195[通信]",  # 永赢国证商用卫星通信产业ETF发起联接C
        "025491[通信]",  # 平安中证卫星产业指数C
        "024749[通信]",  # 博时中证卫星产品指数C
        "009854[通信]",  # 中加优势企业混合C
        "014422[人工智能]",  # 弘毅消费混合
        "011840[人工智能]",  # 天弘中证人工智能C
        "012734[人工智能]",  # 易方达人工智能ETF联接C
        "023565[人工智能]",  # 易方达科创人工智能ETF联接C
        "024663[人工智能]",  # 富国创业板人工智能
        "012183[人工智能]",  # 广发沪港深精选混合
        "023754[人工智能]",  # 永赢信息产业智选混合
        #机器人
        "020256[机器人]",  # 中欧机器人
        "020973[机器人]",  # 易方达机器人
        "015401[机器人]",  # 弘毅甄选混合
        "018125[机器人]",  # 永赢制造混合
        "020608[机器人]",  # 南方中证机器人ETF发起联接C
        "001864[机器人]",  # 中海长江三角混合
        "020982[机器人]",  # 华安国证机器人
        #量化
        "014806[量化]",  # 国金量化混合
        "020902[量化]",  # 招商量化选股
        #新能源
        "017647[新能源]",  # 易方达光伏
        "018419[新能源]",  # 广发碳中和混合
        "011103[新能源]",  # 光天弘中证光伏
        "023852[新能源]",  # 富国上证科创板新能源ETF发起式联接C
        
        "015528[新能源汽车]",  # 弘毅汽车混合
        "017223[储能]",  # 富国中证电池主题ETF发起式联接C
        "013180[储能]",  # 广发国证新能源车电池ETF联接C
        "290014[储能]",  # 泰信现代服务业混合
        "018173[储能]",  # 华泰柏瑞中证电力全指ETF发起式联接C
        "025857[储能]",  # 华夏中证电网设备
        "024738[储能]",  # 永赢新材料智选混合发起C
        "023159[储能]",  # 上银先进制造混合发起式C
        #传统能源
        "016814[传统能源]",  # 国联中证煤炭
        "013275[传统能源]",  # 富国中证煤炭指数(LOF)C
        "023145[传统能源]",  # 汇添富中证油气资源ETF发起式联接C
        "021856[传统能源]",  # 博时中证油气资源ETF发起式联接C
        "015897[化工]",  # 天弘中证化工
        "022328[化工]",  # 宏利高端装备股票C
        "018647[家用电器]",  # 易方达家电龙头
        "011036[稀土]",  # 嘉实中证稀土
        "017193[有色金属]",  # 天弘中证工业有色金属主题指数发起C
        "019875[有色金属]",  # 广发中证稀有金属
        "016708[有色金属]",  # 华夏有色金属ETF联接C
        "017141[有色金属]",  # 华宝中证有色金属
        "023037[有色金属]",  # 中欧资源精选混合发起C
        "023449[有色金属]",  # 上银资源精选混合发起C
        "161715[有色金属]",  # 招商大宗商品(LOF)
        "012725[畜牧养殖]",  # 国泰畜牧养殖
        "010770[农业]",  # 天弘中证农业主题ETF联接C
        "020651[农业]",  # 博时国证粮食产业指数发起式C
        #黄金
        "021959[贵金属]",  # 南方沪深港黄金
        "020412[贵金属]",  # 永赢沪深港黄金
        "008987[贵金属]",  # 广发上海金ETF联接C
        "016582[贵金属]",  # 嘉实上海金ETF联接C
        "000217[贵金属]",  # 华安黄金ETF联接C
        "002963[贵金属]",  # 易方达黄金ETF联接C
        "025446[贵金属]",  # 万家周期视野股票C
        #消费
        "012341[消费]",  # 东财食品饮料指数
        "017870[消费]",  # 光大消费主题股票C
        "018650[消费]",  # 兴银消费混合
        "012414[消费]",  # 招商中证白酒指数
        "018897[消费电子]",  # 易方达消费电子ETF联接C
        "016008[消费电子]",  # 招商中证消费电子主题ETF联接C
        "012769[传媒]",  # 华夏中证动漫游戏ETF发起式联接C
        "001223[传媒]",  # 鹏华文化传媒娱乐股票
        "010677[传媒]",  # 工银瑞信中证传媒
        "004753[传媒]",  # 广发中证传媒
        "016550[传媒]",  # 永赢消费龙头智选混合发起C
        #半导体
        "019571[半导体]",  # 诺安配置混合
        "001665[半导体]",  # 平安鑫安混合
        "014855[半导体]",  # 嘉实中证半导体
        "025209[半导体]",  # 永赢先锋半导体智选混合发起C
        "023829[半导体]",  # 万家中证半导体材料设备主题
        "020671[半导体]",  # 易方达科创板芯片
        "020629[半导体]",  # 汇添富上证科创板芯片ETF联接C
        "011120[半导体]",  # 富国创新科技混合
        "013613[半导体]",  # 宝盈国家安全沪港深股票C
        "013446[半导体]",  # 东财芯片C
        "016874[半导体]",  # 广发远见智选混合
        "020640[半导体]",  # 广发半导体设备ETF联接C
        "007639[半导体]",  # 汇添富竞争优势灵活配置
        "020227[半导体]",  # 国泰中证全指集成电路ETF发起式联接C
        "009883[半导体]",  # 华润元大核心动力混合C
        #指数
        "022435[指数型]",  # 南方中证500
        "022486[指数型]",  # 国金中证A500
        "019919[指数型]",  # 招商中证2000
        "021172[指数型]",  # 华安北证50
        "017518[指数型]",  # 招商北证50
        "025444[指数型]",  # 景顺长城北证50
        "011613[指数型]",  # 华夏科创50
        "023051[指数型]",  # 交银科创100
        "023896[指数型]",  # 天弘科创综合
        "004744[指数型]",  # 易方达创业板ETF联接C
        "025165[指数型]",  # 易方达创业板增强C
        #基建
        "004857[建筑材料]",  # 广发建筑材料
        #港股通科技
        "015740[港股科技]",  # 国泰港股通科技
        "013309[港股科技]",  # 易方达恒生科技
        "012349[港股科技]",  # 天弘恒生科技
        "021378[港股科技]",  # 兴业港股通互联网
        "013172[港股科技]",  # 华夏恒生互联网
        #灵活混合
        "002834[混合型]",  # 华夏新锦绣混合C
        "002833[混合型]",  # 华夏锦绣混合
        "019374[混合型]",  # 广发睿杰精选混合发起式A
        #债基
        "003547[债券型]",  # 鹏华丰禄债券
        "018598[债券型]",  # 兴全招益债券
    ]
    
    # 提取纯基金代码，去掉自定义板块标签
    monitor_pure_codes = [extract_pure_fund_code(code) for code in monitor_fund_codes]
    
    log_info("=== 监控基金信息 ===")
    log_info(f"🔍 并发请求: {max_workers} 线程")
    
    start_time = time.time()
    fund_data_raw = tracker.get_funds_realtime(monitor_pure_codes)
    end_time = time.time()
    
    log_info(f"⏱️  耗时: {end_time - start_time:.1f}s, 成功: {len(fund_data_raw)}/{len(monitor_pure_codes)} 只")
    
    # 转换数据格式
    monitor_fund_data = []
    for fund_info in fund_data_raw:
        if fund_info:
            fund_code = fund_info.get('fundcode', 'N/A')
            fund_name = fund_info.get('name', 'N/A')
            
            # 查找原始基金代码（带自定义标签）
            original_fund_code = ""
            for code in monitor_fund_codes:
                if extract_pure_fund_code(code) == fund_code:
                    original_fund_code = code
                    break
            
            category = tracker.classifier.classify_fund(fund_name, original_fund_code)
            
            # 检查估算净值日期是否与当前交易日一致
            # 如果日期不一致（如港股闭市），使用最新净值作为估算净值，涨跌幅设为0
            gztime = fund_info.get("gztime", "")
            today = datetime.now().strftime('%Y-%m-%d')
            dwjz = fund_info.get("dwjz", "N/A")
            gsz = fund_info.get("gsz", "N/A")
            gszzl = fund_info.get("gszzl", "N/A")
            
            # 提取估算净值日期（从 gztime 中提取日期部分）
            estimate_date = None
            if gztime and gztime != "N/A":
                try:
                    # gztime 格式可能是 "YYYY-MM-DD HH:mm" 或 "YYYY-MM-DD"
                    if ' ' in gztime:
                        estimate_date = gztime.split(' ')[0]
                    else:
                        estimate_date = gztime
                except:
                    pass
            
            # 如果估算净值日期不是今天，且最新净值有效，则使用最新净值
            if estimate_date and estimate_date != today:
                if dwjz != "N/A" and dwjz != "" and dwjz != "NaN" and dwjz != "null":
                    try:
                        # 验证最新净值是有效数字
                        float(dwjz)
                        gsz = dwjz  # 使用最新净值作为估算净值
                        gszzl = "0.00"  # 涨跌幅设为0
                        # 保持 gztime 为当前日期（用于列头显示），但实际净值是上日的
                        gztime = f"{today} 15:00"
                        log_debug(f"📅 {fund_name} ({fund_code}) 估算净值日期({estimate_date})非今日，使用最新净值({dwjz})，涨跌幅设为0")
                    except (ValueError, TypeError):
                        pass  # 如果最新净值无效，保持原值
            
            fund_item = {
                "基金代码": fund_code,
                "基金名称": fund_name,
                "板块分类": category,
                "最新净值": dwjz,
                "估算净值": gsz,
                "估算涨跌率": gszzl,
                "净值日期": fund_info.get("jzrq", "N/A"),
                "估值时间": gztime
            }
            monitor_fund_data.append(fund_item)
    
    # 按板块分类排序基金
    sorted_monitor_funds = sort_funds_by_category(monitor_fund_data)
    
    return sorted_monitor_funds





def save_to_excel(fund_data_dict, monitor_funds=None, filename=None):
    """保存数据到Excel文件，包含多个sheet页"""
    if not filename:
        filename = f"我的基金_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    if not fund_data_dict:
        log_info("\n没有数据可保存")
        return None
    
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # 保存监控基金数据（放在第一位）
            if monitor_funds:
                monitor_df = pd.DataFrame(monitor_funds)
                monitor_df.to_excel(writer, sheet_name='监控基金', index=False)
                log_info(f"✓ 监控基金数据已保存到sheet页（第一位）")
            
            # 保存钞钞的基金数据（按原始定义顺序）
            if fund_data_dict.get('chaochao'):
                chaochao_df = pd.DataFrame(fund_data_dict['chaochao'])
                chaochao_df.to_excel(writer, sheet_name='钞钞的基金', index=False)
                log_info(f"✓ 钞钞的基金数据已保存到sheet页（按原始定义顺序）")
            
            # 保存垚垚的基金数据（按原始定义顺序）
            if fund_data_dict.get('yaoyao'):
                yaoyao_df = pd.DataFrame(fund_data_dict['yaoyao'])
                yaoyao_df.to_excel(writer, sheet_name='垚垚的基金', index=False)
                log_info(f"✓ 垚垚的基金数据已保存到sheet页（按原始定义顺序）")
            
            # 保存境外基金数据（按原始定义顺序）
            if fund_data_dict.get('overseas'):
                overseas_df = pd.DataFrame(fund_data_dict['overseas'])
                overseas_df.to_excel(writer, sheet_name='境外基金', index=False)
                log_info(f"✓ 境外基金数据已保存到sheet页（按原始定义顺序）")
        
            # 保存合并后的基金数据（包含持仓信息）
            if fund_data_dict.get('all'):
                all_df = pd.DataFrame(fund_data_dict['all'])
                all_df.to_excel(writer, sheet_name='全部基金', index=False)
                log_info(f"✓ 全部基金数据（含持仓信息）已保存到sheet页")
        
        log_info(f"💾 Excel: {filename}")
        return filename
    except Exception as e:
        log_info(f"❌ Excel保存失败: {e}")
        return None

def save_to_html_multi_sheet(fund_data_dict, monitor_funds=None, filename=None, profit_results=None):
    """保存数据到HTML文件，支持多sheet页显示"""
    if not filename:
        filename = f"我的基金_{datetime.now().strftime('%Y%m%d')}.html"
    
    if not fund_data_dict:
        print("\n没有数据可保存")
        return None
    
    # HTML模板
    html_template = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>我的基金数据</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        body {{
            font-family: 'Microsoft YaHei', Arial, sans-serif;
            margin: 20px;
            background-color: #f5f5f5;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }}
        h1 {{
            color: #333;
            text-align: center;
            margin-bottom: 30px;
        }}
        .summary {{
            background-color: #f8f9fa;
            padding: 15px;
            border-radius: 5px;
            margin-bottom: 20px;
            border-left: 4px solid #007bff;
        }}
        .tab-container {{
            margin-bottom: 20px;
        }}
        .tab-buttons {{
            display: flex;
            border-bottom: 2px solid #007bff;
            margin-bottom: 20px;
        }}
        .tab-button {{
            background-color: #f8f9fa;
            border: none;
            padding: 12px 24px;
            cursor: pointer;
            font-size: 16px;
            font-weight: bold;
            color: #495057;
            border-radius: 5px 5px 0 0;
            margin-right: 5px;
            transition: all 0.3s;
        }}
        .tab-button:hover {{
            background-color: #e9ecef;
        }}
        .tab-button.active {{
            background-color: #007bff;
            color: white;
        }}
        .tab-content {{
            display: none;
        }}
        .tab-content.active {{
            display: block;
        }}
        .search-box {{
            margin: 15px 0;
            padding: 10px;
            background-color: #f8f9fa;
            border-radius: 8px;
            border: 1px solid #dee2e6;
        }}
        .search-box input {{
            width: 100%;
            padding: 10px 15px;
            font-size: 14px;
            border: 1px solid #ced4da;
            border-radius: 6px;
            box-sizing: border-box;
            transition: border-color 0.3s;
        }}
        .search-box input:focus {{
            outline: none;
            border-color: #007bff;
            box-shadow: 0 0 0 3px rgba(0, 123, 255, 0.1);
        }}
        .search-box input::placeholder {{
            color: #6c757d;
        }}
        .fund-row.hidden {{
            display: none;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
            background-color: white;
        }}
        th {{
            background-color: #007bff;
            color: white;
            padding: 12px 8px;
            text-align: center;
            font-weight: bold;
            position: relative;
        }}
        th.sortable {{
            cursor: pointer;
            user-select: none;
        }}
        th.sortable:hover {{
            background-color: #0056b3;
        }}
        .sort-icons {{
            display: inline-flex;
            flex-direction: column;
            margin-left: 5px;
            vertical-align: middle;
            line-height: 1;
        }}
        .sort-icon {{
            display: inline-block;
            font-size: 8px;
            cursor: pointer;
            opacity: 0.5;
            transition: opacity 0.3s, color 0.3s;
            color: rgba(255, 255, 255, 0.5);
            user-select: none;
        }}
        .sort-icon:hover {{
            opacity: 1;
            color: white;
        }}
        .sort-icon.asc-active {{
            opacity: 1;
            color: white;
        }}
        .sort-icon.desc-active {{
            opacity: 1;
            color: white;
        }}
        .sort-icon-asc {{
            margin-bottom: 1px;
        }}
        .sort-icon-desc {{
            margin-top: 1px;
        }}
        td {{
            padding: 10px 8px;
            text-align: center;
            border-bottom: 1px solid #ddd;
        }}
        tr:nth-child(even) {{
            background-color: #f8f9fa;
        }}
        tr:hover {{
            background-color: #e9ecef;
        }}
        .fund-row {{
            cursor: pointer;
        }}
        .fund-row:hover {{
            background-color: #e3f2fd !important;
        }}
        .positive {{
            color: #dc3545;
            font-weight: bold;
        }}
        .negative {{
            color: #28a745;
            font-weight: bold;
        }}
        .neutral {{
            color: #6c757d;
        }}
        .profit-positive {{
            color: #dc3545;
            font-weight: bold;
        }}
        .profit-negative {{
            color: #28a745;
            font-weight: bold;
        }}
        .profit-neutral {{
            color: #6c757d;
        }}
        .category-tag {{
            padding: 4px 8px;
            border-radius: 12px;
            font-size: 12px;
            color: white;
            font-weight: bold;
        }}
        /* 分类颜色优化：高对比、可读性强 */
        /* 核心科技板块 - 蓝色系 */
        .category-科技 {{
            background-color: #1565c0;
        }}
        .category-半导体 {{
            background-color: #0d47a1;
        }}
        .category-计算机 {{
            background-color: #1976d2;
        }}
        .category-电子 {{
            background-color: #2196f3;
        }}
        .category-通信 {{
            background-color: #42a5f5;
        }}
        .category-人工智能 {{
            background-color: #1e88e5;
        }}
        .category-机器人 {{
            background-color: #455a64;
        }}
        
        /* 新兴产业 - 绿色系 */
        .category-新能源 {{
            background-color: #2e7d32;
        }}
        .category-光伏 {{
            background-color: #388e3c;
        }}
        .category-风电 {{
            background-color: #4caf50;
        }}
        .category-储能 {{
            background-color: #66bb6a;
        }}
        .category-新能源汽车 {{
            background-color: #81c784;
        }}
        .category-消费电子 {{
            background-color: #a5d6a7;
            color: #2e7d32;
        }}
        
        /* 传统优势板块 - 红色系 */
        .category-军工 {{
            background-color: #c62828;
        }}
        .category-医药 {{
            background-color: #d32f2f;
        }}
        .category-消费 {{
            background-color: #ef6c00;
        }}
        .category-食品饮料 {{
            background-color: #f57c00;
        }}
        .category-家电 {{
            background-color: #ff8f00;
            color: #fff;
        }}
        .category-汽车 {{
            background-color: #e65100;
        }}
        
        /* 金融地产 - 紫色系 */
        .category-金融 {{
            background-color: #6a1b9a;
        }}
        .category-地产 {{
            background-color: #ad1457;
        }}
        .category-建筑装饰 {{
            background-color: #c2185b;
        }}
        .category-建筑材料 {{
            background-color: #d81b60;
        }}
        
        /* 周期板块 - 棕色系 */
        .category-化工 {{
            background-color: #8d6e63;
        }}
        .category-钢铁 {{
            background-color: #6d4c41;
        }}
        .category-煤炭 {{
            background-color: #5d4037;
        }}
        .category-电力 {{background-
            background-color: #4e342e;
        }}
        .category-机械设备 {{
            background-color: #3e2723;
        }}
        .category-电气设备 {{
            background-color: #bf360c;
        }}
        
        /* 其他板块 - 多样化颜色 */
        .category-农业 {{
            background-color: #00897b;
        }}
        .category-港股 {{
            background-color: #006064;
        }}
        /* 港股细分板块 - 与对应板块颜色保持一致 */
        .category-港股科技 {{
            background-color: #1565c0;  /* 与科技板块一致 */
        }}
        .category-港股金融 {{
            background-color: #6a1b9a;  /* 与金融板块一致 */
        }}
        .category-港股医药 {{
            background-color: #d32f2f;  /* 与医药板块一致 */
        }}
        
        /* 金属细分板块 - 金色系渐变 */
        .category-贵金属 {{
            background-color: #b58900;  /* 金色 */
            color: #fff;
        }}
        .category-有色金属 {{
            background-color: #cd7f32;  /* 铜色 */
            color: #fff;
        }}
        .category-稀土 {{
            background-color: #b8860b;  /* 暗金色 */
            color: #fff;
        }}
        
        /* 新增自定义板块样式 */
        .category-传统能源 {{
            background-color: #5d4037;  /* 深棕色 */
            color: #fff;
        }}
        .category-食品饮料 {{
            background-color: #f57c00;  /* 橙色 */
            color: #fff;
        }}
        .category-医疗器械 {{
            background-color: #e91e63;  /* 粉红色 */
            color: #fff;
        }}
        .category-混合型 {{
            background-color: #546e7a;  /* 蓝灰色 */
            color: #fff;
        }}
        .category-债券型 {{
            background-color: #1b5e20;  /* 深绿色 */
            color: #fff;
        }}
        .category-指数型 {{
            background-color: #00695c;  /* 深青色 */
            color: #fff;
        }}
        .category-量化 {{
            background-color: #37474f;  /* 深灰色 */
            color: #fff;
        }}
        .category-通信 {{
            background-color: #42a5f5;  /* 浅蓝色 */
            color: #fff;
        }}
        .category-家用电器 {{
            background-color: #ff8f00;  /* 深橙色 */
            color: #fff;
        }}
        .category-畜牧养殖 {{
            background-color: #8bc34a;  /* 浅绿色 */
            color: #2e7d32;
        }}
        .category-农业 {{
            background-color: #8bc34a;  /* 浅绿色 */
            color: #2e7d32;
        }}
        .category-证券 {{
            background-color: #9c27b0;  /* 紫色 */
            color: #fff;
        }}
        .category-云计算 {{
            background-color: #03a9f4;  /* 天蓝色 */
            color: #fff;
        }}
        .category-新能源汽车 {{
            background-color: #81c784;  /* 浅绿色 */
            color: #2e7d32;
        }}
        .category-基建 {{
            background-color: #795548;
            color: #fff;
        }}
        .category-传媒 {{
            background-color: #5d4037;
        }}
        .category-环保 {{
            background-color: #2e7d32;
        }}
        .category-教育 {{
            background-color: #283593;
        }}
        .category-物流 {{
            background-color: #37474f;
        }}
        .category-纺织服装 {{
            background-color: #6a1b9a;
        }}
        .category-轻工制造 {{
            background-color: #8e24aa;
        }}
        .category-公用事业 {{
            background-color: #7b1fa2;
        }}
        .category-交通运输 {{
            background-color: #9c27b0;
        }}
        .category-商业贸易 {{
            background-color: #ba68c8;
            color: #4a148c;
        }}
        .category-休闲服务 {{
            background-color: #ce93d8;
            color: #4a148c;
        }}
        .category-综合 {{
            background-color: #e1bee7;
            color: #4a148c;
        }}
        
        /* 基金类型分类 - 灰色系 */
        .category-混合型 {{
            background-color: #546e7a;
        }}
        .category-股票型 {{
            background-color: #b71c1c;
        }}
        .category-债券型 {{
            background-color: #1b5e20;
        }}
        .category-货币型 {{
            background-color: #795548;
            color: #fff;
        }}
        .category-指数型 {{
            background-color: #00695c;
        }}
        .category-ETF基金 {{
            background-color: #d32f2f;  /* 深红色，更明显 */
            color: #ffffff;
            font-weight: bold;
        }}
        .category-LOF基金 {{
            background-color: #e65100;  /* 橙红色 */
            color: #ffffff;
            font-weight: bold;
        }}
        .category-ETF联接 {{
            background-color: #6a1b9a;  /* 深紫色 */
            color: #ffffff;
            font-weight: bold;
        }}
        
        /* 特殊基金类型 */
        .category-境外基金 {{
            background-color: #4a148c;
        }}
        .category-场内基金 {{
            background-color: #0d47a1;  /* 深蓝色 */
            color: #ffffff;
        }}
        .category-其他 {{
            background-color: #6c757d;
        }}
        .category-未知 {{
            background-color: #6c757d;
        }}
        .timestamp {{
            text-align: center;
            color: #6c757d;
            font-size: 14px;
            margin-bottom: 20px;
        }}
        .section-header {{
            background-color: #e3f2fd;
            padding: 10px;
            margin: 10px 0;
            border-radius: 5px;
            font-weight: bold;
            color: #1976d2;
        }}
        .modal {{
            display: none;
            position: fixed;
            z-index: 1000;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(0,0,0,0.4);
        }}
        .modal-content {{
            background-color: #ffffff;
            color: #333333;
            margin: 2% auto;
            padding: 20px 16px;
            border: 1px solid #ddd;
            width: 92%;
            max-width: 1000px;
            border-radius: 12px;
            position: relative;
            box-shadow: 0 4px 20px rgba(0,0,0,0.3);
            max-height: 92vh;
            overflow-y: auto;
            overflow-x: hidden;
            box-sizing: border-box;
        }}
        .close {{
            color: #aaa;
            float: right;
            font-size: 28px;
            font-weight: bold;
            cursor: pointer;
            position: absolute;
            right: 20px;
            top: 20px;
            transition: color 0.3s;
        }}
        .close:hover {{
            color: #000000;
        }}
        .chart-container {{
            position: relative;
            height: 450px;
            margin: 20px 0 4px;
            width: 100%;
            background-color: #f8f9fa;
            border-radius: 8px;
            padding: 12px;
            border: 1px solid #e9ecef;
            overflow: hidden;
            box-sizing: border-box;
        }}
        #navChart {{ width: 100%; }}
        .nav-data {{
            display: none;
        }}
        .cost-line {{
            border-top: 2px dashed #ff6b6b;
            margin: 10px 0;
            padding-top: 10px;
        }}
        .cost-line strong {{
            color: #ff6b6b;
        }}
        .modal-title {{
            color: #333333;
            margin-bottom: 20px;
            text-align: center;
            font-size: 24px;
        }}
        .fund-info-card {{
            background: rgba(255,255,255,0.95);
            color: #333;
            padding: 12px 16px;
            border-radius: 12px;
            margin: 12px 0;
            border: 1px solid rgba(0,123,255,0.3);
            box-shadow: 0 4px 20px rgba(0,123,255,0.15);
        }}
        .info-row {{
            margin: 8px 0;
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 4px 0;
            gap: 15px;
        }}
        .info-item {{
            flex: 1;
            text-align: left;
            min-width: 0;
        }}
        .info-label {{
            color: #666;
            font-size: 14px;
            font-weight: 500;
            display: inline-block;
            min-width: 70px;
        }}
        .info-value {{
            color: #333;
            font-weight: bold;
            font-size: 14px;
        }}
        .nav-value {{
            color: #007bff;
            font-size: 14px;
            font-weight: bold;
        }}
        .cost-value {{
            color: #ff6b6b;
            font-size: 14px;
            font-weight: bold;
        }}
        .info-value.positive {{
            color: #dc3545;
        }}
        .info-value.negative {{
            color: #28a745;
        }}
        .info-value.neutral {{
            color: #6c757d;
        }}
        .range-buttons {{
            display: flex;
            justify-content: center;
            gap: 8px;
            margin-bottom: 10px;
        }}
        .header-row {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 10px;
            flex-wrap: wrap;
            margin-bottom: 10px;
        }}
        .range-section {{
            display: flex;
            align-items: center;
            gap: 8px;
        }}
        .label-strong {{
            font-weight: 700;
        }}
        .range-btn {{
            border: 1px solid #007bff;
            background: #fff;
            color: #007bff;
            padding: 6px 12px;
            border-radius: 16px;
            cursor: pointer;
            font-size: 14px;
        }}
        .range-btn.active {{
            background: #007bff;
            color: #fff;
        }}
        .legend {{
            display: flex;
            align-items: center;
            color: #555;
            gap: 8px;
            flex-wrap: wrap;
        }}
        .legend-item {{
            display: inline-flex;
            align-items: center;
            margin-left: 6px;
            font-size: 14px;
        }}
        .legend-line {{
            display: inline-block;
            width: 22px;
            height: 0;
            border-top: 3px solid #007bff;
            margin-right: 6px;
        }}
        .legend-line.price {{
            border-top: 3px solid #007bff;
        }}
        .legend-line.cost {{
            border-top: 2px dashed #ff6b6b;
        }}
        .tooltip {{
            position: fixed; /* 改为固定定位，避免被容器裁剪 */
            background: rgba(0,0,0,0.85);
            color: #fff;
            padding: 8px 10px;
            border-radius: 6px;
            font-size: 14px;
            pointer-events: none;
            transform: translate(-50%, -120%);
            white-space: nowrap;
            z-index: 9999;
        }}
    </style>
</head>
<body>
    <div class="container">
        <h1>我的基金数据</h1>
        <div class="timestamp">生成时间: {timestamp}</div>
        
        <!-- 弹框 -->
        <div id="fundModal" class="modal">
            <div class="modal-content">
                <span class="close">&times;</span>
                <h2 class="modal-title" id="modalTitle">基金详情</h2>
                <div class="header-row">
                    <div class="range-section">
                        <span class="label-strong">选择时间</span>
                        <div class="range-buttons">
                            <button class="range-btn active" data-days="30">1月</button>
                            <button class="range-btn" data-days="90">3月</button>
                            <button class="range-btn" data-days="180">6月</button>
                            <button class="range-btn" data-days="365">1年</button>
                        </div>
                    </div>
                    <div class="legend">
                        <span class="legend-item"><span class="legend-line price"></span>净价</span>
                        <span class="legend-item"><span class="legend-line cost"></span>成本单价</span>
                    </div>
                </div>
                
                <!-- 信息显示区域：日期选择按钮下方，图表区域上方 -->
                <div class="fund-info-card">
                    <div class="info-row">
                        <div class="info-item">
                            <span class="info-label">交易日期:</span>
                            <span class="info-value" id="transactionDate">-</span>
                        </div>
                        <div class="info-item">
                            <span class="info-label">单位净值:</span>
                            <span class="nav-value" id="unitNav">-</span>
                        </div>
                        <div class="info-item">
                            <span class="info-label">涨跌幅:</span>
                            <span class="info-value" id="changeRate">-</span>
                        </div>
                    </div>
                    <div class="info-row">
                        <div class="info-item">
                            <span class="info-label">份额变动:</span>
                            <span class="info-value" id="sharesChange">-</span>
                        </div>
                        <div class="info-item">
                            <span class="info-label">变动金额:</span>
                            <span class="info-value" id="changeAmount">-</span>
                        </div>
                        <div class="info-item">
                            <span class="info-label">持仓成本:</span>
                            <span class="cost-value" id="holdingCost">-</span>
                        </div>
                    </div>
                </div>
                <div class="chart-container">
                    <div id="navChart"></div>
                </div>
                <!-- 底部成本价条已移除，改由悬浮展示 -->
            </div>
        </div>
        
        <div class="summary">
            <h3>汇总信息</h3>
            <p>监控基金: <strong>{monitor_count}</strong> 只, 钞钞的基金: <strong>{chaochao_count}</strong> 只, 垚垚的基金: <strong>{yaoyao_count}</strong> 只, 境外基金: <strong>{overseas_count}</strong> 只</p>
            <p>监控基金平均涨跌率: <strong class="{monitor_avg_class}">{avg_change}</strong></p>
            <p>监控基金上涨: <span class="positive">{up_count}</span> 只, 下跌: <span class="negative">{down_count}</span> 只, 平盘: <span class="neutral">{flat_count}</span> 只</p>
            <p>钞钞平均涨跌率: <strong class="{chaochao_avg_class}">{chaochao_avg_change}</strong>，上涨: <span class="positive">{chaochao_up_count}</span> 只, 下跌: <span class="negative">{chaochao_down_count}</span> 只, 平盘: <span class="neutral">{chaochao_flat_count}</span> 只</p>
            <p>垚垚平均涨跌率: <strong class="{yaoyao_avg_class}">{yaoyao_avg_change}</strong>，上涨: <span class="positive">{yaoyao_up_count}</span> 只, 下跌: <span class="negative">{yaoyao_down_count}</span> 只, 平盘: <span class="neutral">{yaoyao_flat_count}</span> 只</p>
            <p>钞钞当日预估收益: <strong class="{chaochao_profit_class}">{chaochao_today_profit}</strong></p>
            <p>垚垚当日预估收益: <strong class="{yaoyao_profit_class}">{yaoyao_today_profit}</strong></p>
        </div>
        
        <div class="tab-container">
            <div class="tab-buttons">
                <button class="tab-button active" onclick="showTab('monitor')">监控基金</button>
                <button class="tab-button" onclick="showTab('chaochao')">钞钞的基金</button>
                <button class="tab-button" onclick="showTab('yaoyao')">垚垚的基金</button>
                <button class="tab-button" onclick="showTab('overseas')">境外基金</button>
                <button class="tab-button" onclick="showTab('etf')">场内基金</button>
            </div>
            
            <div id="monitor" class="tab-content active">
                <h3>监控基金</h3>
                <div class="search-box">
                    <input type="text" id="search-monitor" placeholder="搜索基金代码、基金名称或板块..." oninput="filterFunds('monitor', this.value)">
                </div>
                <table>
                    <thead>
                        <tr>
                            <th>基金代码</th>
                            <th>基金名称</th>
                            <th>板块分类</th>
                            <th>最新净值<br><span style="font-size: 14px; color: #ffffff; background-color: #007bff; padding: 2px 6px; border-radius: 4px; display: inline-block; margin-top: 4px;">{latest_date_header}</span></th>
                            <th>估算净值<br><span style="font-size: 12px; color: #ffffff; background-color: #007bff; padding: 2px 6px; border-radius: 4px; display: inline-block; margin-top: 4px;">{estimate_date_header}</span></th>
                            <th class="sortable" onclick="if(!event.target.classList.contains('sort-icon')) sortTableByHeader('monitor', 5, this)">估算涨跌率<span class="sort-icons"><span class="sort-icon sort-icon-asc" onclick="event.stopPropagation(); sortTableByIcon('monitor', 5, 'asc', this);">▲</span><span class="sort-icon sort-icon-desc" onclick="event.stopPropagation(); sortTableByIcon('monitor', 5, 'desc', this);">▼</span></span></th>
                        </tr>
                    </thead>
                    <tbody>
                        {monitor_table_rows}
                    </tbody>
                </table>
            </div>
            
            <div id="chaochao" class="tab-content">
                <h3>钞钞的基金</h3>
                <div class="search-box">
                    <input type="text" id="search-chaochao" placeholder="搜索基金代码、基金名称或板块..." oninput="filterFunds('chaochao', this.value)">
                </div>
                <table>
                    <thead>
                        <tr>
                            <th>基金代码</th>
                            <th>基金名称</th>
                            <th>板块分类</th>
                            <th>最新净值<br><span style="font-size: 14px; color: #ffffff; background-color: #007bff; padding: 2px 6px; border-radius: 4px; display: inline-block; margin-top: 4px;">{latest_date_header}</span></th>
                            <th>估算净值<br><span style="font-size: 14px; color: #ffffff; background-color: #007bff; padding: 2px 6px; border-radius: 4px; display: inline-block; margin-top: 4px;">{estimate_date_header}</span></th>
                            <th class="sortable" onclick="if(!event.target.classList.contains('sort-icon')) sortTableByHeader('chaochao', 5, this)">估算涨跌率<span class="sort-icons"><span class="sort-icon sort-icon-asc" onclick="event.stopPropagation(); sortTableByIcon('chaochao', 5, 'asc', this);">▲</span><span class="sort-icon sort-icon-desc" onclick="event.stopPropagation(); sortTableByIcon('chaochao', 5, 'desc', this);">▼</span></span></th>
                            <th>成本单价</th>
                            <th class="sortable" onclick="if(!event.target.classList.contains('sort-icon')) sortTableByHeader('chaochao', 7, this)">当日收益<span class="sort-icons"><span class="sort-icon sort-icon-asc" onclick="event.stopPropagation(); sortTableByIcon('chaochao', 7, 'asc', this);">▲</span><span class="sort-icon sort-icon-desc" onclick="event.stopPropagation(); sortTableByIcon('chaochao', 7, 'desc', this);">▼</span></span></th>
                            <th class="sortable" onclick="if(!event.target.classList.contains('sort-icon')) sortTableByHeader('chaochao', 8, this)">持仓收益<span class="sort-icons"><span class="sort-icon sort-icon-asc" onclick="event.stopPropagation(); sortTableByIcon('chaochao', 8, 'asc', this);">▲</span><span class="sort-icon sort-icon-desc" onclick="event.stopPropagation(); sortTableByIcon('chaochao', 8, 'desc', this);">▼</span></span></th>
                            <th class="sortable" onclick="if(!event.target.classList.contains('sort-icon')) sortTableByHeader('chaochao', 9, this)">持仓收益率<span class="sort-icons"><span class="sort-icon sort-icon-asc" onclick="event.stopPropagation(); sortTableByIcon('chaochao', 9, 'asc', this);">▲</span><span class="sort-icon sort-icon-desc" onclick="event.stopPropagation(); sortTableByIcon('chaochao', 9, 'desc', this);">▼</span></span></th>
                        </tr>
                    </thead>
                    <tbody>
                        {chaochao_table_rows}
                    </tbody>
                </table>
            </div>
            
            <div id="yaoyao" class="tab-content">
                <h3>垚垚的基金</h3>
                <div class="search-box">
                    <input type="text" id="search-yaoyao" placeholder="搜索基金代码、基金名称或板块..." oninput="filterFunds('yaoyao', this.value)">
                </div>
                <table>
                    <thead>
                        <tr>
                            <th>基金代码</th>
                            <th>基金名称</th>
                            <th>板块分类</th>
                            <th>最新净值<br><span style="font-size: 14px; color: #ffffff; background-color: #007bff; padding: 2px 6px; border-radius: 4px; display: inline-block; margin-top: 4px;">{latest_date_header}</span></th>
                            <th>估算净值<br><span style="font-size: 14px; color: #ffffff; background-color: #007bff; padding: 2px 6px; border-radius: 4px; display: inline-block; margin-top: 4px;">{estimate_date_header}</span></th>
                            <th class="sortable" onclick="if(!event.target.classList.contains('sort-icon')) sortTableByHeader('yaoyao', 5, this)">估算涨跌率<span class="sort-icons"><span class="sort-icon sort-icon-asc" onclick="event.stopPropagation(); sortTableByIcon('yaoyao', 5, 'asc', this);">▲</span><span class="sort-icon sort-icon-desc" onclick="event.stopPropagation(); sortTableByIcon('yaoyao', 5, 'desc', this);">▼</span></span></th>
                            <th>成本单价</th>
                            <th class="sortable" onclick="if(!event.target.classList.contains('sort-icon')) sortTableByHeader('yaoyao', 7, this)">当日收益<span class="sort-icons"><span class="sort-icon sort-icon-asc" onclick="event.stopPropagation(); sortTableByIcon('yaoyao', 7, 'asc', this);">▲</span><span class="sort-icon sort-icon-desc" onclick="event.stopPropagation(); sortTableByIcon('yaoyao', 7, 'desc', this);">▼</span></span></th>
                            <th class="sortable" onclick="if(!event.target.classList.contains('sort-icon')) sortTableByHeader('yaoyao', 8, this)">持仓收益<span class="sort-icons"><span class="sort-icon sort-icon-asc" onclick="event.stopPropagation(); sortTableByIcon('yaoyao', 8, 'asc', this);">▲</span><span class="sort-icon sort-icon-desc" onclick="event.stopPropagation(); sortTableByIcon('yaoyao', 8, 'desc', this);">▼</span></span></th>
                            <th class="sortable" onclick="if(!event.target.classList.contains('sort-icon')) sortTableByHeader('yaoyao', 9, this)">持仓收益率<span class="sort-icons"><span class="sort-icon sort-icon-asc" onclick="event.stopPropagation(); sortTableByIcon('yaoyao', 9, 'asc', this);">▲</span><span class="sort-icon sort-icon-desc" onclick="event.stopPropagation(); sortTableByIcon('yaoyao', 9, 'desc', this);">▼</span></span></th>
                        </tr>
                    </thead>
                    <tbody>
                        {yaoyao_table_rows}
                    </tbody>
                </table>
            </div>
            
            <div id="overseas" class="tab-content">
                <h3>境外基金</h3>
                <div class="search-box">
                    <input type="text" id="search-overseas" placeholder="搜索基金代码、基金名称或板块..." oninput="filterFunds('overseas', this.value)">
                </div>
                <table>
                    <thead>
                        <tr>
                            <th>基金代码</th>
                            <th>基金名称</th>
                            <th>板块分类</th>
                            <th>最新净值</th>
                            <th>上日净值</th>
                            <th class="sortable" onclick="if(!event.target.classList.contains('sort-icon')) sortTableByHeader('overseas', 5, this)">涨跌幅<span class="sort-icons"><span class="sort-icon sort-icon-asc" onclick="event.stopPropagation(); sortTableByIcon('overseas', 5, 'asc', this);">▲</span><span class="sort-icon sort-icon-desc" onclick="event.stopPropagation(); sortTableByIcon('overseas', 5, 'desc', this);">▼</span></span></th>
                        </tr>
                    </thead>
                    <tbody>
                        {overseas_table_rows}
                    </tbody>
                </table>
            </div>
            
            <div id="etf" class="tab-content">
                <h3>场内基金</h3>
                <div class="search-box">
                    <input type="text" id="search-etf" placeholder="搜索基金代码、基金名称或板块..." oninput="filterFunds('etf', this.value)">
                </div>
                <table>
                    <thead>
                        <tr>
                            <th>基金代码</th>
                            <th>基金名称</th>
                            <th>板块分类</th>
                            <th>最新净价<br><span style="font-size: 12px; color: #ffffff; background-color: #007bff; padding: 2px 6px; border-radius: 4px; display: inline-block; margin-top: 4px;">{etf_latest_time_header}</span></th>
                            <th>上日净价<br><span style="font-size: 14px; color: #ffffff; background-color: #007bff; padding: 2px 6px; border-radius: 4px; display: inline-block; margin-top: 4px;">{etf_prev_date_header}</span></th>
                            <th class="sortable" onclick="if(!event.target.classList.contains('sort-icon')) sortTableByHeader('etf', 5, this)">涨跌率<span class="sort-icons"><span class="sort-icon sort-icon-asc" onclick="event.stopPropagation(); sortTableByIcon('etf', 5, 'asc', this);">▲</span><span class="sort-icon sort-icon-desc" onclick="event.stopPropagation(); sortTableByIcon('etf', 5, 'desc', this);">▼</span></span></th>
                        </tr>
                    </thead>
                    <tbody>
                        {etf_table_rows}
                    </tbody>
                </table>
            </div>
        </div>
    </div>
    
    <script>
        // API 基址自动探测与可配置（支持 ?api= 覆盖与 localStorage 持久化，支持 HTTP/HTTPS 协议自动检测）
        (function() {{
            function resolveApiBase() {{
                try {{
                    var params = new URLSearchParams(window.location.search);
                    var apiFromQuery = params.get('api');
                    if (apiFromQuery) {{
                        try {{ localStorage.setItem('API_BASE', apiFromQuery); }} catch (e) {{}}
                        return apiFromQuery;
                    }}
                    try {{
                        var saved = localStorage.getItem('API_BASE');
                        if (saved) return saved;
                    }} catch (e) {{}}

                    // 默认策略：
                    // - 如果在 http(s) 环境且有同源后端，可用相对路径
                    // - 如果是 GitHub Pages（github.io），建议配置公开 API 地址
                    var origin = window.location.origin || '';
                    if (/^https?:\\/\\//i.test(origin)) {{
                        if (/github\\.io$/i.test(window.location.hostname)) {{
                            // 在 GitHub Pages 上，强烈建议通过 ?api= 或 localStorage 配置公开可访问的 HTTPS API 基址
                            // 临时回退为同源相对路径，若无后端会 404
                            return origin; 
                        }}
                        return origin;
                    }}
                }} catch (e) {{}}
                // 本地文件或未知协议下的回退
                // 根据当前页面协议决定使用 HTTP 还是 HTTPS
                var currentProtocol = window.location.protocol || 'http:';
                if (currentProtocol === 'https:') {{
                    return 'https://127.0.0.1:5000';
                }}
                return 'http://127.0.0.1:5000';
            }}
            window.API_BASE = (resolveApiBase() || '').replace(/\\/$/, '');
            
            // 协议检测和切换函数
            window.getApiUrlWithProtocolCheck = function(endpoint) {{
                var baseUrl = window.API_BASE;
                if (!baseUrl) {{
                    return null;
                }}
                // 如果 API_BASE 已经包含协议，直接使用
                if (/^https?:\\/\\//i.test(baseUrl)) {{
                    return baseUrl + endpoint;
                }}
                // 否则根据当前页面协议添加协议
                var protocol = window.location.protocol || 'http:';
                return protocol + '//' + baseUrl.replace(/^https?:\\/\\//i, '') + endpoint;
            }};
        }})();
        function showTab(tabName) {{
            // 隐藏所有tab内容
            var tabContents = document.getElementsByClassName('tab-content');
            for (var i = 0; i < tabContents.length; i++) {{
                tabContents[i].classList.remove('active');
            }}
            
            // 移除所有tab按钮的active类
            var tabButtons = document.getElementsByClassName('tab-button');
            for (var i = 0; i < tabButtons.length; i++) {{
                tabButtons[i].classList.remove('active');
            }}
            
            // 显示选中的tab内容
            document.getElementById(tabName).classList.add('active');
            
            // 添加active类到选中的按钮
            event.target.classList.add('active');
        }}
        
        // 搜索功能：模糊匹配基金代码、基金名称、板块，支持关键词筛选（红、绿、涨、跌）
        function filterFunds(tabName, searchText) {{
            var tabContent = document.getElementById(tabName);
            if (!tabContent) return;
            
            var table = tabContent.querySelector('table');
            if (!table) return;
            
            var rows = table.querySelectorAll('tbody tr.fund-row');
            var searchLower = searchText.toLowerCase().trim();
            
            if (searchLower === '') {{
                // 如果搜索框为空，显示所有行
                rows.forEach(function(row) {{
                    row.classList.remove('hidden');
                }});
                return;
            }}
            
            // 确定涨跌幅/率列的索引
            var changeRateColIndex = -1;
            if (tabName === 'monitor' || tabName === 'overseas' || tabName === 'etf') {{
                changeRateColIndex = 5; // 估算涨跌率/涨跌幅/涨跌率
            }} else if (tabName === 'chaochao' || tabName === 'yaoyao') {{
                changeRateColIndex = 5; // 估算涨跌率
            }}
            
            rows.forEach(function(row) {{
                // 获取基金代码（第一列）
                var codeCell = row.cells[0];
                var code = codeCell ? codeCell.textContent.trim().toLowerCase() : '';
                
                // 获取基金名称（第二列）
                var nameCell = row.cells[1];
                var name = nameCell ? nameCell.textContent.trim().toLowerCase() : '';
                
                // 获取板块分类（第三列）
                var categoryCell = row.cells[2];
                var category = '';
                if (categoryCell) {{
                    var categoryTag = categoryCell.querySelector('.category-tag');
                    category = categoryTag ? categoryTag.textContent.trim().toLowerCase() : '';
                }}
                
                // 检查关键词筛选（红、绿、涨、跌）
                var keywordMatch = true; // 默认为true，如果没有关键词则不影响
                var hasKeyword = false;
                var keywordText = '';
                
                if (changeRateColIndex >= 0) {{
                    var changeRateCell = row.cells[changeRateColIndex];
                    if (changeRateCell) {{
                        var changeRateText = changeRateCell.textContent.trim();
                        var changeRateValue = parseFloat(changeRateText.replace(/[+%]/g, ''));
                        
                        // 检查是否包含关键词
                        if (searchLower.includes('红') || searchLower.includes('涨')) {{
                            hasKeyword = true;
                            keywordMatch = !isNaN(changeRateValue) && changeRateValue > 0;
                            keywordText = searchLower.replace(/[红涨]/g, '').trim();
                        }} else if (searchLower.includes('绿') || searchLower.includes('跌')) {{
                            hasKeyword = true;
                            keywordMatch = !isNaN(changeRateValue) && changeRateValue < 0;
                            keywordText = searchLower.replace(/[绿跌]/g, '').trim();
                        }}
                    }}
                }}
                
                // 模糊匹配：检查搜索文本是否包含在基金代码、基金名称或板块中
                var searchTextForMatch = hasKeyword ? keywordText : searchLower;
                var textMatch = true;
                if (searchTextForMatch) {{
                    textMatch = code.includes(searchTextForMatch) || 
                               name.includes(searchTextForMatch) || 
                               category.includes(searchTextForMatch);
                }}
                
                // 如果有关键词，需要同时满足关键词和文本匹配；否则只检查文本匹配
                var match = hasKeyword ? (keywordMatch && textMatch) : textMatch;
                
                if (match) {{
                    row.classList.remove('hidden');
                }} else {{
                    row.classList.add('hidden');
                }}
            }});
        }}
        
        // 排序功能
        var sortStates = {{}}; // 存储每个表格的排序状态 {{tabName: {{colIndex: 'asc'|'desc'|null}}}}
        var originalOrders = {{}}; // 存储每个表格的原始行顺序 {{tabName: [row1, row2, ...]}}
        
        // 通用的排序执行函数
        function executeSort(tabName, colIndex, sortDirection, headerElement) {{
            var tabContent = document.getElementById(tabName);
            if (!tabContent) return;
            
            var table = tabContent.querySelector('table');
            if (!table) return;
            
            var tbody = table.querySelector('tbody');
            if (!tbody) return;
            
            // 初始化排序状态
            if (!sortStates[tabName]) {{
                sortStates[tabName] = {{}};
            }}
            
            // 保存原始顺序（仅在第一次排序时保存）
            if (!originalOrders[tabName]) {{
                originalOrders[tabName] = tbody.innerHTML;
            }}
            
            // 如果sortDirection为null，恢复原始顺序
            if (sortDirection === null) {{
                sortStates[tabName][colIndex] = null;
                resetSortIcons(table);
                tbody.innerHTML = originalOrders[tabName];
                return;
            }}
            
            // 更新排序状态
            sortStates[tabName][colIndex] = sortDirection;
            
            // 重置所有排序图标
            resetSortIcons(table);
            
            // 设置当前列的排序图标
            var ascIcon = headerElement.querySelector('.sort-icon-asc');
            var descIcon = headerElement.querySelector('.sort-icon-desc');
            if (sortDirection === 'asc' && ascIcon) {{
                ascIcon.classList.add('asc-active');
            }} else if (sortDirection === 'desc' && descIcon) {{
                descIcon.classList.add('desc-active');
            }}
            
            // 获取所有行（排除汇总行）
            var rows = Array.from(tbody.querySelectorAll('tr.fund-row'));
            var summaryRow = tbody.querySelector('tr:not(.fund-row)');
            
            // 排序
            rows.sort(function(a, b) {{
                var aCell = a.cells[colIndex];
                var bCell = b.cells[colIndex];
                
                if (!aCell || !bCell) return 0;
                
                var aText = aCell.textContent.trim();
                var bText = bCell.textContent.trim();
                
                // 尝试解析为数字
                var aNum = parseFloat(aText.replace(/[+¥,%]/g, ''));
                var bNum = parseFloat(bText.replace(/[+¥,%]/g, ''));
                
                var aValue = isNaN(aNum) ? aText : aNum;
                var bValue = isNaN(bNum) ? bText : bNum;
                
                if (aValue < bValue) return sortDirection === 'asc' ? -1 : 1;
                if (aValue > bValue) return sortDirection === 'asc' ? 1 : -1;
                return 0;
            }});
            
            // 清空tbody并重新插入排序后的行
            tbody.innerHTML = '';
            rows.forEach(function(row) {{
                tbody.appendChild(row);
            }});
            // 如果有汇总行，添加到最后
            if (summaryRow) {{
                tbody.appendChild(summaryRow);
            }}
        }}
        
        // 点击列头进行三态循环排序：null -> desc -> asc -> null
        function sortTableByHeader(tabName, colIndex, headerElement) {{
            var tabContent = document.getElementById(tabName);
            if (!tabContent) return;
            
            var table = tabContent.querySelector('table');
            if (!table) return;
            
            // 初始化排序状态
            if (!sortStates[tabName]) {{
                sortStates[tabName] = {{}};
            }}
            
            // 获取当前排序状态
            var currentSort = sortStates[tabName][colIndex] || null;
            
            // 三态循环：null -> desc -> asc -> null
            var nextSort;
            if (currentSort === null) {{
                nextSort = 'desc';
            }} else if (currentSort === 'desc') {{
                nextSort = 'asc';
            }} else {{
                nextSort = null;
            }}
            
            // 执行排序
            executeSort(tabName, colIndex, nextSort, headerElement);
        }}
        
        // 点击图标直接进行指定排序
        function sortTableByIcon(tabName, colIndex, sortDirection, iconElement) {{
            // 找到对应的表头元素
            var tabContent = document.getElementById(tabName);
            if (!tabContent) return;
            
            var table = tabContent.querySelector('table');
            if (!table) return;
            
            var headers = table.querySelectorAll('th');
            var headerElement = headers[colIndex];
            
            if (!headerElement) return;
            
            // 初始化排序状态
            if (!sortStates[tabName]) {{
                sortStates[tabName] = {{}};
            }}
            
            // 获取当前排序状态
            var currentSort = sortStates[tabName][colIndex] || null;
            
            // 如果点击的是当前已激活的排序方向，则取消排序
            if (currentSort === sortDirection) {{
                executeSort(tabName, colIndex, null, headerElement);
            }} else {{
                // 否则直接设置指定的排序方向
                executeSort(tabName, colIndex, sortDirection, headerElement);
            }}
        }}
        
        // 重置所有排序图标
        function resetSortIcons(table) {{
            var allHeaders = table.querySelectorAll('th.sortable');
            allHeaders.forEach(function(h) {{
                var ascIcon = h.querySelector('.sort-icon-asc');
                var descIcon = h.querySelector('.sort-icon-desc');
                if (ascIcon) {{
                    ascIcon.classList.remove('asc-active');
                }}
                if (descIcon) {{
                    descIcon.classList.remove('desc-active');
                }}
            }});
        }}

        // 弹框相关变量
        var modal = document.getElementById('fundModal');
        var span = document.getElementsByClassName('close')[0];
        var navChart = null;
        var selectedFundCode = null;
        var selectedCostPrice = null;
        var selectedDays = 30;
        var selectedUser = null;
        var selectedTrades = [];
        var rangeButtons = null;

        // 关闭弹框
        span.onclick = function() {{
            modal.style.display = "none";
            if (navChart) {{
                navChart.destroy();
                navChart = null;
            }}
        }}

        // 点击弹框外部关闭
        window.onclick = function(event) {{
            if (event.target == modal) {{
                modal.style.display = "none";
                if (navChart) {{
                    navChart.destroy();
                    navChart = null;
                }}
            }}
        }}

        // 双击基金行显示弹框
        function showFundDetail(fundCode, fundName, costPrice, user) {{
            console.log('显示基金详情:', fundCode, fundName, costPrice, user);
            document.getElementById('modalTitle').textContent = fundName + ' (' + fundCode + ')';
            // 底部成本价区域已移除，改由悬浮提示

            // 显示弹框
            modal.style.display = "block";

            // 记录选择
            selectedFundCode = fundCode;
            selectedCostPrice = costPrice;
            selectedDays = 30;
            selectedUser = user || null;
            selectedTrades = [];

            // 绑定区间按钮
            rangeButtons = document.getElementsByClassName('range-btn');
            for (var i = 0; i < rangeButtons.length; i++) {{
                rangeButtons[i].classList.remove('active');
                if (rangeButtons[i].getAttribute('data-days') === '30') {{
                    rangeButtons[i].classList.add('active');
                }}
                (function(btn) {{
                    btn.onclick = function() {{
                        for (var j = 0; j < rangeButtons.length; j++) {{
                            rangeButtons[j].classList.remove('active');
                        }}
                        btn.classList.add('active');
                        var days = parseInt(btn.getAttribute('data-days')) || 30;
                        selectedDays = days;
                        loadFundData(selectedFundCode, selectedCostPrice, selectedDays, selectedUser);
                    }}
                }})(rangeButtons[i]);
            }}

            // 加载净值数据和图表（默认1月）
            loadFundData(fundCode, costPrice, selectedDays, selectedUser);
        }}

        // 加载基金数据（支持协议自动检测和切换）
        function loadFundData(fundCode, costPrice, days, user) {{
            console.log('加载基金数据:', fundCode, costPrice, user);
            days = days || 30;
            
            // 显示加载状态
            var chartContainer = document.getElementById('navChart');
            if (chartContainer) {{
                chartContainer.innerHTML = '<div style="text-align: center; padding: 50px; color: #666;">正在加载历史净值数据...</div>';
            }}
            
            // 构建 API URL，支持协议检查
            var apiUrl = window.getApiUrlWithProtocolCheck ? 
                window.getApiUrlWithProtocolCheck('/api/fund/history/' + fundCode + '?days=' + days) :
                (window.API_BASE + '/api/fund/history/' + fundCode + '?days=' + days);
            
            // 尝试获取数据，如果失败则尝试切换协议
            function tryFetch(url, tryAlternateProtocol) {{
                return fetch(url)
                    .then(response => {{
                        if (!response.ok) {{
                            // 如果是 400 Bad Request 且可能是协议问题，尝试切换协议
                            if (response.status === 400 && tryAlternateProtocol) {{
                                var alternateUrl = url.replace(/^https?:/, url.startsWith('https:') ? 'http:' : 'https:');
                                console.log('协议可能不匹配，尝试切换协议:', alternateUrl);
                                return tryFetch(alternateUrl, false); // 只尝试一次
                            }}
                            throw new Error(`HTTP ${{response.status}}`);
                        }}
                        return response.json();
                    }})
                    .catch(error => {{
                        // 如果是网络错误且是第一次尝试，尝试切换协议
                        if (tryAlternateProtocol && (error.message.includes('Failed to fetch') || error.message.includes('NetworkError'))) {{
                            var alternateUrl = url.replace(/^https?:/, url.startsWith('https:') ? 'http:' : 'https:');
                            console.log('网络错误，尝试切换协议:', alternateUrl);
                            return fetch(alternateUrl)
                                .then(resp => {{
                                    if (!resp.ok) throw new Error(`HTTP ${{resp.status}}`);
                                    return resp.json();
                                }})
                                .catch(() => {{
                                    throw error; // 如果切换协议也失败，抛出原始错误
                                }});
                        }}
                        throw error;
                    }});
            }}
            
            // 调用真实的历史净值API（近30天）- 支持协议自动检测和切换
            tryFetch(apiUrl, true)
                .then(data => {{
                    if (data.success && data.data && data.data.length > 0) {{
                        console.log('获取到真实历史数据:', data);
                        // 转换数据格式以适配图表显示
                        var chartData = data.data.map(function(item) {{
                            return {{
                                date: item.date,
                                nav: item.nav
                            }};
                        }});
                        // 日期升序排列，保证从左到右递增
                        chartData.sort(function(a, b) {{
                            return new Date(a.date) - new Date(b.date);
                        }});
                        var label = '近' + (days === 30 ? '1月' : days === 90 ? '3月' : days === 180 ? '6月' : '1年') + '净值走势';
                        // 如果有用户信息且是个人持仓或QDII用户，尝试获取交易记录
                        if (user && (user === 'chaochao' || user === 'yaoyao' || user === 'QDII')) {{
                            var tradesUrl = window.getApiUrlWithProtocolCheck ? 
                                window.getApiUrlWithProtocolCheck('/api/fund/trades/' + user + '/' + fundCode) :
                                (window.API_BASE + '/api/fund/trades/' + user + '/' + fundCode);
                            return fetch(tradesUrl)
                                .then(resp => {{
                                    if (!resp.ok) {{
                                        throw new Error(`HTTP ${{resp.status}}`);
                                    }}
                                    return resp.json();
                                }})
                                .then(tradeData => {{
                                    selectedTrades = tradeData.data || [];
                                    displayChart(chartData, fundCode, costPrice, label, selectedTrades);
                                }})
                                .catch(err => {{
                                    console.warn('交易记录获取失败，继续仅绘制净值:', err);
                                    displayChart(chartData, fundCode, costPrice, label, []);
                                }});
                        }} else {{
                            displayChart(chartData, fundCode, costPrice, label, []);
                        }}
                    }} else {{
                        console.error('API返回数据为空:', data);
                        if (chartContainer) {{
                            chartContainer.innerHTML = '<div style="text-align: center; padding: 50px; color: #c00;">未获取到历史净值数据（API无数据）</div>';
                        }}
                    }}
                }})
                .catch(error => {{
                    console.error('获取历史净值数据失败:', error);
                    var errorMsg = '获取历史净值数据失败。<br/>';
                    errorMsg += '当前 API_BASE: ' + (window.API_BASE || '(未设置)') + '<br/>';
                    errorMsg += '尝试的 URL: ' + apiUrl + '<br/>';
                    if (error.message.includes('400')) {{
                        errorMsg += '<strong style="color: #f00;">检测到协议不匹配错误（400 Bad Request）。</strong><br/>';
                        errorMsg += '请检查：<br/>';
                        errorMsg += '1. 如果页面通过 HTTPS 打开，API 服务器需要支持 HTTPS<br/>';
                        errorMsg += '2. 如果页面通过 HTTP 打开，API 服务器应使用 HTTP<br/>';
                        errorMsg += '3. 可通过 URL 参数 ?api=http://或https://你的API地址 手动指定协议<br/>';
                    }} else {{
                        errorMsg += '如在 GitHub Pages，请在 URL 追加 ?api=https://你的API域名 或在本地打开后设置 localStorage.API_BASE';
                    }}
                    if (chartContainer) {{
                        chartContainer.innerHTML = '<div style="text-align: center; padding: 50px; color: #c00;">' + errorMsg + '</div>';
                    }}
                }});
        }}

        /* 模拟数据逻辑已禁用以便排查真实API问题
        function generateMockData(fundCode) {{}}
        */

        // 显示图表
        function displayChart(data, fundCode, costPrice, titleLabel, trades) {{
            trades = trades || [];
            console.log('开始绘制SVG图表:', data, fundCode, costPrice, titleLabel, trades);
            
            // 创建交易记录日期映射（用于快速查找）
            // 注意：trades 已经通过 API 按基金代码过滤，所以这里只包含当前基金的交易记录
            var tradesByDate = {{}};
            var sortedDates = [];  // 提升到外层作用域，供后续使用
            if (trades && trades.length > 0) {{
                trades.forEach(function(trade) {{
                    var date = trade.date;
                    if (!tradesByDate[date]) {{
                        tradesByDate[date] = [];
                    }}
                    tradesByDate[date].push(trade);
                }});
                sortedDates = Object.keys(tradesByDate).sort();
            }}
            
            // 计算指定日期的份额变动（合计值）
            // 支持同一天多种操作（如买入+卖出），会正确合计
            function calculateSharesChange(dateStr) {{
                if (!tradesByDate[dateStr] || tradesByDate[dateStr].length === 0) {{
                    return null;
                }}
                var totalChange = 0;
                tradesByDate[dateStr].forEach(function(trade) {{
                    var action = (trade.action || '').toLowerCase();
                    var shares = parseFloat(trade.shares) || 0;
                    if (action === 'buy' || action === 'convert_in') {{
                        totalChange += shares;  // 买入/转入为正
                    }} else if (action === 'sell' || action === 'convert_out') {{
                        totalChange -= shares;  // 卖出/转出为负
                    }}
                }});
                return totalChange;
            }}
            
            var chartContainer = document.getElementById('navChart');
            if (!chartContainer) {{
                console.error('找不到图表容器元素');
                return;
            }}
            
            // 清空容器
            chartContainer.innerHTML = '';
            
            // 创建动态坐标轴标签容器
            var xAxisLabel = document.createElementNS('http://www.w3.org/2000/svg', 'text');
            var yAxisLabel = document.createElementNS('http://www.w3.org/2000/svg', 'text');
            

            // X轴标签稀疏显示：优先显示每月1日；若无，则显示每月15日；否则按间隔抽样
            var labels = [];
            var labelIndices = [];
            for (var i = 0; i < data.length; i++) {{
                var d = new Date(data[i].date);
                var day = d.getDate();
                if (day === 1 || day === 15) {{
                    labels.push((d.getMonth()+1) + '/' + day);
                    labelIndices.push(i);
                }}
            }}
            if (labels.length < 4) {{
                // 兜底：按固定间隔取4个点
                labels = [];
                labelIndices = [];
                var steps = 3;
                for (var s = 0; s <= steps; s++) {{
                    var idx = Math.round(s * (data.length - 1) / steps);
                    var dd = new Date(data[idx].date);
                    labels.push((dd.getMonth()+1) + '/' + dd.getDate());
                    labelIndices.push(idx);
                }}
            }}
            var values = data.map(function(item) {{
                return parseFloat(item.nav);
            }});
            
            console.log('图表标签:', labels);
            console.log('图表数值:', values);

            // 计算SVG尺寸
            var width = chartContainer.clientWidth - 24;
            var height = 400;
            var paddingLeft = 54, paddingRight = 16, paddingTop = 32, paddingBottom = 46;
            var chartWidth = width - paddingLeft - paddingRight;
            var chartHeight = height - paddingTop - paddingBottom;
            
            // 计算数据范围
            var minValue = Math.min(...values);
            var maxValue = Math.max(...values);
            var valueRange = maxValue - minValue;
            if (valueRange === 0) valueRange = 0.1;
            
            // 如果有成本价，调整范围
            if (costPrice && costPrice !== 'N/A') {{
                try {{
                    var costValue = parseFloat(costPrice);
                    if (!isNaN(costValue)) {{
                        minValue = Math.min(minValue, costValue);
                        maxValue = Math.max(maxValue, costValue);
                        valueRange = maxValue - minValue;
                        if (valueRange === 0) valueRange = 0.1;
                    }}
                }} catch (e) {{
                    console.log('成本价解析失败:', e);
                }}
            }}
            
            // 创建SVG元素
            var svg = document.createElementNS('http://www.w3.org/2000/svg', 'svg');
            svg.setAttribute('width', width);
            svg.setAttribute('height', height);
            svg.setAttribute('viewBox', '0 0 ' + width + ' ' + height);
            
            // 添加网格线（减少密度）
            for (var i = 0; i <= 4; i++) {{
                var y = paddingTop + (i * chartHeight / 4);
                var gridLine = document.createElementNS('http://www.w3.org/2000/svg', 'line');
                gridLine.setAttribute('x1', paddingLeft);
                gridLine.setAttribute('y1', y);
                gridLine.setAttribute('x2', width - paddingRight);
                gridLine.setAttribute('y2', y);
                gridLine.setAttribute('stroke', '#e0e0e0');
                gridLine.setAttribute('stroke-width', '1');
                svg.appendChild(gridLine);
            }}
            
            // 添加Y轴标签（5-6档稀疏显示）
            var yTicksCount = 5;
            for (var i = 0; i <= yTicksCount; i++) {{
                var y = paddingTop + (i * chartHeight / yTicksCount);
                var value = maxValue - (i * valueRange / yTicksCount);
                var text = document.createElementNS('http://www.w3.org/2000/svg', 'text');
                text.setAttribute('x', paddingLeft - 10);
                text.setAttribute('y', y + 4);
                text.setAttribute('text-anchor', 'end');
                text.setAttribute('font-size', '12');
                text.setAttribute('fill', '#666');
                text.textContent = value.toFixed(3);
                svg.appendChild(text);
            }}
            
            // 绘制净值曲线
            var points = [];
            for (var i = 0; i < values.length; i++) {{
                var x = paddingLeft + (i * chartWidth / (values.length - 1));
                var y = paddingTop + ((maxValue - values[i]) * chartHeight / valueRange);
                points.push(x + ',' + y);
            }}
            
            // 生成平滑曲线路径（使用二次贝塞尔曲线，保持圆滑）
            function buildSmoothPath(pts) {{
                if (pts.length < 2) return '';
                if (pts.length === 2) {{
                    return 'M ' + pts[0] + ' L ' + pts[1];
                }}
                var d = 'M ' + pts[0];
                // 使用平滑的二次贝塞尔曲线连接
                for (var i = 0; i < pts.length - 1; i++) {{
                    var p = pts[i].split(',');
                    var x1 = parseFloat(p[0]);
                    var y1 = parseFloat(p[1]);
                    var p2 = pts[i + 1].split(',');
                    var x2 = parseFloat(p2[0]);
                    var y2 = parseFloat(p2[1]);
                    // 使用中点作为控制点，创建平滑曲线
                    var mx = (x1 + x2) / 2;
                    var my = (y1 + y2) / 2;
                    d += ' Q ' + x1 + ' ' + y1 + ' ' + mx + ' ' + my;
                }}
                // 连接到最后一个点
                d += ' L ' + pts[pts.length - 1];
                return d;
            }}
            var path = document.createElementNS('http://www.w3.org/2000/svg', 'path');
            path.setAttribute('d', buildSmoothPath(points));
            path.setAttribute('stroke', '#007bff');
            path.setAttribute('stroke-width', '3');
            path.setAttribute('fill', 'none');
            path.setAttribute('stroke-linecap', 'round');
            path.setAttribute('stroke-linejoin', 'round');
            svg.appendChild(path);
            
            // 悬浮垂直虚线
            var hoverVerticalLine = document.createElementNS('http://www.w3.org/2000/svg', 'line');
            hoverVerticalLine.setAttribute('x1', paddingLeft);
            hoverVerticalLine.setAttribute('y1', paddingTop);
            hoverVerticalLine.setAttribute('x2', paddingLeft);
            hoverVerticalLine.setAttribute('y2', height - paddingBottom);
            hoverVerticalLine.setAttribute('stroke', '#007bff');
            hoverVerticalLine.setAttribute('stroke-width', '1.5');
            hoverVerticalLine.setAttribute('stroke-dasharray', '5,5');
            hoverVerticalLine.style.opacity = 0;
            svg.appendChild(hoverVerticalLine);
            
            // 悬浮水平虚线
            var hoverHorizontalLine = document.createElementNS('http://www.w3.org/2000/svg', 'line');
            hoverHorizontalLine.setAttribute('x1', paddingLeft);
            hoverHorizontalLine.setAttribute('y1', paddingTop);
            hoverHorizontalLine.setAttribute('x2', width - paddingRight);
            hoverHorizontalLine.setAttribute('y2', paddingTop);
            hoverHorizontalLine.setAttribute('stroke', '#007bff');
            hoverHorizontalLine.setAttribute('stroke-width', '1.5');
            hoverHorizontalLine.setAttribute('stroke-dasharray', '5,5');
            hoverHorizontalLine.style.opacity = 0;
            svg.appendChild(hoverHorizontalLine);
            
            // 如果有成本价，添加水平虚线
            if (costPrice && costPrice !== 'N/A') {{
                try {{
                    var costValue = parseFloat(costPrice);
                    if (!isNaN(costValue)) {{
                        var costY = paddingTop + ((maxValue - costValue) * chartHeight / valueRange);
                        var costLine = document.createElementNS('http://www.w3.org/2000/svg', 'line');
                        costLine.setAttribute('x1', paddingLeft);
                        costLine.setAttribute('y1', costY);
                        costLine.setAttribute('x2', width - paddingRight);
                        costLine.setAttribute('y2', costY);
                        costLine.setAttribute('stroke', '#ff6b6b');
                        costLine.setAttribute('stroke-width', '2');
                        costLine.setAttribute('stroke-dasharray', '5,5');
                        svg.appendChild(costLine);
                    }}
                }} catch (e) {{
                    console.log('成本价解析失败:', e);
                }}
            }}
            
            // 绘制交易标记（买入/卖出/转换），如果有
            if (trades && trades.length > 0) {{
                var markersGroup = document.createElementNS('http://www.w3.org/2000/svg', 'g');
                sortedDates.forEach(function(date) {{
                    var dayTrades = tradesByDate[date];
                    if (dayTrades.length === 0) return;
                    
                    try {{
                        var tDate = new Date(date);
                        // 精确匹配：优先查找完全匹配的日期
                        var bestIdx = -1;
                        var bestDiff = Infinity;
                        for (var i = 0; i < data.length; i++) {{
                            var dataDate = new Date(data[i].date);
                            var diff = Math.abs(dataDate - tDate);
                            // 优先完全匹配（diff为0）
                            if (diff === 0) {{
                                bestIdx = i;
                                bestDiff = 0;
                                break;
                            }}
                            if (diff < bestDiff) {{
                                bestDiff = diff;
                                bestIdx = i;
                            }}
                        }}
                        
                        // 如果找不到匹配点，使用第一个或最后一个
                        if (bestIdx < 0) {{
                            bestIdx = 0;
                        }}
                        if (bestIdx >= data.length) {{
                            bestIdx = data.length - 1;
                        }}
                        
                        // 使用points数组中对应索引的坐标，确保标记贴合曲线
                        // points数组的构建方式与曲线路径一致
                        if (bestIdx >= 0 && bestIdx < points.length) {{
                            var point = points[bestIdx].split(',');
                            var x = parseFloat(point[0]);
                            var y = parseFloat(point[1]);
                        }} else {{
                            // 边界情况，使用计算值
                            var x = paddingLeft + (bestIdx * chartWidth / (values.length - 1));
                            var priceVal = parseFloat(data[bestIdx].nav);
                            var y = paddingTop + ((maxValue - priceVal) * chartHeight / valueRange);
                        }}
                        
                        // 确定颜色（同一天有多个交易时，优先显示买入/转入）
                        var action = (dayTrades[0].action || '').toLowerCase();
                        var hasBuy = dayTrades.some(function(t) {{ return t.action === 'buy' || t.action === 'convert_in'; }});
                        var hasSell = dayTrades.some(function(t) {{ return t.action === 'sell' || t.action === 'convert_out'; }});
                        
                        var color = '#6c757d';
                        if (hasBuy && hasSell) {{
                            color = '#9c27b0';  // 同一天有买有卖，用紫色
                        }} else if (hasBuy) {{
                            color = '#007bff';  // 蓝色
                        }} else if (hasSell) {{
                            color = '#ff6b6b';  // 橙色
                        }}
                        
                        var marker = document.createElementNS('http://www.w3.org/2000/svg', 'circle');
                        marker.setAttribute('cx', x);
                        marker.setAttribute('cy', y);
                        marker.setAttribute('r', 5);
                        marker.setAttribute('fill', color);
                        marker.setAttribute('stroke', '#ffffff');
                        marker.setAttribute('stroke-width', '2');
                        markersGroup.appendChild(marker);
                    }} catch (e) {{
                        console.warn('交易标记渲染失败:', e);
                    }}
                }});
                svg.appendChild(markersGroup);
            }}
            
            // 添加X轴标签（使用抽样后的索引定位）
            for (var i = 0; i < labels.length; i++) {{
                var x = paddingLeft + (labelIndices[i] * chartWidth / (values.length - 1));
                var text = document.createElementNS('http://www.w3.org/2000/svg', 'text');
                text.setAttribute('x', x);
                text.setAttribute('y', height - paddingBottom + 20);
                text.setAttribute('text-anchor', 'middle');
                text.setAttribute('font-size', '12');
                text.setAttribute('fill', '#666');
                text.textContent = labels[i];
                svg.appendChild(text);
            }}
            
            // 曲线下方填充淡蓝色区域（使用与曲线相同的平滑路径）
            var areaPath = document.createElementNS('http://www.w3.org/2000/svg', 'path');
            var smoothPath = buildSmoothPath(points);
            // 构建填充路径：曲线路径 + 底部边界 + 闭合
            var firstPoint = points[0].split(',');
            var lastPoint = points[points.length - 1].split(',');
            var areaD = smoothPath + ' L ' + lastPoint[0] + ',' + (height - paddingBottom) + ' L ' + firstPoint[0] + ',' + (height - paddingBottom) + ' Z';
            areaPath.setAttribute('d', areaD);
            areaPath.setAttribute('fill', 'rgba(0,123,255,0.15)');
            areaPath.setAttribute('stroke', 'none');
            svg.insertBefore(areaPath, path);
            
            // 添加标题
            var title = document.createElementNS('http://www.w3.org/2000/svg', 'text');
            title.setAttribute('x', width / 2);
            title.setAttribute('y', paddingTop - 12);
            title.setAttribute('text-anchor', 'middle');
            title.setAttribute('font-size', '16');
            title.setAttribute('font-weight', 'bold');
            title.setAttribute('fill', '#333');
            title.textContent = titleLabel || '净值走势';
            svg.appendChild(title);

            // 默认显示有交易记录的日期信息（优先显示最新交易日期）
            function displayTradeInfoForDate(dateStr) {{
                if (!dateStr || !tradesByDate[dateStr] || tradesByDate[dateStr].length === 0) {{
                    return;
                }}
                
                // 找到对应日期的数据点
                var targetDate = new Date(dateStr);
                var bestIdx = -1;
                var bestDiff = Infinity;
                for (var i = 0; i < data.length; i++) {{
                    var dataDate = new Date(data[i].date);
                    var diff = Math.abs(dataDate - targetDate);
                    if (diff === 0) {{
                        bestIdx = i;
                        break;
                    }}
                    if (diff < bestDiff) {{
                        bestDiff = diff;
                        bestIdx = i;
                    }}
                }}
                
                if (bestIdx < 0 || bestIdx >= data.length) return;
                
                var currentData = data[bestIdx];
                var currentNav = currentData.nav;
                var currentDate = currentData.date;
                
                // 计算涨跌幅
                var changePercent = '';
                var changeClass = '';
                if (bestIdx > 0) {{
                    var prevNav = data[bestIdx - 1].nav;
                    var change = ((currentNav - prevNav) / prevNav * 100);
                    changePercent = (change >= 0 ? '+' : '') + change.toFixed(2) + '%';
                    changeClass = change >= 0 ? 'positive' : 'negative';
                }}
                
                // 计算份额变动和变动金额
                var sharesChange = calculateSharesChange(dateStr);
                var sharesChangeEl = document.getElementById('sharesChange');
                var changeAmountEl = document.getElementById('changeAmount');
                
                // 更新信息显示
                document.getElementById('transactionDate').textContent = currentDate;
                document.getElementById('unitNav').textContent = currentNav.toFixed(4);
                
                if (changePercent) {{
                    var changeElement = document.getElementById('changeRate');
                    changeElement.textContent = changePercent;
                    changeElement.className = 'info-value ' + changeClass;
                }} else {{
                    document.getElementById('changeRate').textContent = '-';
                    document.getElementById('changeRate').className = 'info-value';
                }}
                
                if (sharesChange !== null && sharesChange !== 0) {{
                    var sign = sharesChange > 0 ? '+' : '';
                    sharesChangeEl.textContent = sign + sharesChange.toFixed(2);
                    sharesChangeEl.className = 'info-value ' + (sharesChange > 0 ? 'positive' : 'negative');
                    
                    var changeAmount = sharesChange * currentNav;
                    changeAmountEl.textContent = (changeAmount >= 0 ? '+' : '') + changeAmount.toFixed(2);
                    changeAmountEl.className = 'info-value ' + (changeAmount >= 0 ? 'positive' : 'negative');
                }} else {{
                    sharesChangeEl.textContent = '-';
                    sharesChangeEl.className = 'info-value';
                    changeAmountEl.textContent = '-';
                    changeAmountEl.className = 'info-value';
                }}
                
                // 更新持仓成本信息
                if (costPrice && costPrice !== 'N/A') {{
                    document.getElementById('holdingCost').textContent = parseFloat(costPrice).toFixed(4);
                }}
            }}
            
            // 默认显示最新交易日期（如果有交易记录）
            if (sortedDates.length > 0) {{
                // 找到最新交易日期对应的数据点
                var latestTradeDate = sortedDates[sortedDates.length - 1];
                displayTradeInfoForDate(latestTradeDate);
            }} else {{
                // 如果没有交易记录，显示最新净值数据
                if (data.length > 0) {{
                    var latestData = data[data.length - 1];
                    document.getElementById('transactionDate').textContent = latestData.date;
                    document.getElementById('unitNav').textContent = latestData.nav.toFixed(4);
                    if (data.length > 1) {{
                        var prevNav = data[data.length - 2].nav;
                        var change = ((latestData.nav - prevNav) / prevNav * 100);
                        var changePercent = (change >= 0 ? '+' : '') + change.toFixed(2) + '%';
                        var changeElement = document.getElementById('changeRate');
                        changeElement.textContent = changePercent;
                        changeElement.className = 'info-value ' + (change >= 0 ? 'positive' : 'negative');
                    }}
                    if (costPrice && costPrice !== 'N/A') {{
                        document.getElementById('holdingCost').textContent = parseFloat(costPrice).toFixed(4);
                    }}
                }}
            }}
            
            // 悬浮交互（动态显示虚线）
            svg.addEventListener('mouseleave', function() {{
                hoverVerticalLine.style.opacity = 0;
                hoverHorizontalLine.style.opacity = 0;
                xAxisLabel.style.opacity = 0;
                yAxisLabel.style.opacity = 0;
                
                // 移除圆标记
                var existingCircle = svg.querySelector('.hover-circle');
                if (existingCircle) {{
                    svg.removeChild(existingCircle);
                }}
                
                // 恢复默认显示（最新交易日期或最新净值）
                if (sortedDates.length > 0) {{
                    var latestTradeDate = sortedDates[sortedDates.length - 1];
                    displayTradeInfoForDate(latestTradeDate);
                }} else if (data.length > 0) {{
                    var latestData = data[data.length - 1];
                    document.getElementById('transactionDate').textContent = latestData.date;
                    document.getElementById('unitNav').textContent = latestData.nav.toFixed(4);
                    if (data.length > 1) {{
                        var prevNav = data[data.length - 2].nav;
                        var change = ((latestData.nav - prevNav) / prevNav * 100);
                        var changePercent = (change >= 0 ? '+' : '') + change.toFixed(2) + '%';
                        var changeElement = document.getElementById('changeRate');
                        changeElement.textContent = changePercent;
                        changeElement.className = 'info-value ' + (change >= 0 ? 'positive' : 'negative');
                    }}
                    document.getElementById('sharesChange').textContent = '-';
                    document.getElementById('sharesChange').className = 'info-value';
                    document.getElementById('changeAmount').textContent = '-';
                    document.getElementById('changeAmount').className = 'info-value';
                    if (costPrice && costPrice !== 'N/A') {{
                        document.getElementById('holdingCost').textContent = parseFloat(costPrice).toFixed(4);
                    }}
                }}
            }});
            
            svg.addEventListener('mousemove', function(evt) {{
                var rect = svg.getBoundingClientRect();
                var x = evt.clientX - rect.left;
                var y = evt.clientY - rect.top;
                
                // 限定到绘图区域
                var cx = Math.max(paddingLeft, Math.min(width - paddingRight, x));
                var cy = Math.max(paddingTop, Math.min(height - paddingBottom, y));
                
                // 计算对应的数据索引
                var ratio = (cx - paddingLeft) / chartWidth;
                var fIndex = ratio * (values.length - 1);
                var idx = Math.round(fIndex); // 四舍五入到最近的数据点
                
                // 计算更精确的曲线位置
                // 使用points数组中对应索引的坐标，确保悬浮圆贴合曲线
                var exactX = cx;
                var exactY;
                
                if (idx >= 0 && idx < points.length) {{
                    // 直接使用points数组中的坐标，确保与曲线路径一致
                    var point = points[idx].split(',');
                    exactX = parseFloat(point[0]);
                    exactY = parseFloat(point[1]);
                }} else if (fIndex >= 0 && fIndex < values.length - 1) {{
                    // 在两个数据点之间进行线性插值（作为备选方案）
                    var idx1 = Math.floor(fIndex);
                    var idx2 = Math.ceil(fIndex);
                    var weight = fIndex - idx1;
                    
                    if (idx1 >= 0 && idx1 < points.length && idx2 >= 0 && idx2 < points.length) {{
                        var p1 = points[idx1].split(',');
                        var p2 = points[idx2].split(',');
                        var x1 = parseFloat(p1[0]);
                        var y1 = parseFloat(p1[1]);
                        var x2 = parseFloat(p2[0]);
                        var y2 = parseFloat(p2[1]);
                        exactX = cx;
                        exactY = y1 + (y2 - y1) * weight;
                    }} else {{
                        var nav1 = data[idx1].nav;
                        var nav2 = data[idx2].nav;
                        var interpolatedNav = nav1 + (nav2 - nav1) * weight;
                        exactY = paddingTop + ((maxValue - interpolatedNav) * chartHeight / valueRange);
                    }}
                }} else {{
                    // 边界情况，使用最近的数据点
                    if (idx >= 0 && idx < points.length) {{
                        var point = points[idx].split(',');
                        exactX = parseFloat(point[0]);
                        exactY = parseFloat(point[1]);
                    }} else {{
                        var currentNav = data[idx].nav;
                        exactY = paddingTop + ((maxValue - currentNav) * chartHeight / valueRange);
                    }}
                }}
                
                // 获取当前数据点信息
                var currentData = data[idx];
                var currentNav = currentData.nav;
                var currentDate = currentData.date;
                
                // 计算涨跌幅（如果有前一天数据）
                var changePercent = '';
                var changeClass = '';
                if (idx > 0) {{
                    var prevNav = data[idx - 1].nav;
                    var change = ((currentNav - prevNav) / prevNav * 100);
                    changePercent = (change >= 0 ? '+' : '') + change.toFixed(2) + '%';
                    changeClass = change >= 0 ? 'positive' : 'negative';
                }}
                
                // 设置垂直虚线位置
                hoverVerticalLine.setAttribute('x1', cx);
                hoverVerticalLine.setAttribute('x2', cx);
                hoverVerticalLine.style.opacity = 1;
                
                // 设置水平虚线位置（使用插值计算的精确位置）
                hoverHorizontalLine.setAttribute('y1', exactY);
                hoverHorizontalLine.setAttribute('y2', exactY);
                hoverHorizontalLine.style.opacity = 1;
                
                // 创建空心圆标记（确保XY轴交点落在曲线上）
                // 移除之前的圆标记
                var existingCircle = svg.querySelector('.hover-circle');
                if (existingCircle) {{
                    svg.removeChild(existingCircle);
                }}
                
                // 创建空心圆组
                var circleGroup = document.createElementNS('http://www.w3.org/2000/svg', 'g');
                circleGroup.setAttribute('class', 'hover-circle');
                
                // 空心圆（与趋势线颜色相近，边框加粗）
                var hollowCircle = document.createElementNS('http://www.w3.org/2000/svg', 'circle');
                hollowCircle.setAttribute('cx', exactX);
                hollowCircle.setAttribute('cy', exactY);
                hollowCircle.setAttribute('r', '8');
                hollowCircle.setAttribute('fill', 'none');
                hollowCircle.setAttribute('stroke', '#007bff');
                hollowCircle.setAttribute('stroke-width', '4');
                circleGroup.appendChild(hollowCircle);
                
                // 将空心圆组添加到SVG
                svg.appendChild(circleGroup);
                
                // 更新信息显示区域
                document.getElementById('transactionDate').textContent = currentDate;
                document.getElementById('unitNav').textContent = currentNav.toFixed(4);
                
                if (changePercent) {{
                    var changeElement = document.getElementById('changeRate');
                    changeElement.textContent = changePercent;
                    changeElement.className = 'info-value ' + changeClass;
                }} else {{
                    document.getElementById('changeRate').textContent = '-';
                    document.getElementById('changeRate').className = 'info-value';
                }}
                
                // 更新持仓成本信息（如果鼠标移动时有成本价）
                if (costPrice && costPrice !== 'N/A') {{
                    document.getElementById('holdingCost').textContent = parseFloat(costPrice).toFixed(4);
                }}
                
                // 计算并显示份额变动和变动金额（仅显示当前日期的交易记录）
                var sharesChange = calculateSharesChange(currentDate);
                var sharesChangeEl = document.getElementById('sharesChange');
                var changeAmountEl = document.getElementById('changeAmount');
                if (sharesChange !== null && sharesChange !== 0) {{
                    var sign = sharesChange > 0 ? '+' : '';
                    sharesChangeEl.textContent = sign + sharesChange.toFixed(2);
                    sharesChangeEl.className = 'info-value ' + (sharesChange > 0 ? 'positive' : 'negative');
                    
                    // 计算变动金额：份额变动 * 当日净值
                    var changeAmount = sharesChange * currentNav;
                    changeAmountEl.textContent = (changeAmount >= 0 ? '+' : '') + changeAmount.toFixed(2);
                    changeAmountEl.className = 'info-value ' + (changeAmount >= 0 ? 'positive' : 'negative');
                }} else {{
                    sharesChangeEl.textContent = '-';
                    sharesChangeEl.className = 'info-value';
                    changeAmountEl.textContent = '-';
                    changeAmountEl.className = 'info-value';
                }}
                
                // 显示动态坐标轴标签
                // X轴标签（交易日期）
                xAxisLabel.setAttribute('x', cx);
                xAxisLabel.setAttribute('y', height - paddingBottom + 35);
                xAxisLabel.setAttribute('text-anchor', 'middle');
                xAxisLabel.setAttribute('font-size', '12');
                xAxisLabel.setAttribute('fill', '#007bff');
                xAxisLabel.setAttribute('font-weight', 'bold');
                xAxisLabel.textContent = currentDate;
                xAxisLabel.style.opacity = 1;
                if (!svg.contains(xAxisLabel)) {{
                    svg.appendChild(xAxisLabel);
                }}
                
                // Y轴标签（单位净值）
                yAxisLabel.setAttribute('x', paddingLeft - 15);
                yAxisLabel.setAttribute('y', exactY + 4);
                yAxisLabel.setAttribute('text-anchor', 'end');
                yAxisLabel.setAttribute('font-size', '12');
                yAxisLabel.setAttribute('fill', '#007bff');
                yAxisLabel.setAttribute('font-weight', 'bold');
                yAxisLabel.textContent = currentNav.toFixed(4);
                yAxisLabel.style.opacity = 1;
                if (!svg.contains(yAxisLabel)) {{
                    svg.appendChild(yAxisLabel);
                }}
            }});
            
            // 初始化信息显示区域
            document.getElementById('transactionDate').textContent = '-';
            document.getElementById('unitNav').textContent = '-';
            document.getElementById('changeRate').textContent = '-';
            document.getElementById('changeRate').className = 'info-value';
            
            if (costPrice && costPrice !== 'N/A') {{
                try {{
                    document.getElementById('holdingCost').textContent = parseFloat(costPrice).toFixed(4);
                }} catch (e) {{
                    document.getElementById('holdingCost').textContent = '-';
                }}
            }} else {{
                document.getElementById('holdingCost').textContent = '-';
            }}
            
            // 将SVG添加到容器
            chartContainer.appendChild(svg);
            console.log('SVG图表创建成功');
        }}
        
        function updateFundInfoCard(date, value, costPrice, data) {{
            // 更新交易日期
            document.getElementById('transactionDate').textContent = date;
            
            // 更新单位净值
            document.getElementById('unitNav').textContent = value.toFixed(4);
            
            // 更新持仓成本
            var costElement = document.getElementById('holdingCost');
            if (costPrice && costPrice !== 'N/A' && costPrice !== '-') {{
                costElement.textContent = parseFloat(costPrice).toFixed(4);
            }} else {{
                costElement.textContent = '-';
            }}
            
            // 计算并更新涨跌幅
            var changeRateElement = document.getElementById('changeRate');
            var currentIndex = data.findIndex(function(item) {{ return item.date === date; }});
            
            if (currentIndex > 0) {{
                var prevValue = parseFloat(data[currentIndex - 1].nav);
                var changeRate = ((value - prevValue) / prevValue) * 100;
                var changeText = (changeRate >= 0 ? '+' : '') + changeRate.toFixed(2) + '%';
                
                changeRateElement.textContent = changeText;
                changeRateElement.className = 'info-value ' + (changeRate > 0 ? 'positive' : changeRate < 0 ? 'negative' : 'neutral');
            }} else {{
                changeRateElement.textContent = '-';
                changeRateElement.className = 'info-value neutral';
            }}
        }}
    </script>
</body>
</html>"""
    
    # 计算收益颜色类的辅助函数
    def get_profit_color_class_for_table(value_str):
        if value_str == 'N/A':
            return "neutral"
        try:
            # 提取数字部分
            value = float(str(value_str).replace(',', '').replace('¥', '').replace('%', ''))
            if value > 0:
                return "positive"
            elif value < 0:
                return "negative"
            else:
                return "neutral"
        except:
            return "neutral"
    
    # 生成表格行的函数（用于自选基金，包含持仓信息）
    def generate_table_rows(fund_list, user_key):
        rows = ""
        total_today_profit = 0.0
        total_holdings_profit = 0.0
        
        for fund in fund_list:
            change_rate = fund['估算涨跌率']
            try:
                change_value = float(change_rate)
                if change_value > 0:
                    change_class = "positive"
                    change_symbol = "+"
                elif change_value < 0:
                    change_class = "negative"
                    change_symbol = ""
                else:
                    change_class = "neutral"
                    change_symbol = ""
            except:
                change_class = "neutral"
                change_symbol = ""
            
            # 获取持仓信息（如果存在）
            cost_price = fund.get('成本单价', 'N/A')
            today_profit = fund.get('当日收益', 'N/A')
            total_profit = fund.get('持仓收益', 'N/A')
            total_profit_rate = fund.get('持仓收益率', 'N/A')
            
            # 累计收益数据用于汇总行
            if today_profit != 'N/A':
                try:
                    profit_str = str(today_profit).replace('¥', '').replace(',', '')
                    total_today_profit += float(profit_str)
                except:
                    pass
            
            if total_profit != 'N/A':
                try:
                    profit_str = str(total_profit).replace('¥', '').replace(',', '')
                    total_holdings_profit += float(profit_str)
                except:
                    pass
            
            # 格式化日期显示
            jzrq = fund.get('净值日期', '')
            gztime = fund.get('估值时间', '')
            
            # 最新净值显示昨日日期
            latest_date_display = ""
            if jzrq:
                try:
                    # 提取mm-dd格式
                    date_obj = datetime.strptime(jzrq, '%Y-%m-%d')
                    latest_date_display = date_obj.strftime('%m-%d')
                except:
                    latest_date_display = jzrq
            
            # 估算净值列标题显示日期，明细行显示时间
            estimate_date_display = ""
            estimate_time_display = ""
            
            # 对于境外基金，使用净值日期显示日期，确保与境外基金tab页一致
            if fund.get('板块分类') == '境外基金':
                # 境外基金：最新净价显示最新净值日期，估值净价显示上日净值日期
                if jzrq and jzrq != 'N/A':
                    try:
                        date_obj = datetime.strptime(jzrq, '%Y-%m-%d')
                        latest_date_display = date_obj.strftime('%m-%d')  # 最新净值日期
                    except:
                        latest_date_display = jzrq
                else:
                    latest_date_display = "N/A"
                
                # 估值净价显示上日净值日期
                if gztime and gztime != 'N/A':
                    try:
                        if ' ' in gztime:
                            date_part = gztime.split(' ')[0]
                            date_obj = datetime.strptime(date_part, '%Y-%m-%d')
                            estimate_date_display = date_obj.strftime('%m-%d')
                        else:
                            date_obj = datetime.strptime(gztime, '%Y-%m-%d')
                            estimate_date_display = date_obj.strftime('%m-%d')
                        estimate_time_display = ""  # 境外基金不显示时间
                    except:
                        estimate_date_display = gztime
                        estimate_time_display = ""
                else:
                    estimate_date_display = "N/A"
                    estimate_time_display = ""
            else:
                # 对于其他基金，使用原有的逻辑
                if gztime:
                    try:
                        # 解析估值时间，格式为 "YYYY-MM-DD HH:mm" 或 "HH:mm"
                        if ' ' in gztime:
                            date_part, time_part = gztime.split(' ', 1)
                            date_obj = datetime.strptime(date_part, '%Y-%m-%d')
                            estimate_date_display = date_obj.strftime('%m-%d')
                            estimate_time_display = time_part
                        else:
                            # 如果只有时间部分，使用当前日期
                            estimate_date_display = datetime.now().strftime('%m-%d')
                            estimate_time_display = gztime
                    except:
                        estimate_date_display = datetime.now().strftime('%m-%d')
                        estimate_time_display = ""
                else:
                    estimate_date_display = datetime.now().strftime('%m-%d')
                    estimate_time_display = ""
            
            today_profit_class = get_profit_color_class_for_table(today_profit)
            # 修正：持仓收益与持仓收益率按自身红涨绿跌
            total_profit_class = get_profit_color_class_for_table(total_profit)
            total_profit_rate_class = get_profit_color_class_for_table(total_profit_rate)
            
            rows += f"""
                <tr class="fund-row" ondblclick="showFundDetail('{fund['基金代码']}', '{fund['基金名称']}', '{cost_price}', '{user_key}')">
                    <td>{fund['基金代码']}</td>
                    <td>{fund['基金名称']}</td>
                    <td><span class="category-tag category-{fund['板块分类']}">{fund['板块分类']}</span></td>
                    <td>
                        {fund['最新净值']}
                        <div style="font-size: 12px; color: #6c757d; margin-top: 2px;">{latest_date_display}</div>
                    </td>
                    <td>
                        {fund['估算净值']}
                        <div style="font-size: 12px; color: #6c757d; margin-top: 2px;">{estimate_date_display if fund.get('板块分类') == '境外基金' else estimate_time_display}</div>
                    </td>
                    <td class="{change_class}">{change_symbol}{change_rate}%</td>
                    <td>{cost_price}</td>
                    <td class="{today_profit_class}">{today_profit}</td>
                    <td class="{total_profit_class}">{total_profit}</td>
                    <td class="{total_profit_rate_class}">{total_profit_rate}</td>
                </tr>"""
        
        # 添加汇总行
        if len(fund_list) > 0:
            # 计算汇总行的CSS类
            today_profit_summary_class = get_profit_color_class_for_table(f"{total_today_profit:,.2f}")
            holdings_profit_summary_class = get_profit_color_class_for_table(f"{total_holdings_profit:,.2f}")
            
            rows += f"""
                <tr style="background-color: #f8f9fa; font-weight: bold;">
                    <td colspan="7" style="text-align: left;">汇总:</td>
                    <td class="{today_profit_summary_class}">{total_today_profit:,.2f}</td>
                    <td class="{holdings_profit_summary_class}">{total_holdings_profit:,.2f}</td>
                    <td>-</td>
                </tr>"""
        
        return rows
    
    # 生成监控基金表格行的函数（不包含持仓信息）
    def generate_monitor_table_rows(fund_list):
        rows = ""
        for fund in fund_list:
            change_rate = fund['估算涨跌率']
            try:
                change_value = float(change_rate)
                if change_value > 0:
                    change_class = "positive"
                    change_symbol = "+"
                elif change_value < 0:
                    change_class = "negative"
                    change_symbol = ""
                else:
                    change_class = "neutral"
                    change_symbol = ""
            except:
                change_class = "neutral"
                change_symbol = ""
            
            # 格式化日期显示
            jzrq = fund.get('净值日期', '')
            gztime = fund.get('估值时间', '')
            
            # 最新净值显示昨日日期
            latest_date_display = ""
            if jzrq:
                try:
                    # 提取mm-dd格式
                    date_obj = datetime.strptime(jzrq, '%Y-%m-%d')
                    latest_date_display = date_obj.strftime('%m-%d')
                except:
                    latest_date_display = jzrq
            
            # 估算净值列标题显示日期，明细行显示时间
            estimate_date_display = ""
            estimate_time_display = ""
            if gztime:
                try:
                    # 解析估值时间，格式为 "YYYY-MM-DD HH:mm" 或 "HH:mm"
                    if ' ' in gztime:
                        date_part, time_part = gztime.split(' ', 1)
                        date_obj = datetime.strptime(date_part, '%Y-%m-%d')
                        estimate_date_display = date_obj.strftime('%m-%d')
                        estimate_time_display = time_part
                    else:
                        # 如果只有时间部分，使用当前日期
                        estimate_date_display = datetime.now().strftime('%m-%d')
                        estimate_time_display = gztime
                except:
                    estimate_date_display = datetime.now().strftime('%m-%d')
                    estimate_time_display = ""
            else:
                estimate_date_display = datetime.now().strftime('%m-%d')
                estimate_time_display = ""
            
            rows += f"""
                <tr class="fund-row" ondblclick="showFundDetail('{fund['基金代码']}', '{fund['基金名称']}', 'N/A', 'monitor')">
                    <td>{fund['基金代码']}</td>
                    <td>{fund['基金名称']}</td>
                    <td><span class="category-tag category-{fund['板块分类']}">{fund['板块分类']}</span></td>
                    <td>
                        {fund['最新净值']}
                        <div style="font-size: 12px; color: #6c757d; margin-top: 2px;">{latest_date_display}</div>
                    </td>
                    <td>
                        {fund['估算净值']}
                        <div style="font-size: 12px; color: #6c757d; margin-top: 2px;">{estimate_time_display}</div>
                    </td>
                    <td class="{change_class}">{change_symbol}{change_rate}%</td>
                </tr>"""
        return rows
    
    # 生成境外基金表格行的函数（不包含持仓信息，但支持双击绘制图表）
    # 改版：移除"估算净值"，新增"上日净值"，并计算涨跌幅
    def generate_overseas_table_rows(fund_list):
        rows = ""
        
        for fund in fund_list:
            change_rate = fund['估算涨跌率']
            try:
                change_value = float(change_rate)
                if change_value > 0:
                    change_class = "positive"
                    change_symbol = "+"
                elif change_value < 0:
                    change_class = "negative"
                    change_symbol = ""
                else:
                    change_class = "neutral"
                    change_symbol = ""
            except:
                change_class = "neutral"
                change_symbol = ""
            
            # 直接使用基金数据中的日期字段，格式化为MM-dd
            latest_nav_value = fund.get('最新净值', 'N/A')  # 最新净值
            prev_nav_value = fund.get('估算净值', 'N/A')   # 上日净值
            
            # 格式化日期显示
            latest_date_display = ""
            prev_date_display = ""
            
            # 处理净值日期（最新净值日期）
            jzrq = fund.get('净值日期', '')
            if jzrq and jzrq != 'N/A':
                try:
                    date_obj = datetime.strptime(jzrq, '%Y-%m-%d')
                    latest_date_display = date_obj.strftime('%m-%d')
                except:
                    latest_date_display = ""
            
            # 处理估值时间（上日净值日期）
            gztime = fund.get('估值时间', '')
            if gztime and gztime != 'N/A':
                try:
                    # 如果估值时间包含日期部分，提取并格式化
                    if ' ' in gztime:
                        date_part = gztime.split(' ')[0]
                        date_obj = datetime.strptime(date_part, '%Y-%m-%d')
                        prev_date_display = date_obj.strftime('%m-%d')
                    else:
                        # 如果只有日期，直接格式化
                        date_obj = datetime.strptime(gztime, '%Y-%m-%d')
                        prev_date_display = date_obj.strftime('%m-%d')
                except:
                    prev_date_display = ""
            
            rows += f"""
                <tr class="fund-row" ondblclick="showFundDetail('{fund['基金代码']}', '{fund['基金名称']}', 'N/A', 'QDII')">
                    <td>{fund['基金代码']}</td>
                    <td>{fund['基金名称']}</td>
                    <td><span class="category-tag category-境外基金">境外基金</span></td>
                    <td>
                        {latest_nav_value}
                        <div style="font-size: 12px; color: #6c757d; margin-top: 2px;">{latest_date_display}</div>
                    </td>
                    <td>
                        {prev_nav_value}
                        <div style="font-size: 12px; color: #6c757d; margin-top: 2px;">{prev_date_display}</div>
                    </td>
                    <td class="{change_class}">{change_symbol}{change_rate}%</td>
                </tr>"""
        
        return rows
    
    # 生成场内基金表格行的函数（不包含持仓信息，但支持双击绘制图表）
    def generate_etf_table_rows(fund_list):
        rows = ""
        
        for fund in fund_list:
            change_rate = fund['估算涨跌率']
            try:
                change_value = float(change_rate)
                if change_value > 0:
                    change_class = "positive"
                    change_symbol = "+"
                elif change_value < 0:
                    change_class = "negative"
                    change_symbol = ""
                else:
                    change_class = "neutral"
                    change_symbol = ""
            except:
                change_class = "neutral"
                change_symbol = ""
            
            # ETF基金：最新净价显示当前价格，上日净价显示昨收盘价
            jzrq = fund.get('净值日期', '')
            gztime = fund.get('估值时间', '')
            prev_trade_date = fund.get('昨收盘价日期', '')
            
            # 最新净价显示时间 HH:mm（净价公布时间）
            latest_time_display = ""
            if gztime:
                try:
                    # 解析估值时间，格式为 "YYYY-MM-DD HH:mm" 或 "HH:mm"
                    if ' ' in gztime:
                        time_part = gztime.split(' ', 1)[1]
                        latest_time_display = time_part
                    else:
                        # 如果只有时间部分，直接使用
                        latest_time_display = gztime
                except:
                    latest_time_display = ""
            
            # 上日净价显示日期 MM-DD（昨收盘价日期，即上一交易日）
            prev_date_display = ""
            if prev_trade_date and prev_trade_date != 'N/A':
                try:
                    # 提取mm-dd格式
                    date_obj = datetime.strptime(prev_trade_date, '%Y-%m-%d')
                    prev_date_display = date_obj.strftime('%m-%d')
                except:
                    prev_date_display = prev_trade_date
            elif jzrq:
                try:
                    # 如果没有昨收盘价日期，使用净值日期
                    date_obj = datetime.strptime(jzrq, '%Y-%m-%d')
                    prev_date_display = date_obj.strftime('%m-%d')
                except:
                    prev_date_display = jzrq
            
            rows += f"""
                <tr class="fund-row" ondblclick="showFundDetail('{fund['基金代码']}', '{fund['基金名称']}', 'N/A', 'etf')">
                    <td>{fund['基金代码']}</td>
                    <td>{fund['基金名称']}</td>
                    <td><span class="category-tag category-ETF基金">ETF基金</span></td>
                    <td>
                        {fund['最新净值']}
                        <div style="font-size: 12px; color: #6c757d; margin-top: 2px;">{latest_time_display}</div>
                    </td>
                    <td>
                        {fund['估算净值']}
                        <div style="font-size: 12px; color: #6c757d; margin-top: 2px;">{prev_date_display}</div>
                    </td>
                    <td class="{change_class}">{change_symbol}{change_rate}%</td>
                </tr>"""
        
        return rows
    
    # 获取各组数据（已按原始定义顺序排序）
    chaochao_data = fund_data_dict.get('chaochao', [])
    yaoyao_data = fund_data_dict.get('yaoyao', [])
    overseas_fund_data = fund_data_dict.get('overseas', [])
    etf_fund_data = fund_data_dict.get('etf', [])
    monitor_data = monitor_funds if monitor_funds else []
    
    # 统计函数
    def compute_stats(fund_list):
        up = down = flat = 0
        total = 0.0
        count = 0
        for fund in fund_list:
            try:
                change_value = float(fund.get('估算涨跌率'))
                total += change_value
                count += 1
                if change_value > 0:
                    up += 1
                elif change_value < 0:
                    down += 1
                else:
                    flat += 1
            except Exception:
                continue
        avg = f"{(total / count):.2f}%" if count > 0 else "0.00%"
        return avg, up, down, flat

    # 各组统计
    avg_change, up_count, down_count, flat_count = compute_stats(monitor_data)
    chaochao_avg_change, chaochao_up_count, chaochao_down_count, chaochao_flat_count = compute_stats(chaochao_data)
    yaoyao_avg_change, yaoyao_up_count, yaoyao_down_count, yaoyao_flat_count = compute_stats(yaoyao_data)
    
    # 计算当日预估收益
    def compute_today_profit(fund_list):
        total_profit = 0.0
        for fund in fund_list:
            try:
                today_profit = fund.get('当日收益', 'N/A')
                if today_profit != 'N/A':
                    # 提取数字部分
                    profit_str = str(today_profit).replace('¥', '').replace(',', '')
                    total_profit += float(profit_str)
            except:
                continue
        return f"{total_profit:,.2f}" if total_profit != 0 else "0.00"
    
    chaochao_today_profit = compute_today_profit(chaochao_data)
    yaoyao_today_profit = compute_today_profit(yaoyao_data)
    
    # 计算持仓收益汇总信息（如果提供了profit_results）
    chaochao_holdings_summary = ""
    yaoyao_holdings_summary = ""
    if profit_results:
        try:
            chaochao_result = profit_results.get('chaochao', {})
            yaoyao_result = profit_results.get('yaoyao', {})
            
            chaochao_holdings_summary = (
                f"投入:{chaochao_result.get('total_cost', 0):,.0f} "
                f"市值:{chaochao_result.get('total_current_value', 0):,.0f} "
                f"总收益:{chaochao_result.get('total_profit', 0):,.0f}({chaochao_result.get('total_profit_rate', 0):+.1f}%) "
                f"今日:{chaochao_result.get('today_profit', 0):,.0f}({chaochao_result.get('today_profit_rate', 0):+.1f}%)"
            )
            
            yaoyao_holdings_summary = (
                f"投入:{yaoyao_result.get('total_cost', 0):,.0f} "
                f"市值:{yaoyao_result.get('total_current_value', 0):,.0f} "
                f"总收益:{yaoyao_result.get('total_profit', 0):,.0f}({yaoyao_result.get('total_profit_rate', 0):+.1f}%) "
                f"今日:{yaoyao_result.get('today_profit', 0):,.0f}({yaoyao_result.get('today_profit_rate', 0):+.1f}%)"
            )
        except Exception as e:
            log_info(f"⚠️ 计算持仓汇总信息失败: {e}")
            chaochao_holdings_summary = "数据不足"
            yaoyao_holdings_summary = "数据不足"
    
    # 计算收益CSS类
    def get_profit_css_class(profit_str):
        if profit_str == '0.00':
            return 'profit-neutral'
        try:
            profit_value = float(profit_str.replace(',', ''))
            if profit_value > 0:
                return 'profit-positive'
            elif profit_value < 0:
                return 'profit-negative'
            else:
                return 'profit-neutral'
        except:
            return 'profit-neutral'
    
    chaochao_profit_class = get_profit_css_class(chaochao_today_profit)
    yaoyao_profit_class = get_profit_css_class(yaoyao_today_profit)
    
    # 计算平均涨跌率CSS类
    def get_avg_change_css_class(avg_change_str):
        try:
            # 提取数字部分，去掉%符号
            avg_value = float(avg_change_str.replace('%', ''))
            if avg_value > 0:
                return 'positive'
            elif avg_value < 0:
                return 'negative'
            else:
                return 'neutral'
        except:
            return 'neutral'
    
    monitor_avg_class = get_avg_change_css_class(avg_change)
    chaochao_avg_class = get_avg_change_css_class(chaochao_avg_change)
    yaoyao_avg_class = get_avg_change_css_class(yaoyao_avg_change)
    
    # 计算日期标题
    latest_date_header = ""
    estimate_date_header = ""
    
    # 场内基金表头变量
    etf_latest_time_header = ""
    etf_prev_date_header = ""
    
    # 从第一个基金数据中获取日期信息
    if chaochao_data:
        first_fund = chaochao_data[0]
        jzrq = first_fund.get('净值日期', '')
        gztime = first_fund.get('估值时间', '')
        
        # 最新净值标题显示昨日日期
        if jzrq:
            try:
                date_obj = datetime.strptime(jzrq, '%Y-%m-%d')
                latest_date_header = date_obj.strftime('%m-%d')
            except:
                latest_date_header = jzrq
        
        # 估算净值标题显示今日日期
        if gztime:
            try:
                if ' ' in gztime:
                    date_part = gztime.split(' ')[0]
                    date_obj = datetime.strptime(date_part, '%Y-%m-%d')
                    estimate_date_header = date_obj.strftime('%m-%d')
                else:
                    estimate_date_header = datetime.now().strftime('%m-%d')
            except:
                estimate_date_header = datetime.now().strftime('%m-%d')
        else:
            estimate_date_header = datetime.now().strftime('%m-%d')
    
            # 计算ETF基金表头
        if etf_fund_data:
            first_etf = etf_fund_data[0]
            etf_jzrq = first_etf.get('净值日期', '')
            etf_gztime = first_etf.get('估值时间', '')
            etf_prev_trade_date = first_etf.get('昨收盘价日期', '')
            
            # ETF基金最新净价标题显示日期（净价公布日期）
            if etf_jzrq:
                try:
                    date_obj = datetime.strptime(etf_jzrq, '%Y-%m-%d')
                    etf_latest_time_header = date_obj.strftime('%m-%d')
                except:
                    etf_latest_time_header = etf_jzrq
            
            # ETF基金上日净价标题显示日期（昨收盘价日期，即上一交易日）
            if etf_prev_trade_date and etf_prev_trade_date != 'N/A':
                try:
                    date_obj = datetime.strptime(etf_prev_trade_date, '%Y-%m-%d')
                    etf_prev_date_header = date_obj.strftime('%m-%d')
                except:
                    etf_prev_date_header = etf_prev_trade_date
            elif etf_jzrq:
                try:
                    date_obj = datetime.strptime(etf_jzrq, '%Y-%m-%d')
                    etf_prev_date_header = date_obj.strftime('%m-%d')
                except:
                    etf_prev_date_header = etf_jzrq
    
    # 生成HTML文件
    html_content = html_template.format(
        timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        total_funds=len(monitor_data),
        chaochao_count=len(chaochao_data),
        yaoyao_count=len(yaoyao_data),
        overseas_count=len(overseas_fund_data),
        monitor_count=len(monitor_data),
        avg_change=avg_change,
        up_count=up_count,
        down_count=down_count,
        flat_count=flat_count,
        chaochao_avg_change=chaochao_avg_change,
        chaochao_up_count=chaochao_up_count,
        chaochao_down_count=chaochao_down_count,
        chaochao_flat_count=chaochao_flat_count,
        yaoyao_avg_change=yaoyao_avg_change,
        yaoyao_up_count=yaoyao_up_count,
        yaoyao_down_count=yaoyao_down_count,
        yaoyao_flat_count=yaoyao_flat_count,
        chaochao_today_profit=chaochao_today_profit,
        yaoyao_today_profit=yaoyao_today_profit,
        chaochao_profit_class=chaochao_profit_class,
        yaoyao_profit_class=yaoyao_profit_class,
        chaochao_holdings_summary=chaochao_holdings_summary,
        yaoyao_holdings_summary=yaoyao_holdings_summary,
        monitor_avg_class=monitor_avg_class,
        chaochao_avg_class=chaochao_avg_class,
        yaoyao_avg_class=yaoyao_avg_class,
        latest_date_header=latest_date_header,
        estimate_date_header=estimate_date_header,
        etf_latest_time_header=etf_latest_time_header,
        etf_prev_date_header=etf_prev_date_header,
            chaochao_table_rows=generate_table_rows(chaochao_data, 'chaochao'),
            yaoyao_table_rows=generate_table_rows(yaoyao_data, 'yaoyao'),
        monitor_table_rows=generate_monitor_table_rows(monitor_data),
        overseas_table_rows=generate_overseas_table_rows(overseas_fund_data),
        etf_table_rows=generate_etf_table_rows(etf_fund_data)
    )
    
    with open(filename, 'w', encoding='utf-8-sig') as htmlfile:
        htmlfile.write(html_content)
    
    log_info(f"🌐 HTML: {filename}")
    return filename



def analyze_by_category(fund_data):
    """按板块分析基金表现"""
    if not fund_data:
        return
    
    print("\n=== 板块分析 ===")
    
    # 按板块分组
    category_groups = {}
    for fund in fund_data:
        category = fund['板块分类']
        if category not in category_groups:
            category_groups[category] = []
        category_groups[category].append(fund)
    
    # 分析每个板块并收集数据
    category_stats = []
    for category, funds in category_groups.items():
        valid_changes = []
        for fund in funds:
            try:
                change = float(fund['估算涨跌率'])
                valid_changes.append(change)
            except:
                pass
        
        if valid_changes:
            avg_change = sum(valid_changes) / len(valid_changes)
            up_count = len([x for x in valid_changes if x > 0])
            down_count = len([x for x in valid_changes if x < 0])
            
            category_stats.append({
                'category': category,
                'avg_change': avg_change,
                'up_count': up_count,
                'down_count': down_count
            })
    
    # 按平均涨跌幅降序排列（涨幅最大的在前面）
    category_stats.sort(key=lambda x: x['avg_change'], reverse=True)
    
    # 输出排序后的板块分析
    for stat in category_stats:
        print(f"🏷️  {stat['category']}: {stat['avg_change']:+.2f}% (↑{stat['up_count']} ↓{stat['down_count']})")

def _clean_remote_url(remote_url):
    """清理和验证远程URL格式"""
    if not remote_url or not isinstance(remote_url, str):
        return None
    
    # 检查URL是否包含重复的令牌或错误的格式
    if remote_url.count('@') > 1 or 'x-access-token:' in remote_url:
        print("⚠️ 检测到损坏的远程URL，正在修复...")
        # 提取正确的仓库路径
        if 'github.com' in remote_url:
            # 找到github.com后的部分
            github_part = remote_url.split('github.com')
            if len(github_part) > 1:
                repo_path = 'github.com' + github_part[1]
                clean_url = f"https://{repo_path}"
                print(f"✅ 已修复远程URL: {clean_url}")
                return clean_url
    
    return remote_url

def update_github_pages(html_filename):
    """更新GitHub Pages的index.html和.nojekyll文件，先更新到fund_tool_clean文件夹，再推送"""
    try:
        import os
        import subprocess
        import shutil
        
        # 检查fund_tool_clean文件夹是否存在
        clean_folder = '../fund_tool_clean'
        if not os.path.exists(clean_folder):
            print(f"❌ 未找到 {clean_folder} 文件夹，请先创建该文件夹")
            return False
        
        print(f"📁 正在更新 {clean_folder} 文件夹...")
        
        # 复制最新的HTML文件到fund_tool_clean文件夹
        if os.path.exists(html_filename):
            clean_html_path = os.path.join(clean_folder, 'index.html')
            shutil.copy2(html_filename, clean_html_path)
            print(f"✓ 已更新 {clean_html_path}（从 {html_filename}）")
        else:
            print(f"❌ 源文件 {html_filename} 不存在")
            return False
        
        # 在fund_tool_clean文件夹中创建.nojekyll文件
        nojekyll_path = os.path.join(clean_folder, '.nojekyll')
        if not os.path.exists(nojekyll_path):
            with open(nojekyll_path, 'w') as f:
                pass
            print(f"✓ 已创建 {nojekyll_path} 文件")
        
        # 切换到fund_tool_clean文件夹进行推送
        original_dir = os.getcwd()
        os.chdir(clean_folder)
        print(f"📂 已切换到 {clean_folder} 目录")
        
        # 检查是否在git仓库中
        try:
            result = subprocess.run(['git', 'status'], capture_output=True, text=True, timeout=10)
            if result.returncode == 0:
                print("🔍 检测到Git仓库，开始自动更新...")
                
                # 配置Git用户信息（如果未配置）
                try:
                    subprocess.run(['git', 'config', '--local', 'user.name', 'SevenKale'], check=True, capture_output=True)
                    subprocess.run(['git', 'config', '--local', 'user.email', 'sevenkale@example.com'], check=True, capture_output=True)
                    print("✓ Git用户信息已配置")
                except:
                    print("ℹ️ Git用户信息配置跳过（可能已配置）")
                
                # 处理可能的文件冲突：以本地为准，强制覆盖
                print("🔄 处理文件冲突：以本地为准...")
                try:
                    # 强制添加所有本地更改，确保本地文件被包含
                    subprocess.run(['git', 'add', '.'], check=True, capture_output=True)
                    print("✓ 已添加所有本地更改到暂存区")
                except:
                    print("ℹ️ Git添加跳过（可能无更改）")
                
                # 添加文件到暂存区（已经在上一步添加了所有文件，这里可以跳过）
                print("📁 文件已添加到Git暂存区...")
                
                # 提交更改
                commit_message = f'Update fund report - {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
                print(f"💾 提交更改: {commit_message}")
                subprocess.run(['git', 'commit', '-m', commit_message], check=True)
                
                # 获取远程仓库信息
                try:
                    remote_result = subprocess.run(['git', 'remote', '-v'], capture_output=True, text=True, timeout=10)
                    if remote_result.returncode == 0:
                        print("🌐 远程仓库信息:")
                        print(remote_result.stdout.strip())
                        
                        # 检查并清理损坏的远程URL
                        remote_url_result = subprocess.run(['git', 'remote', 'get-url', 'origin'], capture_output=True, text=True, timeout=10)
                        if remote_url_result.returncode == 0:
                            current_remote_url = remote_url_result.stdout.strip()
                            cleaned_url = _clean_remote_url(current_remote_url)
                            if cleaned_url and cleaned_url != current_remote_url:
                                print("🔧 正在修复损坏的远程URL...")
                                subprocess.run(['git', 'remote', 'set-url', 'origin', cleaned_url], check=True)
                                print("✅ 远程URL已修复")
                    else:
                        print("⚠️ 无法获取远程仓库信息")
                except:
                    pass
                
                # 推送到GitHub（使用SSH或HTTPS，无令牌）
                print("🚀 开始推送到GitHub...")
                
                # 方法1：尝试直接推送
                try:
                    subprocess.run(['git', 'push'], check=True, timeout=60)
                    print("✅ 推送成功！")
                    print("🌐 GitHub Pages访问地址: https://SevenKale.github.io/fund-report/")
                    return True
                except subprocess.CalledProcessError as e1:
                    print(f"⚠️ 直接推送失败: {e1}")
                    
                    # 方法2：尝试设置上游分支并推送
                    try:
                        print("🔄 尝试设置上游分支...")
                        subprocess.run(['git', 'push', '--set-upstream', 'origin', 'main'], check=True, timeout=60)
                        print("✅ 设置上游分支并推送成功！")
                        print("🌐 GitHub Pages访问地址: https://SevenKale.github.io/fund-report/")
                        return True
                    except subprocess.CalledProcessError as e2:
                        print(f"⚠️ 设置上游分支失败: {e2}")
                        
                        # 方法3：强制推送（以本地为准）
                        try:
                            print("💪 尝试强制推送（以本地为准）...")
                            subprocess.run(['git', 'push', '--force-with-lease', 'origin', 'main'], check=True, timeout=60)
                            print("✅ 强制推送成功！")
                            print("🌐 GitHub Pages访问地址: https://SevenKale.github.io/fund-report/")
                            return True
                        except subprocess.CalledProcessError as e3:
                            print(f"⚠️ 强制推送失败: {e3}")
                        
                        # 方法4：提供手动推送指导
                        print("\n💡 自动推送失败，请手动执行以下命令:")
                        print("   1. 检查Git状态: git status")
                        print("   2. 手动推送: git push origin main")
                        print("   3. 如果有冲突，强制推送（以本地为准）:")
                        print("      git push --force-with-lease origin main")
                        print("   4. 如果仍有问题，尝试:")
                        print("      git push -u origin main")
                        print("\n⚠️  重要提醒：本地代码保持不变，推送失败不影响本地文件")
                        print("💡 本地文件已保存，可以稍后手动推送或下次运行时自动重试")
                        
                except subprocess.TimeoutExpired:
                    print("⚠️  Git推送超时，可能是网络问题")
                    print("💡 建议稍后手动推送: git push origin main")
                    print("⚠️  重要提醒：本地代码保持不变，推送失败不影响本地文件")
                except KeyboardInterrupt:
                    print("⚠️  推送被中断（KeyboardInterrupt），已跳过自动推送")
                    print("⚠️  重要提醒：本地代码保持不变，推送失败不影响本地文件")
            else:
                print("ℹ️  不在git仓库中，已更新本地文件")
                print("💡 如需自动推送，请确保在Git仓库目录中运行")
        except (subprocess.TimeoutExpired, subprocess.CalledProcessError, FileNotFoundError) as e:
            print(f"ℹ️  Git操作失败: {e}")
            print("💡 已更新本地文件，请手动推送到GitHub")
        
        # 切换回原目录
        os.chdir(original_dir)
        print(f"📂 已切换回原目录: {original_dir}")
            
    except Exception as e:
        print(f"⚠️  更新GitHub Pages时出错: {e}")
        import traceback
        traceback.print_exc()
        # 确保切换回原目录
        try:
            os.chdir(original_dir)
            print(f"📂 已切换回原目录: {original_dir}")
        except:
            pass
    
    return False

def update_fund_values():
    """手动更新基金净值（当今日净值公布后）"""
    print("🔄 更新基金净值...")
    print("=" * 50)
    
    # 获取自选基金数据
    log_info("\n🔍 获取自选基金数据...")
    self_selected_dict = get_self_selected_funds(max_workers=10)
    
    # 获取监控基金数据
    log_info("\n🔍 获取监控基金数据...")
    monitor_funds = get_monitor_funds(max_workers=10)
    
    if self_selected_dict and monitor_funds:
        # 保存Excel文件（包含多个sheet页）
        excel_filename = save_to_excel(self_selected_dict, monitor_funds)
        # 保存HTML文件（多sheet页显示）
        html_filename = save_to_html_multi_sheet(self_selected_dict, monitor_funds)
        
        print("\n=== 净值更新完成 ===")
        print(f"✅ Excel文件: {excel_filename}")
        print(f"✅ HTML文件: {html_filename}")
        
        # 持仓收益计算
        print("\n=== 持仓收益计算 ===")
        try:
            calculator = HoldingsProfitCalculator()
            holdings_data = calculator.load_holdings_from_excel()
            
            if holdings_data:
                print("💰 计算持仓收益...")
                profit_results = calculator.calculate_holdings_profit(holdings_data, self_selected_dict)
                
                # 显示汇总结果
                print("\n=== 收益汇总 ===")
                for user, result in profit_results.items():
                    user_name = "钞钞" if user == "chaochao" else "垚垚"
                    print(f"\n{user_name}的持仓:")
                    print(f"  总投入: {result['total_cost']:,.2f}")
                    print(f"  当前市值: {result['total_current_value']:,.2f}")
                    print(f"  总收益: {result['total_profit']:,.2f} ({result['total_profit_rate']:+.2f}%)")
                    print(f"  今日收益: {result['today_profit']:,.2f} ({result['today_profit_rate']:+.2f}%)")
                
                print("✅ 持仓收益计算完成")
            else:
                print("⚠️  未能加载持仓数据")
        except Exception as e:
            print(f"⚠️  持仓收益计算失败: {e}")
        
        # 自动更新GitHub Pages
        print("\n=== 自动更新GitHub Pages ===")
        update_github_pages(html_filename)
        
    else:
        print("❌ 未获取到完整的基金数据")

def calculate_holdings_profit():
    """计算持仓收益（独立模式）"""
    log_info("💰 正在计算持仓收益...")
    log_info("=" * 60)
    
    # 获取自选基金数据
    log_info("\n正在获取自选基金数据...")
    self_selected_dict = get_self_selected_funds(max_workers=10)
    
    if not self_selected_dict:
        log_info("❌ 未能获取到基金数据")
        return
    
    # 计算持仓收益
    log_info("\n正在计算持仓收益...")
    try:
        calculator = HoldingsProfitCalculator()
        holdings_data = calculator.load_holdings_from_excel()
        
        if not holdings_data:
            log_info("❌ 未能加载持仓数据")
            return None
        
        profit_results = calculator.calculate_holdings_profit(holdings_data, self_selected_dict)
        
        # 显示汇总结果
        log_info("\n=== 持仓收益汇总 ===")
        for user, result in profit_results.items():
            user_name = "钞钞" if user == "chaochao" else "垚垚"
            log_info(f"\n{user_name}的持仓:")
            log_info(f"  总投入: {result['total_cost']:,.2f}")
            log_info(f"  当前市值: {result['total_current_value']:,.2f}")
            log_info(f"  总收益: {result['total_profit']:,.2f} ({result['total_profit_rate']:+.2f}%)")
            log_info(f"  今日收益: {result['today_profit']:,.2f} ({result['today_profit_rate']:+.2f}%)")
        
        log_info(f"\n✅ 持仓收益计算完成!")
        
        return profit_results
            
    except Exception as e:
        log_info(f"❌ 持仓收益计算失败: {e}")
        import traceback
        traceback.print_exc()
        return None

def update_holdings_from_holdings_file(holdings_file='holdings_data.xlsx', auto_confirm=False, fund_data_dict=None):
    """从持仓文件中检测操作数据并更新持仓"""
    try:
        # 检查持仓文件是否存在
        log_info(f"🔍 检查文件路径: {os.path.abspath(holdings_file)}")
        if not os.path.exists(holdings_file):
            log_info(f"📄 持仓文件 {holdings_file} 不存在，跳过持仓更新")
            return True
        
        # 加载当前持仓数据
        calculator = HoldingsProfitCalculator()
        holdings_data = calculator.load_holdings_from_excel(holdings_file)
        
        if not holdings_data:
            log_info("❌ 无法加载持仓数据，请检查持仓文件格式")
            return False
        
        # 检查持仓文件中是否有操作数据字段
        # 需要检查是否有 buy_amount, sell_shares, convert_shares 等字段
        trading_data = []
        has_operations = False
        
        try:
            excel_file = pd.ExcelFile(holdings_file)
            log_info(f"📋 工作表: {excel_file.sheet_names}")
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(holdings_file, sheet_name=sheet_name)
                log_info(f"📊 工作表 {sheet_name} 的列: {df.columns.tolist()}")
                
                # 检查是否有操作字段
                operation_fields = ['buy_amount', 'sell_shares', 'convert_shares', 'convert_from_fund_code', 'convert_from_fund_name', 'convert_ratio']
                has_operation_fields = any(field in df.columns for field in operation_fields)
                log_info(f"🔍 工作表 {sheet_name} 是否有操作字段: {has_operation_fields}")
                
                if not has_operation_fields:
                    continue
                
                for _, row in df.iterrows():
                    # 检查是否有任何操作（买入、卖出、转换）
                    buy_amount = float(row.get('buy_amount', 0)) if 'buy_amount' in df.columns else 0
                    sell_shares = float(row.get('sell_shares', 0)) if 'sell_shares' in df.columns else 0
                    convert_shares = float(row.get('convert_shares', 0)) if 'convert_shares' in df.columns else 0
                    
                    if buy_amount > 0 or sell_shares > 0 or convert_shares > 0:
                        has_operations = True
                        # 确保基金代码是6位数字格式
                        fund_code = str(row.get('fund_code', '')).strip()
                        if fund_code and fund_code != 'nan':
                            fund_code = fund_code.split('.')[0].zfill(6)  # 去掉小数部分并补0
                        else:
                            fund_code = '000000'
                        
                        convert_from_code = str(row.get('convert_from_fund_code', '')).strip()
                        if convert_from_code and convert_from_code != 'nan':
                            convert_from_code = convert_from_code.split('.')[0].zfill(6)
                        else:
                            convert_from_code = ''
                        
                        trade = {
                            'user': sheet_name,
                            'fund_code': fund_code,
                            'fund_name': row.get('fund_name', ''),
                            'shares': float(row.get('shares', 0)),
                            'cost_price': float(row.get('cost_price', 0)),
                            'cost_amount': float(row.get('cost_amount', 0)),
                            'buy_amount': buy_amount,
                            'sell_shares': sell_shares,
                            'convert_shares': convert_shares,
                            'convert_from_fund_code': convert_from_code,
                            'convert_from_fund_name': row.get('convert_from_fund_name', ''),
                            'convert_ratio': float(row.get('convert_ratio', 1))
                        }
                        trading_data.append(trade)
        except Exception as e:
            log_info(f"读取持仓文件失败: {e}")
            return False
        
        if not has_operations:
            log_info("📄 持仓文件中没有操作数据字段，跳过持仓更新")
            return False
        
        if not trading_data:
            log_info("📄 持仓文件中没有有效的操作数据，跳过持仓更新")
            return False
        
        # 显示检测到的交易操作
        log_info(f"🔍 检测到 {len(trading_data)} 笔交易操作:")
        for trade in trading_data:
            user_name = "钞钞" if trade['user'] == 'chaochao' else "垚垚"
            fund_name = trade.get('fund_name', '')
            if trade['buy_amount'] > 0:
                log_info(f"  - {user_name}: 买入 {trade['fund_code']}({fund_name}) {trade['buy_amount']:.2f}元")
            if trade['sell_shares'] > 0:
                log_info(f"  - {user_name}: 卖出 {trade['fund_code']}({fund_name}) {trade['sell_shares']:.2f}份")
            if trade['convert_shares'] > 0:
                convert_from_name = trade.get('convert_from_fund_name', '')
                log_info(f"  - {user_name}: 转换 {trade['convert_from_fund_code']}({convert_from_name}) -> {trade['fund_code']}({fund_name}) {trade['convert_shares']:.2f}份")
        
        # 询问用户是否更新
        if not auto_confirm:
            while True:
                response = input("\n❓ 是否更新持仓数据？(Y/N): ").strip().upper()
                if response in ['Y', 'YES', '是']:
                    break
                elif response in ['N', 'NO', '否']:
                    log_info("⏭️ 用户取消更新，跳过持仓更新")
                    return True
                else:
                    print("请输入 Y 或 N")
        
        log_info("🔄 开始更新持仓数据...")
        
        # 先更新持仓文件中的基金名称（使用最新的基金数据）
        if fund_data_dict:
            try:
                calculator = HoldingsProfitCalculator()
                name_updated = calculator.update_fund_names_in_holdings(holdings_data, fund_data_dict, verbose=False)
                if name_updated:
                    # 重新加载更新后的持仓数据
                    holdings_data = calculator.load_holdings_from_excel(holdings_file)
                    log_info("✓ 基金名称已更新到最新版本")
            except Exception as e:
                log_info(f"⚠️  更新基金名称失败: {e}")
        
        # 获取基金净值数据 - 优先使用已有的基金数据
        fund_codes = set()
        for trade in trading_data:
            fund_codes.add(trade['fund_code'])
            if trade['convert_from_fund_code']:
                fund_codes.add(trade['convert_from_fund_code'])
        
        fund_nav_data = {}
        if fund_data_dict:
            # 使用已有的基金数据，避免重复获取
            log_info("📊 使用已有基金数据获取净值...")
            for fund_code in fund_codes:
                # 在自选基金数据中查找
                found_nav = None
                for user_key in ['chaochao', 'yaoyao', 'all']:
                    if user_key in fund_data_dict:
                        for fund in fund_data_dict[user_key]:
                            if fund.get('基金代码') == fund_code:
                                try:
                                    nav_price = float(fund.get('估算净值', 0))
                                    if nav_price > 0:
                                        found_nav = nav_price
                                        break
                                except:
                                    pass
                    if found_nav:
                        break
                
                if found_nav:
                    fund_nav_data[fund_code] = found_nav
                else:
                    log_info(f"⚠️ {fund_code} 在已有数据中未找到")
        else:
            # 如果没有已有数据，则重新获取
            log_info("📊 重新获取基金净值数据...")
            tracker = OptimizedFundTracker()
            for fund_code in fund_codes:
                try:
                    fund_data = tracker.get_funds_realtime([fund_code])
                    if fund_data and fund_data[0]:
                        nav_price = float(fund_data[0].get('gsz', 0))
                        if nav_price > 0:
                            fund_nav_data[fund_code] = nav_price
                except Exception as e:
                    log_info(f"❌ 获取基金 {fund_code} 净值失败: {e}")
        
        # 显示净值获取汇总
        log_info(f"📊 净值: 今日 {len(fund_nav_data)} 只, 估算 {len(fund_nav_data)} 只")
        
        # 更新持仓数据
        success = calculator.update_holdings_from_trading_data(holdings_data, trading_data, fund_nav_data, fund_data_dict)
        
        if success:
            log_info("✅ 持仓数据更新成功")
        else:
            log_info("❌ 持仓数据更新失败")
        
        return success
        
    except Exception as e:
        log_info(f"更新持仓数据失败: {e}")
        return False

def main():
    log_info("=== 基金数据汇总工具 ===")
    log_info(f"🕐 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log_info("=" * 50)
    
    # 检查是否有命令行参数
    import sys
    if len(sys.argv) > 1:
        if sys.argv[1] == "--update":
            # 净值更新模式
            update_fund_values()
            return
        elif sys.argv[1] == "--profit":
            # 持仓收益计算模式
            calculate_holdings_profit()
            return
        elif sys.argv[1] == "--import-trades":
            # 从Excel导入交易记录模式
            log_info("📥 从Excel导入交易记录...")
            calculator = HoldingsProfitCalculator()
            excel_file = sys.argv[2] if len(sys.argv) > 2 else 'trade_data.xlsx'
            # 如果文件不存在，先创建模板
            if not os.path.exists(excel_file):
                log_info(f"📄 文件 {excel_file} 不存在，正在创建模板...")
                calculator.create_trade_data_template(excel_file)
                log_info("✅ 模板已创建，请填写交易记录后重新运行导入命令")
                return
            success = calculator.import_trades_from_excel(excel_file)
            if success:
                log_info("✅ 导入完成")
            else:
                log_info("❌ 导入失败")
            return
        elif sys.argv[1] == "--create-trade-template":
            # 创建交易记录模板
            log_info("📄 创建交易记录模板...")
            calculator = HoldingsProfitCalculator()
            excel_file = sys.argv[2] if len(sys.argv) > 2 else 'trade_data.xlsx'
            success = calculator.create_trade_data_template(excel_file)
            if success:
                log_info(f"✅ 模板已创建: {excel_file}")
            else:
                log_info("❌ 创建模板失败")
            return
        elif sys.argv[1] == "--export-trades":
            # 从JSON导出交易记录到Excel
            log_info("📤 从JSON导出交易记录到Excel...")
            calculator = HoldingsProfitCalculator()
            excel_file = sys.argv[2] if len(sys.argv) > 2 else 'trade_data.xlsx'
            success = calculator.export_trades_to_excel(excel_file)
            if success:
                log_info("✅ 导出完成")
            else:
                log_info("❌ 导出失败")
            return
        elif sys.argv[1] == "--restore-trades":
            # 从备份恢复交易记录
            log_info("🔄 从备份恢复交易记录...")
            calculator = HoldingsProfitCalculator()
            excel_file = sys.argv[2] if len(sys.argv) > 2 else 'trade_data.xlsx'
            success = calculator.restore_trades_from_backup(excel_file)
            if success:
                log_info("✅ 恢复完成")
            else:
                log_info("❌ 恢复失败")
            return
        elif sys.argv[1] == "--help":
            log_info("使用方法:")
            log_info("  python combined_fund_tracker.py                        # 标准模式：获取基金数据并生成报告")
            log_info("  python combined_fund_tracker.py --update                # 更新模式：更新基金净值")
            log_info("  python combined_fund_tracker.py --profit                # 收益模式：计算持仓收益")
            log_info("  python combined_fund_tracker.py --import-trades [file]   # 从Excel全量导入交易记录到JSON（默认trade_data.xlsx）")
            log_info("  python combined_fund_tracker.py --export-trades [file]   # 从JSON全量导出交易记录到Excel（默认trade_data.xlsx）")
            log_info("  python combined_fund_tracker.py --create-trade-template [file] # 创建交易记录模板（默认trade_data.xlsx）")
            log_info("  python combined_fund_tracker.py --restore-trades [file]  # 从备份恢复交易记录（默认trade_data.xlsx）")
            log_info("  python combined_fund_tracker.py --help                  # 显示帮助信息")
            return
    
    # 可以通过参数调整并发数，默认10个并发
    max_workers = 10
    
    # 获取自选基金数据
    log_info("\n🔍 获取自选基金数据...")
    self_selected_dict = get_self_selected_funds(max_workers=max_workers)
    
    # 获取监控基金数据
    log_info("\n🔍 获取监控基金数据...")
    monitor_funds = get_monitor_funds(max_workers=max_workers)
    
    # 初始化 profit_results（在条件块外初始化，避免作用域问题）
    profit_results = None
    
    if self_selected_dict and monitor_funds:
        # 加载持仓数据
        log_info("\n💰 加载持仓数据...")
        calculator = HoldingsProfitCalculator()
        holdings_data = None
        
        try:
            holdings_data = calculator.load_holdings_from_excel()
            if not holdings_data:
                print("❌  持仓数据加载失败，请检查持仓文件格式后重新运行程序")
                print("⚠️  跳过持仓收益计算，继续生成报告...")
            else:
                # 验证持仓数据与基金数据的匹配性
                if not calculator.validate_holdings_data(holdings_data, self_selected_dict):
                    print("❌  持仓数据与基金数据不匹配，请检查持仓数据文件！")
                    print("⚠️  跳过持仓收益计算，继续生成报告...")
                    holdings_data = None
        except Exception as e:
            log_info(f"⚠️  加载持仓数据失败: {e}")
            print("⚠️  跳过持仓收益计算，继续生成报告...")
        
        # 检测并更新持仓数据（如果有操作数据）
        if holdings_data:
            log_info("\n🔍 检查持仓文件中的操作数据...")
            updated_holdings = update_holdings_from_holdings_file('holdings_data.xlsx', fund_data_dict=self_selected_dict)
            if updated_holdings:
                # 重新加载更新后的持仓数据
                holdings_data = calculator.load_holdings_from_excel()
                log_info("✅ 持仓数据已更新，使用最新数据计算收益")
        
        # 计算持仓收益（使用最新的持仓数据）
        
        if holdings_data:
            log_info("\n💰 计算持仓收益...")
            try:
                profit_results = calculator.calculate_holdings_profit(holdings_data, self_selected_dict)
                # 将持仓收益信息添加到基金数据中（仅自选基金）
                self_selected_dict = calculator.enhance_fund_data_with_holdings(self_selected_dict, profit_results)
                log_info("✅ 持仓信息已添加到自选基金数据中")
            except Exception as e:
                log_info(f"⚠️  计算持仓收益失败: {e}")
                profit_results = None
        else:
            log_info("⚠️  跳过持仓收益计算")
        
        # 保存Excel文件（包含多个sheet页）
        excel_filename = save_to_excel(self_selected_dict, monitor_funds)
        # 保存HTML文件（多sheet页显示）
        html_filename = save_to_html_multi_sheet(self_selected_dict, monitor_funds, profit_results=profit_results)
        
        log_info("\n=== 汇总信息 ===")
        # 自选基金简洁汇总（每个用户一行）
        try:
            # profit_results 由上文收益计算阶段生成
            if 'profit_results' not in locals():
                pass
        except Exception:
            pass

        def format_user_summary(user_key, user_label):
            try:
                ur = profit_results.get(user_key, {})
                return (
                    f"{user_label}的持仓: "
                    f"投入:{ur.get('total_cost', 0):,.0f} "
                    f"市值:{ur.get('total_current_value', 0):,.0f} "
                    f"总收益:{ur.get('total_profit', 0):,.0f}({ur.get('total_profit_rate', 0):+.1f}%) "
                    f"今日:{ur.get('today_profit', 0):,.0f}({ur.get('today_profit_rate', 0):+.1f}%)"
                )
            except Exception:
                return f"{user_label}的持仓:数据不足"

        if 'profit_results' in locals() and isinstance(profit_results, dict):
            log_info(format_user_summary('chaochao', '钞钞'))
            log_info(format_user_summary('yaoyao', '垚垚'))
        
        # 计算监控基金统计
        monitor_valid_changes = []
        for fund in monitor_funds:
            try:
                change = float(fund["估算涨跌率"])
                monitor_valid_changes.append(change)
            except:
                pass
        
        if monitor_valid_changes:
            monitor_avg = sum(monitor_valid_changes) / len(monitor_valid_changes)
            monitor_up = len([x for x in monitor_valid_changes if x > 0])
            monitor_down = len([x for x in monitor_valid_changes if x < 0])
            monitor_flat = len([x for x in monitor_valid_changes if x == 0])
            
            log_info(f"📈 监控基金: {monitor_avg:+.2f}% (↑{monitor_up} ↓{monitor_down} →{monitor_flat})")
        
        # 按板块分析（仅分析监控基金）
        log_info("\n=== 板块分析 ===")
        analyze_by_category(monitor_funds)
        
        # 移除重复的持仓收益计算日志块（已在前面完成计算与渲染），避免双重输出
        
        # 自动更新GitHub Pages
        print("\n=== 自动更新GitHub Pages ===")
        update_github_pages(html_filename)
        
    else:
        print("❌ 未获取到完整的基金数据")

# Flask API服务器 - 用于提供基金历史净值数据
def create_flask_app():
    """创建Flask应用"""
    try:
        from flask import Flask, jsonify, request
        from flask_cors import CORS
        
        app = Flask(__name__)
        CORS(app)  # 允许跨域请求
        
        def _load_trades_from_file(user):
            """读取交易JSON文件"""
            try:
                filepath = os.path.join('trades', f"{user}.json")
                if not os.path.exists(filepath):
                    return []
                with open(filepath, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                return data.get('trades', []) if isinstance(data, dict) else []
            except Exception as e:
                log_info(f"⚠️ 读取交易文件失败: {e}")
                return []
        
        @app.route('/api/fund/history/<fund_code>')
        def get_fund_history(fund_code):
            """获取基金历史净值API端点"""
            try:
                days = request.args.get('days', 7, type=int)
                tracker = OptimizedFundTracker()
                
                # 处理ETF代码和境外基金代码映射
                original_fund_code = fund_code
                processed_fund_code = fund_code
                
                if '.' in fund_code:
                    # 检查是否是ETF代码（以数字开头，以.SZ或.SH结尾）
                    if fund_code.split('.')[0].isdigit() and fund_code.split('.')[1] in ['SZ', 'SH']:
                        # 对于ETF代码，移除交易所后缀
                        base_code = fund_code.split('.')[0]
                        print(f"ETF代码处理: {fund_code} -> {base_code}")
                        processed_fund_code = base_code
                    else:
                        # 对于境外基金代码，使用映射
                        mapped_code = tracker._map_overseas_fund_code(fund_code)
                        if mapped_code:
                            print(f"境外基金代码映射: {fund_code} -> {mapped_code}")
                            processed_fund_code = mapped_code
                
                print(f"最终使用的基金代码: {processed_fund_code}")
                
                data = tracker.get_fund_history_nav(processed_fund_code, days)
                
                if data:
                    return jsonify({
                        'success': True,
                        'fund_code': original_fund_code,
                        'data': data,
                        'count': len(data),
                        'message': f'成功获取{original_fund_code}的历史净值数据'
                    })
                else:
                    return jsonify({
                        'success': False,
                        'fund_code': original_fund_code,
                        'data': [],
                        'count': 0,
                        'message': f'未找到{original_fund_code}的历史净值数据'
                    }), 404
                    
            except Exception as e:
                return jsonify({
                    'success': False,
                    'fund_code': original_fund_code,
                    'error': str(e),
                    'message': f'获取{original_fund_code}历史净值数据时发生错误'
                }), 500
        
        @app.route('/api/fund/realtime/<fund_code>')
        def get_fund_realtime(fund_code):
            """获取基金实时数据API端点"""
            try:
                tracker = OptimizedFundTracker()
                data = tracker._fetch_single_fund(fund_code)
                
                if data:
                    return jsonify({
                        'success': True,
                        'fund_code': fund_code,
                        'data': data,
                        'message': f'成功获取{fund_code}的实时数据'
                    })
                else:
                    return jsonify({
                        'success': False,
                        'fund_code': fund_code,
                        'data': None,
                        'message': f'未找到{fund_code}的实时数据'
                    }), 404
                    
            except Exception as e:
                return jsonify({
                    'success': False,
                    'fund_code': fund_code,
                    'error': str(e),
                    'message': f'获取{fund_code}实时数据时发生错误'
                }), 500
        
        @app.route('/api/fund/batch', methods=['POST'])
        def get_batch_fund_data():
            """批量获取基金数据API端点"""
            try:
                data = request.get_json()
                fund_codes = data.get('fund_codes', [])
                
                if not fund_codes:
                    return jsonify({
                        'success': False,
                        'error': '请提供基金代码列表',
                        'message': 'fund_codes参数不能为空'
                    }), 400
                
                tracker = OptimizedFundTracker()
                fund_data = tracker.get_funds_realtime(fund_codes)
                
                return jsonify({
                    'success': True,
                    'fund_codes': fund_codes,
                    'data': fund_data,
                    'count': len(fund_data),
                    'message': f'成功获取{len(fund_data)}只基金的数据'
                })
                
            except Exception as e:
                return jsonify({
                    'success': False,
                    'error': str(e),
                    'message': '批量获取基金数据时发生错误'
                }), 500
        
        @app.route('/api/fund/trades/<user>/<fund_code>')
        def get_trades_by_fund(user, fund_code):
            """按用户&基金获取交易记录"""
            try:
                trades = _load_trades_from_file(user)
                fund_trades = [t for t in trades if t.get('fund_code') == fund_code]
                return jsonify({
                    'success': True,
                    'user': user,
                    'fund_code': fund_code,
                    'data': fund_trades,
                    'count': len(fund_trades)
                })
            except Exception as e:
                return jsonify({
                    'success': False,
                    'error': str(e),
                    'message': f'获取 {user}/{fund_code} 交易记录失败'
                }), 500
        
        @app.route('/api/fund/trades/<user>')
        def get_trades_by_user(user):
            """按用户获取全部交易记录"""
            try:
                trades = _load_trades_from_file(user)
                return jsonify({
                    'success': True,
                    'user': user,
                    'data': trades,
                    'count': len(trades)
                })
            except Exception as e:
                return jsonify({
                    'success': False,
                    'error': str(e),
                    'message': f'获取 {user} 交易记录失败'
                }), 500
        
        @app.route('/api/health')
        def health_check():
            """健康检查端点"""
            return jsonify({
                'success': True,
                'status': 'healthy',
                'timestamp': datetime.now().isoformat(),
                'message': '基金数据API服务运行正常'
            })
        
        return app
        
    except ImportError as e:
        print(f"⚠️  无法创建Flask应用: {e}")
        print("💡 请安装Flask: pip install flask flask-cors")
        return None

def run_api_server(host='127.0.0.1', port=5000, debug=True, ssl_cert=None, ssl_key=None, use_https=False):
    """运行API服务器，支持HTTP和HTTPS
    
    Args:
        host: 服务器地址
        port: 服务器端口
        debug: 是否开启调试模式
        ssl_cert: SSL证书文件路径（可选）
        ssl_key: SSL私钥文件路径（可选）
        use_https: 是否使用HTTPS（如果为True，需要提供证书和密钥）
    """
    app = create_flask_app()
    if app:
        print(f"🚀 启动基金数据API服务器...")
        
        # 检查是否使用HTTPS
        ssl_context = None
        if use_https:
            if ssl_cert and ssl_key:
                import os
                if os.path.exists(ssl_cert) and os.path.exists(ssl_key):
                    ssl_context = (ssl_cert, ssl_key)
                    print(f"🔒 使用HTTPS模式（证书: {ssl_cert}）")
                else:
                    print(f"⚠️  证书文件不存在，回退到HTTP模式")
                    use_https = False
            else:
                print(f"⚠️  未提供SSL证书，回退到HTTP模式")
                use_https = False
        
        protocol = 'https' if use_https else 'http'
        print(f"🌐 服务地址: {protocol}://{host}:{port}")
        print(f"📊 API端点:")
        print(f"   - GET  /api/fund/history/<fund_code> - 获取基金历史净值")
        print(f"   - GET  /api/fund/realtime/<fund_code> - 获取基金实时数据")
        print(f"   - POST /api/fund/batch - 批量获取基金数据")
        print(f"   - GET  /api/fund/trades/<user>/<fund_code> - 获取交易记录")
        print(f"   - GET  /api/health - 健康检查")
        if not use_https:
            print(f"💡 提示: 如需使用HTTPS，请提供SSL证书和密钥文件")
        print(f"💡 按 Ctrl+C 停止服务器")
        
        try:
            app.run(host=host, port=port, debug=debug, ssl_context=ssl_context)
        except KeyboardInterrupt:
            print("\n🛑 服务器已停止")
        except Exception as e:
            print(f"\n❌ 服务器启动失败: {e}")
            if use_https:
                print("💡 提示: 如果HTTPS启动失败，请检查证书文件路径和格式")
    else:
        print("❌ 无法启动API服务器")

# 如果直接运行此文件且安装了Flask，则启动API服务器
if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        if sys.argv[1] == "--api":
            # 启动API服务器模式
            # 支持命令行参数：--api [--https] [--cert cert.pem] [--key key.pem] [--host 127.0.0.1] [--port 5000]
            use_https = '--https' in sys.argv
            ssl_cert = None
            ssl_key = None
            host = '127.0.0.1'
            port = 5000
            
            # 解析命令行参数
            if '--cert' in sys.argv:
                idx = sys.argv.index('--cert')
                if idx + 1 < len(sys.argv):
                    ssl_cert = sys.argv[idx + 1]
            if '--key' in sys.argv:
                idx = sys.argv.index('--key')
                if idx + 1 < len(sys.argv):
                    ssl_key = sys.argv[idx + 1]
            if '--host' in sys.argv:
                idx = sys.argv.index('--host')
                if idx + 1 < len(sys.argv):
                    host = sys.argv[idx + 1]
            if '--port' in sys.argv:
                idx = sys.argv.index('--port')
                if idx + 1 < len(sys.argv):
                    port = int(sys.argv[idx + 1])
            
            run_api_server(host=host, port=port, use_https=use_https, ssl_cert=ssl_cert, ssl_key=ssl_key)
        elif sys.argv[1] == "--update":
            # 净值更新模式
            update_fund_values()
        elif sys.argv[1] == "--profit":
            # 持仓收益计算模式
            calculate_holdings_profit()
        elif sys.argv[1] == "--update-holdings":
            # 持仓更新模式
            update_holdings_from_holdings_file('holdings_data.xlsx')
        elif sys.argv[1] == "--help":
            log_info("使用方法:")
            log_info("  python combined_fund_tracker.py              # 标准模式：获取基金数据并生成报告")
            log_info("  python combined_fund_tracker.py --api       # API模式：启动基金数据API服务器")
            log_info("  python combined_fund_tracker.py --update    # 更新模式：更新基金净值")
            log_info("  python combined_fund_tracker.py --profit    # 收益模式：计算持仓收益")
            log_info("  python combined_fund_tracker.py --update-holdings # 持仓更新模式：根据交易数据更新持仓")
            log_info("  python combined_fund_tracker.py --help      # 显示帮助信息")
        else:
            main()
    else:
        main()
